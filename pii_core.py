
import os
import re
import logging
from datetime import datetime
from typing import Tuple, List, Dict, Any
from pathlib import Path
import pandas as pd
from pii_database import PIIRulesManager
from utils.gemini_client import GeminiPIIDetector
from utils.pii_excel_processor import PIIExcelProcessor

logger = logging.getLogger(__name__)


class RuleBasedPIIDetector:    
    def __init__(self, conn):
        if conn is None:
            raise ValueError("Database connection is required")
        
        self.conn = conn
        self.custom_rules = []
        self.initialize_patterns_from_db()
        
    def initialize_patterns_from_db(self):
        """Load all PII detection patterns from PostgreSQL database"""
        logger.info("Loading PII detection rules from database")
        
        try:
            with PIIRulesManager(self.conn) as manager:
                rules = manager.load_all_rules()
            
            self.pii_field_patterns = rules['pii_field_patterns']
            self.non_pii_patterns = rules['non_pii_patterns']
            self.sensitive_keywords = rules['sensitive_keywords']
            self.value_patterns = rules['value_patterns']
            self.datatype_hints = rules['datatype_hints']
            
            total_patterns = sum(len(p) for p in self.pii_field_patterns.values())
            logger.info(f"Loaded {len(self.pii_field_patterns)} PII categories, "
                       f"{total_patterns} patterns")
            
        except Exception as e:
            logger.error(f"Failed to load PII rules from database: {e}")
            self.pii_field_patterns = {}
            self.non_pii_patterns = []
            self.sensitive_keywords = []
            self.value_patterns = {}
            self.datatype_hints = {}
            raise
    
    def reload_rules_from_db(self):
        logger.info("Reloading PII rules from database")
        self.initialize_patterns_from_db()
    
    def set_custom_rules(self, custom_rules_list):
        self.custom_rules = custom_rules_list or []
        
        if self.custom_rules:
            logger.info(f"Custom rules loaded: {len(self.custom_rules)} rule(s)")
            for rule in self.custom_rules:
                action = rule.get('action', 'PII')
                scope = rule.get('scope', 'fieldname')
                pattern_count = len(rule.get('patterns', []))
                pii_type = rule.get('pii_type', 'Custom_PII')
                logger.info(f"  Rule: {pattern_count} patterns, Scope={scope}, "
                          f"Action={action}, Type={pii_type}")
    
    def _normalize_field_name(self, field_name):
        """Normalize field name to handle camelCase, snake_case, etc."""
        if not field_name:
            return ""
        
        normalized = re.sub(r'([a-z0-9])([A-Z])', r'\1 \2', field_name)
        normalized = normalized.lower()
        normalized = re.sub(r'[_\-.]', ' ', normalized)
        normalized = re.sub(r'\s+', ' ', normalized)
        
        return normalized.strip()
    
    def _check_custom_rules(self, field_info):
        """Check field against custom rules (HIGHEST PRIORITY)"""
        if not self.custom_rules:
            return None
        
        field_name = str(field_info.get('name', ''))
        description = str(field_info.get('description', ''))
        sample_value = str(field_info.get('sample_value', ''))
        table_name = str(field_info.get('table_name', ''))
        
        # Normalize ALL fields for consistent matching
        normalized_name = self._normalize_field_name(field_name)
        normalized_description = self._normalize_field_name(description)
        normalized_table = self._normalize_field_name(table_name)
        normalized_sample = sample_value.lower()
        
        for rule in self.custom_rules:
            patterns = rule.get('patterns', [])
            scope = rule.get('scope', 'fieldname')
            action = rule.get('action', 'PII')
            pii_type = rule.get('pii_type', 'Custom_PII')
            
            for pattern in patterns:
                # Normalize pattern based on scope
                if scope in ['fieldname', 'description', 'tablename', 'both']:
                    pattern_normalized = self._normalize_field_name(pattern)
                else:
                    pattern_normalized = pattern.lower()
                
                matched = False
                
                if scope == 'fieldname':
                    matched = pattern_normalized in normalized_name
                elif scope == 'description':
                    matched = pattern_normalized in normalized_description
                elif scope == 'samplevalue':
                    matched = pattern_normalized in normalized_sample
                elif scope == 'tablename':
                    matched = pattern_normalized in normalized_table
                elif scope == 'both':
                    matched = (pattern_normalized in normalized_name or pattern_normalized in normalized_description)
                
                if matched:
                    logger.debug(f"Custom rule matched: '{pattern}' in {scope} for field '{field_name}'")
                    
                    if action == 'PII':
                        return {
                            'is_pii': True,
                            'pii_types': [pii_type],
                            'confidence': 'HIGH',
                            'detection_method': 'CUSTOM_RULE',
                            'reason': f'Custom rule: pattern "{pattern}" in {scope} (PII)'
                        }
                    else:
                        return {
                            'is_pii': False,
                            'pii_types': [],
                            'confidence': 'HIGH',
                            'detection_method': 'CUSTOM_RULE',
                            'reason': f'Custom rule: pattern "{pattern}" in {scope} (Non-PII)'
                        }
        
        return None
    
    def detect_pii(self, field_info):
        """Detect PII with priority: Custom rules -> System PII -> System Non-PII -> Ambiguous"""
        # STEP 1: Check custom rules
        custom_result = self._check_custom_rules(field_info)
        if custom_result is not None:
            return custom_result
        
        # STEP 2: Apply system rules
        original_field_name = str(field_info.get('name', ''))
        field_name = self._normalize_field_name(original_field_name)
        description = str(field_info.get('description', '')).lower()
        datatype = str(field_info.get('datatype', '')).lower()
        sample_value = str(field_info.get('sample_value', ''))
        
        logger.debug(f"Checking field: '{original_field_name}'")
        
        pii_types_from_name = self._check_field_name(field_name)
        pii_types_from_value = self._check_sample_value(sample_value)
        has_sensitive_keywords = self._check_description(description)
        pii_types_from_datatype = self._check_datatype(datatype, field_name)
        
        all_pii_types = list(set(
            pii_types_from_name + 
            pii_types_from_value + 
            pii_types_from_datatype
        ))
        
        # System PII detected
        if all_pii_types:
            reasons = []
            confidence = 'HIGH'
            
            if pii_types_from_name:
                reasons.append(f"Field name: {', '.join(pii_types_from_name)}")
            if pii_types_from_value:
                reasons.append(f"Sample value: {', '.join(pii_types_from_value)}")
            if pii_types_from_datatype:
                reasons.append(f"Datatype: {', '.join(pii_types_from_datatype)}")
            
            return {
                'is_pii': True,
                'pii_types': all_pii_types,
                'confidence': confidence,
                'detection_method': 'RULE_BASED',
                'reason': '; '.join(reasons)
            }
        
        # Definitely NOT PII
        if self._is_definitely_not_pii(field_name, original_field_name.lower()):
            return {
                'is_pii': False,
                'pii_types': [],
                'confidence': 'HIGH',
                'detection_method': 'RULE_BASED',
                'reason': 'Field name matches non-PII pattern'
            }
        
        # Sensitive keywords but no clear PII pattern
        if has_sensitive_keywords:
            return {
                'is_pii': None,
                'pii_types': [],
                'confidence': 'UNKNOWN',
                'detection_method': 'NEEDS_AI',
                'reason': 'Contains sensitive keywords but no clear PII pattern'
            }
        
        # No clear indicators - ambiguous
        return {
            'is_pii': None,
            'pii_types': [],
            'confidence': 'UNKNOWN',
            'detection_method': 'NEEDS_AI',
            'reason': 'No clear PII patterns detected, requires contextual analysis'
        }
    
    def _is_definitely_not_pii(self, normalized_name, original_name):
        financial_terms = ['account', 'bank', 'card', 'payment', 'financial', 'routing', 'iban', 'swift']
        for term in financial_terms:
            if term in normalized_name:
                return False
        
        for pattern in self.non_pii_patterns:
            if re.search(pattern, normalized_name, re.IGNORECASE):
                return True
        
        return False
    
    def _check_field_name(self, field_name):
        """Check field name against system PII patterns"""
        detected_types = []
        for pii_type, patterns in self.pii_field_patterns.items():
            for pattern in patterns:
                if pattern.lower() in field_name.lower():
                    detected_types.append(pii_type)
                    break
        return detected_types
    
    def _check_sample_value(self, sample_value):
        """Check sample value against regex patterns"""
        if not sample_value or sample_value == 'N/A':
            return []
        
        detected_types = []
        for pii_type, pattern in self.value_patterns.items():
            if re.search(pattern, str(sample_value)):
                detected_types.append(pii_type)
        
        return detected_types
    
    def _check_description(self, description):
        """Check if description contains sensitive keywords"""
        if not description:
            return False
        
        for keyword in self.sensitive_keywords:
            if keyword in description:
                return True
        
        return False
    
    def _check_datatype(self, datatype, field_name):
        """Check if datatype suggests PII"""
        detected_types = []
        for pii_type, type_hints in self.datatype_hints.items():
            for hint in type_hints:
                if hint in datatype:
                    if self._check_field_name(field_name):
                        detected_types.append(pii_type)
                        break
        return detected_types
    
    def batch_detect(self, fields_list):
        """Batch detect PII for multiple fields"""
        results = {
            'definite_pii': [],
            'definite_non_pii': [],
            'ambiguous': []
        }
        
        for field_info in fields_list:
            detection = self.detect_pii(field_info)
            
            if detection['is_pii'] is True:
                results['definite_pii'].append({
                    'field_info': field_info,
                    'detection': detection
                })
            elif detection['is_pii'] is False:
                results['definite_non_pii'].append({
                    'field_info': field_info,
                    'detection': detection
                })
            else:
                results['ambiguous'].append({
                    'field_info': field_info,
                    'detection': detection
                })
        
        logger.info(
            f"Rule-based: {len(results['definite_pii'])} PII, "
            f"{len(results['definite_non_pii'])} non-PII, "
            f"{len(results['ambiguous'])} ambiguous"
        )
        
        return results


class PIIDetectionOrchestrator:    
    def __init__(
        self, 
        gemini_api_key: str,
        db_connection,
        batch_size: int = 12, 
        custom_rules: List = None
    ):

        if db_connection is None:
            raise ValueError("Database connection is required")
        
        self.rule_detector = RuleBasedPIIDetector(db_connection)
        
        if custom_rules:
            self.rule_detector.set_custom_rules(custom_rules)
            logger.info(f"Custom rules applied: {len(custom_rules)} rule(s)")
        
        self.ai_detector = GeminiPIIDetector(gemini_api_key)
        self.processor = PIIExcelProcessor()
        self.batch_size = batch_size
        
        logger.info(f"PII Detection Orchestrator initialized")
        logger.info(f"- Rule-based detector: Ready (using provided DB connection)")
        logger.info(f"- AI detector: Gemini, batch size: {batch_size}")
    
    def detect_header_row_for_sheet(self, file_path: str, sheet_name: str, max_rows=15) -> int:
        """
        Detect which row contains headers for a specific sheet.
        Returns the row index (0-based) where headers are found.
        """
        expected_patterns = [
            'target field name', 'targetfieldname', 'field name',
            'target field description', 'description',
            'legacy field name', 'source field name',
            'data type', 'type',
            'example', 'sample value', 'column number'
        ]
        
        try:
            for header_row in range(max_rows):
                try:
                    test_df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, nrows=1)
                    columns_lower = [str(col).strip().lower() for col in test_df.columns]
                    matches = 0
                    for pattern in expected_patterns:
                        for col in columns_lower:
                            if pattern in col.replace(' ', '').replace('_', ''):
                                matches += 1
                                break
                    
                    # Found headers
                    if matches >= 2:
                        logger.debug(f"Sheet '{sheet_name}': Headers detected at row {header_row} ({matches} matches)")
                        return header_row
                        
                except Exception:
                    continue
            
            # Default to row 0
            logger.debug(f"Sheet '{sheet_name}': No headers detected, using row 0")
            return 0
            
        except Exception as e:
            logger.debug(f"Sheet '{sheet_name}': Error detecting headers: {e}")
            return 0
    
    def read_excel_for_pii(self, file_path: str) -> Tuple[List[Tuple], bool]:
        """Read Excel file and validate sheets for PII detection"""
        try:
            filename = Path(file_path).name
            logger.info(f"Reading Excel file: '{filename}'")
            
            extension = Path(file_path).suffix.lower()
            if extension not in ('.xlsx', '.xls', '.xlsm'):
                msg = f"Unsupported file type: {extension}"
                logger.error(msg)
                return [], False
            
            # Get list of sheet names
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names
            logger.info(f"Excel file has {len(sheet_names)} sheet(s)")
            
            valid_sheets = []
            for sheet_name in sheet_names:
                # Detect header row for this sheet
                header_row = self.detect_header_row_for_sheet(file_path, sheet_name)
                
                # Read sheet with correct header row
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                
                # Validate: check if it has field name columns
                has_field_column = any(
                    col.lower().replace(' ', '').replace('_', '') in 
                    ['targetfieldname', 'legacyfieldname', 'fieldname']
                    for col in df.columns
                )
                
                if has_field_column:
                    valid_sheets.append((sheet_name, df))
                    logger.info(f"Sheet '{sheet_name}' validated (header row: {header_row}, {len(df)} data rows)")
                else:
                    logger.warning(f"Sheet '{sheet_name}' skipped - no field name column found")
            
            if not valid_sheets:
                logger.error("No valid sheets found")
                return [], False
            
            file_stem = Path(file_path).stem
            sheet_info_list = []
            
            if len(valid_sheets) == 1:
                sheet_name, df = valid_sheets[0]
                sheet_info_list.append((sheet_name, df, file_stem))
            else:
                for sheet_name, df in valid_sheets:
                    sheet_info_list.append((sheet_name, df, sheet_name))
            
            return sheet_info_list, True
            
        except Exception as e:
            logger.error(f"Error reading file: {str(e)}")
            return [], False
    
    def process_file(self, uploaded_file, progress_callback=None) -> Tuple[str, int, int, pd.DataFrame]:
        try:
            if isinstance(uploaded_file, str):
                file_path = uploaded_file
                file_display_name = os.path.basename(uploaded_file)
            else:
                file_path = uploaded_file.name
                file_display_name = file_path
            
            logger.info("="*80)
            logger.info("MULTI-SHEET PII DETECTION")
            logger.info("="*80)
            
            sheet_info_list, validation_ok = self.read_excel_for_pii(file_path)
            
            if not validation_ok:
                raise ValueError("No valid sheets found")
            
            total_sheets = len(sheet_info_list)
            logger.info(f"Processing {total_sheets} sheet(s)")
            
            all_sheets_results = []
            combined_stats = {
                'total_fields': 0,
                'total_pii': 0,
                'rule_based_pii': 0,
                'ai_based_pii': 0,
                'total_api_calls': 0
            }
            
            for sheet_idx, (sheet_name, df, identifier) in enumerate(sheet_info_list, 1):
                logger.info(f"Processing sheet {sheet_idx}/{total_sheets}: '{sheet_name}'")
                
                df['Sheet_Name'] = sheet_name
                df['Sheet_Identifier'] = identifier
                
                sheet_results, sheet_stats = self._process_single_sheet(
                    df, sheet_name, identifier, sheet_idx, total_sheets, progress_callback
                )
                
                all_sheets_results.append({
                    'sheet_name': sheet_name,
                    'identifier': identifier,
                    'df': df,
                    'results': sheet_results,
                    'stats': sheet_stats
                })
                
                for key in combined_stats:
                    if key in sheet_stats:
                        combined_stats[key] += sheet_stats[key]
                    elif key == 'total_api_calls':
                        combined_stats[key] += sheet_stats['api_calls']
            
            # Create output
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            original_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_filename = f"{original_filename}_PII_detected_{timestamp}.xlsx"
            
            output_path = self._create_multi_sheet_output(
                all_sheets_results,
                output_filename,
                combined_stats
            )
            
            # Create preview
            all_pii_rows = []
            for sheet_result in all_sheets_results:
                df = sheet_result['df']
                pii_df = df[df['PII_Flag'] == True].copy()
                all_pii_rows.append(pii_df)
            
            preview_df = pd.concat(all_pii_rows, ignore_index=True).head(10) if all_pii_rows else pd.DataFrame()
            
            logger.info("="*80)
            logger.info(f"COMPLETE: {combined_stats['total_pii']}/{combined_stats['total_fields']} PII fields found")
            logger.info("="*80)
            
            return output_path, combined_stats['total_pii'], combined_stats['total_fields'], preview_df
            
        except Exception as e:
            logger.error(f"PII detection failed: {str(e)}")
            raise
    
    def _process_single_sheet(self, df, sheet_name, identifier, sheet_idx, total_sheets, progress_callback):
        total_rows = len(df)
        
        # Phase 1: Rule-based
        logger.info("Phase 1: Rule-based classification")
        
        fields_list = []
        for idx, row in df.iterrows():
            field_info = {
                'row_index': idx,
                'name': row.get('Field Name') or row.get('targetfieldname') or row.get('Target Field Name') or 
                        row.get('legacyfieldname') or row.get('Legacy Field Name', ''),
                'description': row.get('Field Description') or row.get('targetfielddescription') or row.get('Target Field Description', ''),
                'datatype': row.get('datatype') or row.get('Data Type', ''),
                'sample_value': row.get('exampletargetdatavalue') or row.get('Example Target Data Value', ''),
                'table_name': row.get('Table Name') or row.get('targettablename') or row.get('Target Table Name', '') or row.get('Legacy Table Name', '')
            }
            fields_list.append(field_info)
        
        rule_results = self.rule_detector.batch_detect(fields_list)
        ambiguous_count = len(rule_results['ambiguous'])
        
        # Phase 2: AI detection
        ai_results = []
        if ambiguous_count > 0:
            logger.info(f"Phase 2: Gemini API detection ({ambiguous_count} fields)")
            
            ambiguous_data = []
            for item in rule_results['ambiguous']:
                row_idx = item['field_info']['row_index']
                ambiguous_data.append({
                    'row_index': row_idx,
                    'field_data': df.iloc[row_idx].to_dict()
                })
            
            num_batches = (len(ambiguous_data) + self.batch_size - 1) // self.batch_size
            
            for batch_num in range(num_batches):
                start_idx = batch_num * self.batch_size
                end_idx = min(start_idx + self.batch_size, len(ambiguous_data))
                
                batch_data = [item['field_data'] for item in ambiguous_data[start_idx:end_idx]]
                batch_results = self.ai_detector.detect_pii_batch(batch_data)
                
                for i, result in enumerate(batch_results):
                    result['row_index'] = ambiguous_data[start_idx + i]['row_index']
                
                ai_results.extend(batch_results)
        
        # Phase 3: Merge
        all_results = self._merge_results(rule_results, ai_results)
        
        # Add to dataframe
        df['PII_Flag'] = False
        df['PII_Types'] = ''
        df['Confidence_Level'] = ''
        df['Detection_Method'] = ''
        
        for result in all_results:
            idx = result['row_index']
            if idx < len(df):
                df.at[idx, 'PII_Flag'] = result['is_pii']
                df.at[idx, 'PII_Types'] = ', '.join(result['pii_types'])
                df.at[idx, 'Confidence_Level'] = result['confidence']
                
                method = result.get('detection_method', 'Unknown')
                if method == 'RULE_BASED':
                    df.at[idx, 'Detection_Method'] = 'Rule-Based'
                elif method == 'CUSTOM_RULE':
                    df.at[idx, 'Detection_Method'] = 'Custom Rule'
                elif method == 'AI_BASED':
                    df.at[idx, 'Detection_Method'] = 'Gemini API'
        
        # Stats
        total_pii = sum(1 for r in all_results if r['is_pii'])
        rule_based_pii = sum(1 for r in all_results if r['is_pii'] and 
                           r.get('detection_method') in ['RULE_BASED', 'CUSTOM_RULE'])
        ai_based_pii = sum(1 for r in all_results if r['is_pii'] and 
                         r.get('detection_method') == 'AI_BASED')
        
        return all_results, {
            'total_fields': total_rows,
            'total_pii': total_pii,
            'rule_based_pii': rule_based_pii,
            'ai_based_pii': ai_based_pii,
            'api_calls': ambiguous_count
        }
    
    def _merge_results(self, rule_results, ai_results):
        """Merge rule-based and AI results"""
        ai_map = {r['row_index']: r for r in ai_results}
        all_results = []
        
        for item in rule_results['definite_pii']:
            all_results.append({
                'row_index': item['field_info']['row_index'],
                'is_pii': True,
                'pii_types': item['detection']['pii_types'],
                'confidence': item['detection']['confidence'],
                'reason': item['detection']['reason'],
                'detection_method': item['detection']['detection_method']
            })
        
        for item in rule_results['definite_non_pii']:
            all_results.append({
                'row_index': item['field_info']['row_index'],
                'is_pii': False,
                'pii_types': [],
                'confidence': item['detection']['confidence'],
                'reason': item['detection']['reason'],
                'detection_method': item['detection']['detection_method']
            })
        
        for item in rule_results['ambiguous']:
            row_idx = item['field_info']['row_index']
            if row_idx in ai_map:
                ai_result = ai_map[row_idx]
                all_results.append({
                    'row_index': row_idx,
                    'is_pii': ai_result['is_pii'],
                    'pii_types': ai_result['pii_types'],
                    'confidence': ai_result['confidence'],
                    'reason': ai_result.get('reason', 'AI analysis'),
                    'detection_method': 'AI_BASED'
                })
        
        all_results.sort(key=lambda x: x['row_index'])
        return all_results
    
    def _create_multi_sheet_output(self, all_sheets_results, output_filename, combined_stats):
        """Create multi-sheet Excel output"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        summary_ws = wb.create_sheet("PII Detection Summary", 0)
        summary_ws['A1'] = "PII Detection Summary"
        summary_ws['A1'].font = Font(bold=True, size=16)
        summary_ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        row = 4
        stats = [
            ("Total Sheets:", len(all_sheets_results)),
            ("Total Fields:", combined_stats['total_fields']),
            ("Total PII Fields:", combined_stats['total_pii']),
            ("", ""),
            ("Rule-Based PII:", combined_stats['rule_based_pii']),
            ("Gemini API PII:", combined_stats['ai_based_pii']),
            ("", ""),
            ("API Calls:", combined_stats['total_api_calls']),
            ("Cost Reduction:", f"{(1-combined_stats['total_api_calls']/combined_stats['total_fields'])*100:.1f}%")
        ]
        
        for label, value in stats:
            summary_ws[f'A{row}'] = label
            summary_ws[f'B{row}'] = value
            if label:
                summary_ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Per-sheet statistics
        row += 2
        summary_ws[f'A{row}'] = "Per-Sheet Results"
        summary_ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        # Headers
        headers = ["Sheet Name", "Total Fields", "PII Fields", "Detection Rate", "Rule-Based PII", "Gemini API PII"]
        for col_idx, header in enumerate(headers, 1):
            cell = summary_ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="FF4472C4")
            cell.font = Font(bold=True, color="FFFFFFFF")
        
        row += 1
        
        # Data
        for sheet_result in all_sheets_results:
            stats = sheet_result['stats']
            detection_rate = (stats['total_pii'] / stats['total_fields'] * 100) if stats['total_fields'] > 0 else 0
            
            data = [
                sheet_result['sheet_name'],
                stats['total_fields'],
                stats['total_pii'],
                f"{detection_rate:.1f}%",
                stats['rule_based_pii'],
                stats['ai_based_pii']
            ]
            
            for col_idx, value in enumerate(data, 1):
                summary_ws.cell(row=row, column=col_idx, value=value)
            
            row += 1
        
        # Auto-adjust column widths
        for col_idx in range(1, 7):
            summary_ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data sheets
        for sheet_result in all_sheets_results:
            sheet_name = sheet_result['sheet_name'][:31]
            df = sheet_result['df']
            
            ws = wb.create_sheet(sheet_name)
            
            # Headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", fgColor="FF4472C4")
                cell.font = Font(bold=True, color="FFFFFFFF")
            
            # Data
            for row_idx, row_data in enumerate(df.itertuples(index=False, name=None), 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            ws.freeze_panes = "A2"
        
        wb.save(output_filename)
        return output_filename