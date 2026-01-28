import os
import logging
import pandas as pd
from tkinter import ttk, filedialog, messagebox
import tkinter as tk
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import RealDictCursor
import questionary
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import configparser
from psycopg2.extras import RealDictCursor, execute_values
from psycopg2 import sql
from psycopg2.extras import execute_values
from questionary import Choice
import base64
from io import StringIO
import configparser
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from cryptography.fernet import Fernet
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sys
from logging.handlers import RotatingFileHandler
from tkinter import font as tkFont
from PIL import Image
from pathlib import Path
from datetime import datetime
import hashlib
import time
import math
import numpy as np
import tabulate
from pii_core import PIIDetectionOrchestrator
from dotenv import load_dotenv

load_dotenv()


def init_logging(app_name="MapMaestro", log_level=logging.INFO):
    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(__file__)
    log_file = os.path.join(base_dir, f"{app_name}.log")

    # Create a rotating handler: keep up to 3 x 5MB files
    handler = RotatingFileHandler(
        log_file, maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8"
    )
    fmt = "%(asctime)s %(levelname)-8s %(name)s: %(message)s"
    handler.setFormatter(logging.Formatter(fmt))

    # Configure root logger
    root = logging.getLogger()
    root.setLevel(log_level)
    root.addHandler(handler)

    # Optional: also echo INFO+ to console if you ever run without --windowed
    console = logging.StreamHandler()
    console.setFormatter(logging.Formatter(fmt))
    root.addHandler(console)

    root.info("=" * 80)
    root.info(f"Logging initialized — writing to {log_file}")

init_logging()


def derive_key(master: bytes, salt: bytes) -> bytes:
    """Derive a 32-byte Fernet key from the master password and salt."""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=390_000,
    )
    return base64.urlsafe_b64encode(kdf.derive(master))


def decrypt_blob(master: bytes, blob: str) -> str:
    """
    Given the base64(salt+token) blob (no ENC()), decode, derive the key,
    decrypt and return the plaintext string.
    """
    data = base64.urlsafe_b64decode(blob)
    salt, token = data[:16], data[16:]
    key = derive_key(master, salt)
    return Fernet(key).decrypt(token).decode()


# ——————— Main startup ———————

# 1) Grab the master from environment

master_key = os.getenv("MASTER_KEY")
if not master_key:
    raise RuntimeError("MASTER_KEY not set in environment")
master = master_key.encode()
# 2) Read the encrypted blobs from properties file
CONFIG_FILE2 = "app.properties"
config2 = configparser.ConfigParser()
if not os.path.exists(CONFIG_FILE2):
    raise FileNotFoundError(f"Configuration file '{CONFIG_FILE2}' not found.")
config2.read(CONFIG_FILE2)
if "postgres" not in config2:
    raise KeyError("Section 'postgres' not found in configuration file.")
pg = config2["postgres"]

raw = open("app.properties", "r").read()
cfg = configparser.ConfigParser()
cfg.read_file(StringIO(raw))

# 3) Decrypt each field
postgres_url = decrypt_blob(master, cfg["postgres"]["host"][4:-1])
postgres_password = decrypt_blob(master, cfg["postgres"]["password"][4:-1])
postgres_outputurl = decrypt_blob(master, cfg["data_lineage"]["host"][4:-1])
postgres_outputpassword = decrypt_blob(master, cfg["data_lineage"]["password"][4:-1])
postgres_outputtable = cfg["data_lineage"]["tablename"]

# 4) Now you can use them, e.g. construct your connection string:
# Load database configuration from properties file
CONFIG_FILE = "app.properties"
config = configparser.ConfigParser()
if not os.path.exists(CONFIG_FILE):
    raise FileNotFoundError(f"Configuration file '{CONFIG_FILE}' not found.")
config.read(CONFIG_FILE)
if "postgres" not in config:
    raise KeyError("Section 'postgres' not found in configuration file.")

def get_db_config():
    pg = config["postgres"]
    return {
        "host": postgres_url,
        "database": pg.get("database", ""),
        "user": pg.get("user", ""),
        "password": postgres_password,
        "port": pg.getint("port", 5432),
    }

def get_outputdb_config():
    pg = config["data_lineage"]
    return {
        "host": postgres_outputurl,
        "database": pg.get("database", ""),
        "user": pg.get("user", ""),
        "password": postgres_outputpassword,
        "port": pg.getint("port", 5432),
    }


# Columns to filter (excluding industry)
COLUMNS = [
    "legacysystem",
    "projectname",
    "domainname",
    "subdomainname",
    "targetsystem",
    "filename",
    "systemintegrator",
    "golivedate",
]
FILTER_COLUMNS = [
    "legacysystem",
    "targetsystem",
    "projectname",
    "domainname",
    "subdomainname",
    "filename",
]
# Columns to fetch from table_a (including projectid)
TABLE_A_COLUMNS = ["projectid"] + FILTER_COLUMNS

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def get_connection():
    """
    Connect to PostgreSQL and return connection. Raises on failure.
    """
    db_cfg = get_db_config()
    try:
        conn = psycopg2.connect(
            host=db_cfg["host"],
            database=db_cfg["database"],
            user=db_cfg["user"],
            password=db_cfg["password"],
            port=db_cfg["port"],
        )
        logger.info("Connected to database.")
        return conn
    except psycopg2.Error as e:
        logger.error(f"Database connection failed: {e.pgerror or e}")
        raise


def get_outputconnection():
    """
    Connect to PostgreSQL and return connection. Raises on failure.
    """
    db_cfg = get_outputdb_config()
    try:
        conn = psycopg2.connect(
            host=db_cfg["host"],
            database=db_cfg["database"],
            user=db_cfg["user"],
            password=db_cfg["password"],
            port=db_cfg["port"],
        )
        logger.info("Connected to database.")
        return conn
    except psycopg2.Error as e:
        logger.error(f"Database connection failed: {e.pgerror or e}")
        raise


def fetch_table_a(conn) -> list:
    """
    Fetch rows and relevant columns from parenttable, including projectid.
    Returns list of dicts.
    """
    cols = ["projectid"] + COLUMNS
    sql = f"SELECT {', '.join(cols)} FROM map_maestro.parentprojectdetails;"
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql)
            rows = cur.fetchall()
            logger.info(f"Fetched {len(rows)} rows from parenttable.")
            return rows
    except psycopg2.Error as e:
        logger.error(f"Error fetching parenttable: {e.pgerror or e}")
        raise


def prompt_cascading_filters(rows: list) -> dict:
    """
    Prompt the user with cascading multiselects. Each selection filters the next.
    'All' is available but not pre-selected.
    Returns a dict mapping each column to selected values (empty list means no filter).
    """
    filters = {}
    current_rows = rows
    for col in FILTER_COLUMNS:
        values = sorted({r[col] for r in current_rows if r.get(col) is not None})
        choices = ["All"] + [str(v) for v in values]
        answer = (
            questionary.checkbox(
                f"Select one or more {col} values (space to select, enter to confirm; 'All' to skip):",
                choices=choices,
            ).ask()
            or []
        )
        if "All" in answer or not answer:
            filters[col] = []
        else:
            filters[col] = answer
            current_rows = [r for r in current_rows if str(r.get(col)) in answer]
    print("\nYou selected the following filters:")
    for col, vals in filters.items():
        print(f" - {col}: {vals if vals else 'All'}")
    return filters


def build_projectid_query(filters: dict) -> tuple:
    """
    Build SQL to fetch distinct projectid from parenttable based on filters.
    Returns (query_str, params).
    """
    logger.info(f"Building projectid query with filters: {filters}")
    base = "SELECT DISTINCT projectid FROM map_maestro.parentprojectdetails"
    clauses = []
    params = []
    for col, vals in filters.items():
        if vals:
            placeholders = ",".join(["%s"] * len(vals))
            clauses.append(f"{col} IN ({placeholders})")
            params.extend(vals)
    if clauses:
        base += " WHERE " + " AND ".join(clauses)
    logger.info(f"ProjectID Query: {base} Params: {params}")
    return base, params


def build_tableb_query(project_ids: list) -> tuple:
    """
    Build SQL to fetch rows from childtable filtered by projectid.
    Returns (query_str, params).
    """
    base = (
        "SELECT c.*, p.filename, p.targetsystem, p.legacysystem, "
        "p.projectname, p.domainname, p.subdomainname "
        "FROM map_maestro.childattributemappingdetails c "
        "JOIN map_maestro.parentprojectdetails p ON c.projectid = p.projectid"
    )
    params = []
    if project_ids:
        placeholders = ",".join(["%s"] * len(project_ids))
        base += f" WHERE c.projectid IN ({placeholders})"
        params = project_ids
    logger.info(f"childattributemappingdetails Query: {base} Params: {params}")
    return base, params


def execute_query(conn, query: str, params: list = None) -> list:
    """
    Execute query and return list of dict rows.
    """
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query, params)
            results = cur.fetchall()
            logger.info(f"Retrieved {len(results)} rows.")
            return results
    except psycopg2.Error as e:
        logger.error(f"Error executing query: {e.pgerror or e}")
        raise


def export_standard_excel(rows: list, filename: str):
    """
    Export list of dict rows to a standard Excel file, separate sheet per project ID.
    """
    if not rows:
        logger.warning("No data to export.")
        return
    df = pd.DataFrame(rows)
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)
    for project_id, group in df.groupby("projectid"):
        ws = wb.create_sheet(title=str(project_id))
        for col_idx, col_name in enumerate(group.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name).font = Font(bold=True)
            max_length = max(group[col_name].astype(str).map(len).max(), len(col_name))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        for row_idx, row in enumerate(group.itertuples(index=False, name=None), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        ws.freeze_panes = "A2"
    wb.save(filename)
    logger.info(f"Exported standard Excel file: {filename}")
    print(f"Excel written: {filename}")


def export_summarized_excel(rows: list, filename: str):
    """
    Export rows to Excel with a summary sheet listing distinct target tables and their attributes,
    then sheets per target table with all rows.
    """
    if not rows:
        logger.warning("No data to export.")
        return
    df = pd.DataFrame(rows)
    # Columns for summary: rename 'description' to 'targettabledescription', and add 'Dependent Tabs'

    summary_cols = [
        "targettablename",
        "targettabledescription",
        "targettableinscope",
        "targettablerequired",
        "primarytable",
        "functionaltargettablename",
        "dependenttabs",
        "mappingid",
    ]
    summary_df = (
        df[summary_cols]
        .drop_duplicates(subset=["targettablename"])
        .reset_index(drop=True)
    )
    # Rename 'description' column
    summary_df = summary_df.rename(columns={"targettabledescription": "Description"})
    # Add blank 'Dependent Tabs' column

    target_names = sorted(summary_df["targettablename"].unique())
    wb = Workbook()
    # Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_df.insert(0, "No.", range(1, len(summary_df) + 1))
    summary_df.sort_values("mappingid", inplace=True)
    summary_df.drop(columns=["mappingid"], inplace=True, errors="ignore")

    header_row = 3
    data_start = header_row + 1
    fill = PatternFill(fill_type="solid", fgColor="FF16365C")  # Dark blue fill
    for col_idx, col_name in enumerate(summary_df.columns, 1):
        # Map to display names
        display_map = {
            "No.": "No.",
            "targettablename": "Tab",
            "targettabledescription": "Description",
            "targettableinscope": "In Scope?",
            "targettablerequired": "Basic Setup",
            "dependenttabs": "Dependent Tabs",
            "primarytable": "Primary Table",
            "functionaltargettablename": "Functional Table Name",
        }
        display_name = display_map.get(col_name, col_name)
        cell = summary_ws.cell(row=header_row, column=col_idx, value=display_name)
        cell.font = Font(bold=True, size=14, color="FFFFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        max_length = max(
            summary_df[col_name].astype(str).map(len).max(), len(display_name)
        )
        summary_ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    for i, row in enumerate(
        summary_df.itertuples(index=False, name=None), start=data_start
    ):
        for col_idx, value in enumerate(row, 1):
            cell = summary_ws.cell(row=i, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left")

    # Freeze so headings remain visible
    summary_ws.freeze_panes = f"A{data_start}"
    # Create sheets per target table
    for tname in target_names:
        group = df[df["targettablename"] == tname]
        ws = wb.create_sheet(title=str(tname))
        # Write mapping specs table at top
        # Row 1: Conversion Name | <tname>
        ws.cell(row=1, column=1, value="Conversion Name:").font = Font(bold=True)
        ws.cell(row=1, column=2, value=tname)
        # Row 2: Description | fetch from summary_df
        desc = summary_df.loc[
            summary_df["targettablename"] == tname, "Description"
        ].values
        desc_val = desc[0] if len(desc) > 0 else ""
        ws.cell(row=2, column=1, value="Description:").font = Font(bold=True)
        ws.cell(row=2, column=2, value=desc_val)
        # Row 3: Conversion Cycle | blank
        ws.cell(row=3, column=1, value="Conversion Cycle:").font = Font(bold=True)
        ws.cell(row=3, column=2, value="")
        # Row 4: Selection Rules | selectionlogic value
        sel = (
            group["selectionlogic"].iloc[0] if "selectionlogic" in group.columns else ""
        )
        ws.cell(row=4, column=1, value="Selection Rules:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=sel)
        # Leave a blank row at 5
        # Data start at row 6
        data_start = 7
        # Define columns for data
        data_cols = [
            "columnnumber",
            "targetfieldname",
            "targetfielddescription",
            "datatype",
            "required",
            "LOV",
            "exampletargetdatavalue",
            "legacytablename",
            "legacyfieldname",
            "legacyfielddescription",
            "translationrule",
            "crossreferences",
            "comments",
            "columnstatus",
            "additionalinformation",
        ]
        # Create subset dataframe sorted by targetfieldname
        if "columnnumber" in group.columns:
            group = group.sort_values("columnnumber").reset_index(drop=True)
        else:
            group.insert(0, "columnnumber", range(1, len(group) + 1))

        # Write header with display names and styling
        data_display_map = {
            "columnnumber": "Column Number",
            "targetfieldname": "Field Name",
            "targetfielddescription": "Field Description",
            "datatype": "Data Type",
            "required": "Required?",
            "LOV": "LOV?",
            "exampletargetdatavalue": "Example Target Data Value",
            "legacytablename": "Legacy Table Name",
            "legacyfieldname": "Legacy Field Name",
            "legacyfielddescription": "Legacy Field Description",
            "translationrule": "Translation Rules",
            "crossreferences": "Cross References",
            "comments": "Comments",
            "columnstatus": "Column Status",
            "additionalinformation": "Additional Information",
        }
        for col_idx, col_name in enumerate(data_cols, 1):
            display_name = data_display_map.get(col_name, col_name)
            cell = ws.cell(row=data_start, column=col_idx, value=display_name)
            cell.font = Font(bold=True, size=14, color="FFFFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            if col_name in group.columns:
                max_length = max(
                    group[col_name].astype(str).map(len).max(), len(display_name)
                )
            else:
                max_length = len(display_name)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        # Write rows
        for row_idx, row in enumerate(
            group[data_cols].itertuples(index=False, name=None), data_start + 1
        ):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = f"A{data_start + 1}"
    wb.save(filename)
    logger.info(f"Exported summarized Excel file: {filename}")


COLUMN_MAPPING = {
    # Primary key
    "mappingid": "id",
    # Target table/field mappings
    "targettablename": "targettable",
    "targetfieldname": "targetfield",
    "targetfielddescription": "targetfielddescription",
    # Data attributes
    "datatype": "datatype",
    "required": "required_flag",
    "LOV": "lov_flag",
    "exampletargetdatavalue": "ex_target_data_value",
    # Legacy table/field mappings
    "legacytablename": "legacytablename",
    "legacyfieldname": "legacyfieldname",
    "legacyfielddescription": "legacyfielddescription",
    "legacytabledescription": "legacytabledescription",
    # Rules and references
    "translationrule": "translation_rules",
    "crossreferences": "cross_references",
    "dataqualityrule": "dataqualityrule",
    "comments": "comments",
    # Metadata
    "columnstatus": "column_status",
    "columnnumber": "columnnumber",
    "additionalinformation": "additional_information",
    # Domain information
    "domainname": "domainname",
    "subdomainname": "subdomainname",
    # Table metadata
    "targettabledescription": "targettabledescription",
    "targettableinscope": "targettableinscope",
    "targettablerequired": "targettablerequired",
    "primarytable": "primarytable",
    "selectionlogic": "selectionlogic",
    "functionaltargettablename": "functionaltargettablename",
    "dependenttabs": "dependenttabs",
    "datasensitivitytags": "datasensitivitytags",
    # Flags
    "isCDE": "iscde",
    "isMasterData": "ismasterdata",
    "isDataMigrationInternal": "isdatamigrationinternal",
}


def store_output_to_db(
    rows: list, conn, table_name: str, column_mapping: dict, columns_to_insert: list
):
    """
    Insert rows into Postgres.
    """
    if not rows:
        logger.info("No rows to insert.")
        return

    logger.info(f"Target table: {table_name}")
    logger.info(f"Input rows: {len(rows)}")

    # 1) Filter rows with non-null mappingid
    filtered = [r for r in rows if r.get("mappingid") is not None]
    dropped = len(rows) - len(filtered)

    if dropped > 0:
        logger.warning(f"Dropped {dropped} rows with null mappingid.")
        # Log samples of dropped rows
        null_samples = [r for r in rows if r.get("mappingid") is None][:2]
        for idx, sample in enumerate(null_samples, 1):
            logger.warning(
                f"Null mappingid sample {idx}: "
                f"targettable={sample.get('targettablename')}, "
                f"targetfield={sample.get('targetfieldname')}"
            )

    if not filtered:
        logger.error("No rows left after filtering null mappingid; aborting.")
        return

    logger.info(f"Rows with valid mappingid: {len(filtered)}")

    # 2) Fetch actual columns from target table
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_schema='public' AND table_name=%s;",
                (table_name,),
            )
            actual_cols = {row[0] for row in cur.fetchall()}
            logger.info(f"Target table '{table_name}' has {len(actual_cols)} columns")
    except Exception as e:
        logger.error(f"Failed to fetch table schema: {e}")
        raise

    # 3) Map source columns to target column names
    mapped_intended = [column_mapping.get(src, src) for src in columns_to_insert]
    logger.info(f"Mapped {len(mapped_intended)} source columns to target columns")

    # 4) Determine which mapped columns actually exist
    valid_cols = [c for c in mapped_intended if c in actual_cols]
    missing_cols = [c for c in mapped_intended if c not in actual_cols]

    if missing_cols:
        logger.warning(
            f"These mapped columns do not exist in {table_name}: {missing_cols}"
        )

    if not valid_cols:
        logger.error("No valid columns to insert; aborting.")
        return

    logger.info(f"Valid columns for insert: {len(valid_cols)}")
    logger.info(f"Valid columns: {valid_cols}")

    # 5) Build rows for insertion
    prepared = []
    skipped_no_id = 0

    INTEGER_COLUMNS = {
        "id",
        "columnnumber",
        "required_flag",
        "lov_flag",
        "iscde",
        "ismasterdata",
        "isdatamigrationinternal",
        "targettableinscope",
        "targettablerequired",
        "primarytable",
    }

    for src in filtered:
        mapped = {column_mapping.get(k, k): v for k, v in src.items()}
        if mapped.get("id") is None:
            skipped_no_id += 1
            continue

        row_pruned = {}
        for col in valid_cols:
            val = mapped.get(col)

            # Convert types for PostgreSQL compatibility
            if val is not None:
                if isinstance(val, (np.integer, np.floating)):
                    val = val.item()

                # Handle float values
                if isinstance(val, float):
                    if math.isnan(val) or math.isinf(val):
                        val = None
                    elif col in INTEGER_COLUMNS and val == int(val):
                        val = int(val)

            row_pruned[col] = val

        prepared.append(row_pruned)

    if skipped_no_id > 0:
        logger.warning(f"Skipped {skipped_no_id} rows with null 'id' after mapping")

    logger.info(f"Prepared {len(prepared)} rows for upsert")

    if not prepared:
        logger.error("No rows to insert after mapping; aborting.")
        return

    # 6) Build upsert SQL
    cols_sql = sql.SQL(", ").join(map(sql.Identifier, valid_cols))
    updates_sql = sql.SQL(", ").join(
        sql.SQL("{col}=EXCLUDED.{col}").format(col=sql.Identifier(c))
        for c in valid_cols
        if c != "id"
    )

    upsert_sql = sql.SQL(
        """
        INSERT INTO {table} ({cols})
        VALUES %s
        ON CONFLICT (id) DO UPDATE SET {updates}
        """
    ).format(table=sql.Identifier(table_name), cols=cols_sql, updates=updates_sql)

    # 7) Execute bulk insert
    try:
        values = [[row[c] for c in valid_cols] for row in prepared]

        # DEBUG: Check columnnumber values specifically
        if "columnnumber" in valid_cols:
            col_idx = valid_cols.index("columnnumber")
            logger.info(f"columnnumber is at index {col_idx} in valid_cols")

            for idx, value_row in enumerate(values[:5]):  # Check first 5 rows
                col_val = value_row[col_idx]
                logger.info(
                    f"Row {idx}: columnnumber = {col_val} (type: {type(col_val).__name__})"
                )

                if col_val is not None:
                    try:
                        num = int(col_val)
                        if num > 2147483647 or num < -2147483648:
                            logger.error(
                                f" ROW {idx}: columnnumber {num} is OUT OF RANGE!"
                            )
                    except (ValueError, TypeError) as e:
                        logger.error(
                            f" ROW {idx}: columnnumber '{col_val}' cannot convert to int: {e}"
                        )
        else:
            logger.warning("columnnumber is NOT in valid_cols!")

        # Also check 'id' column
        if "id" in valid_cols:
            id_idx = valid_cols.index("id")
            id_val = values[0][id_idx] if values else None
            logger.info(
                f"First row 'id' value: {id_val} (type: {type(id_val).__name__})"
            )

        with conn.cursor() as cur:
            execute_values(cur, upsert_sql, values)
        conn.commit()
        logger.info(f" Successfully upserted {len(prepared)} rows into `{table_name}`")
    except Exception as e:
        logger.error(f"Error during bulk insert: {e}", exc_info=True)
        conn.rollback()
        raise


def read_input_file(file_path: str):
    logger = logging.getLogger(__name__)
    try:
        filename = Path(file_path).name
        logger.info(f"Reading input file '{filename}'")

        # Check file extension
        extension = Path(file_path).suffix.lower()
        if extension not in (".xlsx", ".xls", ".xlsm"):
            msg = f"Unsupported file type: {extension}. Only Excel files (.xlsx, .xls, .xlsm) are supported."
            print(msg)
            logger.error(msg)
            return [], False

        required_columns = ["targettablename", "targetfieldname"]

        sheets = pd.read_excel(file_path, sheet_name=None)
        logger.info(f"Excel file has {len(sheets)} sheet(s)")
        print(f"Read Excel file: {filename}")
        valid_sheets = []
        for sheet_name, df in sheets.items():
            logger.info(
                f"Inspecting sheet '{sheet_name}' "
                f"({len(df)} rows, {len(df.columns)} columns)"
            )

            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                logger.info(
                    f"Sheet '{sheet_name}' skipped. "
                    f"Missing: {', '.join(missing_columns)}"
                )
                continue
            logger.info(
                f"Sheet '{sheet_name}' accepted with all required columns present."
            )
            valid_sheets.append((sheet_name, df))

        if not valid_sheets:
            msg = (
                "No sheets in the Excel file contain the required columns: "
                f"{', '.join(required_columns)}"
            )
            print(msg)
            logger.error(msg)
            return [], False

        file_stem = Path(file_path).stem
        sheet_info_list = []

        if len(valid_sheets) == 1:
            # Single valid sheet : legacytablename = FILE NAME
            sheet_name, df = valid_sheets[0]
            legacy_table = file_stem
            print(f"Legacy table extracted from filename: {legacy_table}")
            logger.info(
                f"Single valid sheet '{sheet_name}'. "
                f"Using file name '{file_stem}' as legacytablename."
            )
            sheet_info_list.append((sheet_name, df, legacy_table))
        else:
            # Multiple valid sheets : legacytablename = SHEET NAME
            logger.info(
                "Multiple valid sheets detected. Using sheet names as legacytablename."
            )
            for sheet_name, df in valid_sheets:
                legacy_table = sheet_name
                print(f"Sheet '{sheet_name}' : legacy table: {legacy_table}")
                logger.info(f"Sheet '{sheet_name}' : legacytablename='{legacy_table}'")
                sheet_info_list.append((sheet_name, df, legacy_table))
        return sheet_info_list, True

    except Exception as e:
        print(f"Error reading file: {str(e)}")
        logger.error(f"Error reading file '{file_path}': {str(e)}", exc_info=True)
        return [], False


def extract_target_pairs(input_df: pd.DataFrame):
    # Use drop_duplicates to truly get unique pairs
    pairs = (
        input_df[["targettablename", "targetfieldname"]]
        .drop_duplicates()
        .itertuples(index=False, name=None)
    )
    target_pairs = list(pairs)
    logger = logging.getLogger(__name__)
    logger.info(
        f"Extracted {len(target_pairs)} unique target table/field pairs from input file"
    )
    return target_pairs


def process_mapping_request_from_file(
    file_path: str,
    legacy_system: str,
    target_system: str,
    export_type: str,
    output_path: str = None,
    project_name: str = None,
) -> tuple:

    logger = logging.getLogger(__name__)
    logger.info("Starting enhanced mapping process")
    logger.info(
        f"Input: '{Path(file_path).name}', "
        f"Legacy: '{legacy_system}', Target: '{target_system}', "
        f"Export: '{export_type}', Project: '{project_name}'"
    )

    print("ENHANCED MAPPING PROCESS")
    print("\n Reading input file...")
    sheet_info_list, validation_ok = read_input_file(file_path)

    if not validation_ok or not sheet_info_list:
        print("File validation failed")
        logger.error("Process aborted: file validation failed")
        return False, None, 0

    print(f"  Found {len(sheet_info_list)} valid sheet(s) to process")

    print("\n Connecting to database...")
    conn = None
    try:
        conn = get_connection()
        print("Database connection established")
        logger.info("Database connection established")

        print("\n Fetching project IDs...")
        print(f"  Legacy System: {legacy_system}")
        print(f"  Target System: {target_system}")

        filters = {col: [] for col in FILTER_COLUMNS}
        filters["legacysystem"] = [legacy_system]
        filters["targetsystem"] = [target_system]

        if project_name:
            filters["projectname"] = [project_name]
            logger.info(f"Applying project filter: {project_name}")

        pj_query, pj_params = build_projectid_query(filters)
        project_rows = execute_query(conn, pj_query, pj_params)
        project_ids = [r["projectid"] for r in project_rows]

        logger.info(f"Found project IDs: {project_ids}")
        print(f"Found {len(project_ids)} matching project(s)")

        if not project_ids:
            print("No matching projects found for the selected systems")
            logger.warning("No projects found for selected systems")
            return False, None, 0

        print("\n Fetching mappings from child table...")
        tb_query, tb_params = build_tableb_query(project_ids)
        all_child_rows = execute_query(conn, tb_query, tb_params)
        logger.info(f"Child table query returned {len(all_child_rows)} row(s)")

        rows = []

        # Process each sheet separately with its own legacy_table
        for sheet_name, input_df, legacy_table in sheet_info_list:
            print(f"  Processing sheet: {sheet_name} (Legacy table: {legacy_table})")
            logger.info(
                f"Processing sheet '{sheet_name}' with legacytablename='{legacy_table}'"
            )

            # Extract target pairs from this sheet
            target_pairs = extract_target_pairs(input_df)
            print(f"    Found {len(target_pairs)} unique target table/field pair(s)")
            logger.info(
                f"Sheet '{sheet_name}' -> {len(target_pairs)} target pair(s) extracted"
            )

            if not target_pairs:
                continue

            target_pair_set = {
                (str(t).lower(), str(f).lower()) for (t, f) in target_pairs
            }
            legacy_lower = str(legacy_table).lower()

            # Filter child rows for this legacy table and target pairs
            matched_for_sheet = 0
            for row in all_child_rows:
                lt = str(row.get("legacytablename", "")).lower()
                tt = str(row.get("targettablename", "")).lower()
                tf = str(row.get("targetfieldname", "")).lower()

                # Match on legacy table and target field pairs
                if lt == legacy_lower and (tt, tf) in target_pair_set:
                    rows.append(row)
                    matched_for_sheet += 1

            print(f"     Matched {matched_for_sheet} mapping record(s) for this sheet")
            logger.info(
                f"Sheet '{sheet_name}' -> {matched_for_sheet} row(s) matched "
                "after filtering"
            )

        logger.info(f"Total filtered mapping rows across all sheets: {len(rows)}")
        print(f"Found {len(rows)} mapping record(s) after filtering all sheets")

        if not rows:
            print("No mappings found for the given file and systems")
            logger.warning("No mapping records found")
            return False, None, 0

        # Step 5: Export results using existing exporters
        print("\n Exporting results...")

        input_path = Path(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if output_path:
            output_file = Path(output_path)
        else:
            if export_type == "standard":
                name = f"mapping_standard_{input_path.stem}_{timestamp}.xlsx"
            else:
                name = f"mapping_spec_{input_path.stem}_{timestamp}.xlsx"
            output_file = input_path.parent / name

        if export_type == "standard":
            export_standard_excel(rows, str(output_file))
        else:
            export_summarized_excel(rows, str(output_file))

        print("\n" + "=" * 70)
        print("MAPPING PROCESS COMPLETED SUCCESSFULLY!")
        print("=" * 70)
        print(f"\nInput file: {file_path}")
        print(f"Output file: {output_file}")
        print(f"Export type: {export_type.capitalize()}")
        print(f"Total records: {len(rows)}")

        logger.info("Enhanced mapping process completed successfully")
        logger.info(f"Output file: {output_file}")

        return True, str(output_file), len(rows)

    except Exception as e:
        print(f"\nUnexpected error: {str(e)}")
        logger.error("Unexpected error during mapping process", exc_info=True)
        return False, None, 0

    finally:
        if conn:
            conn.close()
            print("Database connection closed")
            logger.info("Database connection closed")


def run_dynamic_query(conn_params: dict, excel_path: str = "output.xlsx") -> None:
    """
    End-to-end: fetch parenttable, prompt filters, fetch project IDs, query childtable,
    and export results based on user choice: standard or summarized.
    """
    conn = None
    try:
        conn = get_connection()
        conn_output = get_outputconnection()
        rows_a = fetch_table_a(conn)
        filters = prompt_cascading_filters(rows_a)
        pj_query, pj_params = build_projectid_query(filters)
        project_rows = execute_query(conn, pj_query, pj_params)
        project_ids = [r["projectid"] for r in project_rows]
        print(f"Selected project ID(s): {project_ids}")
        tb_query, tb_params = build_tableb_query(project_ids)
        child_rows = execute_query(conn, tb_query, tb_params)
        # Convert to DataFrame for grouping
        action = questionary.select(
            "What would you like to do with the output?",
            choices=[
                Choice(
                    "Generate mapping spec excels for all subdomains",
                    "excel_per_subdomain",
                ),
                Choice("Export standard Excel", "standard_excel"),
                Choice("Store output in Postgres", "to_postgres"),
            ],
        ).ask()

        if action == "excel_per_subdomain":
            df = pd.DataFrame(child_rows)
            for subdomain, group in df.groupby("subdomainname"):
                fname = f"{subdomain}.xlsx"
                print(f"Exporting mapping spec for '{subdomain}' → {fname}")
                export_summarized_excel(group.to_dict(orient="records"), fname)

        elif action == "standard_excel":
            print(f"Exporting standard Excel → {excel_path}")
            export_standard_excel(child_rows, excel_path)

        elif action == "to_postgres":
            print("Storing output in Postgres…")
            desired = [
                "mappingid",
                "domainname",
                "subdomainname",
                "targettablename",  # was targettablename in source
                "targettabledescription",
                "targettableinscope",
                "targettablerequired",
                "primarytable",
                "selectionlogic",
                "columnnumber",
                "targetfieldname",
                "targetfielddescription",
                "datatype",
                "required",
                "legacytablename",
                "legacytabledescription",
                "legacyfieldname",
                "legacyfielddescription",
                "dataqualityrule",
                "translationrule",
                "crossreferences",
                "comments",
                "datasensitivitytags",
                "functionaltargettablename",
                "columnstatus",
                "exampletargetdatavalue",
                "LOV",
                "isCDE",
                "additionalinformation",
                "dependenttabs",
                "isMasterData",
                "isDataMigrationInternal",
            ]
            store_output_to_db(
                child_rows,
                conn_output,
                table_name=postgres_outputtable,
                column_mapping=COLUMN_MAPPING,
                columns_to_insert=desired,
            )
            print("Done. Check your_output_table for inserted rows.")

        else:
            print("No valid action selected. Exiting.")

    finally:
        if conn:
            conn.close()
            logger.info("Connection closed.")


desired = [
    "mappingid",
    "domainname",
    "subdomainname",
    "targettablename",  # was targettablename in source
    "targettabledescription",
    "targettableinscope",
    "targettablerequired",
    "primarytable",
    "selectionlogic",
    "columnnumber",
    "targetfieldname",
    "targetfielddescription",
    "datatype",
    "required",
    "legacytablename",
    "legacytabledescription",
    "legacyfieldname",
    "legacyfielddescription",
    "dataqualityrule",
    "translationrule",
    "crossreferences",
    "comments",
    "datasensitivitytags",
    "functionaltargettablename",
    "columnstatus",
    "exampletargetdatavalue",
    "LOV",
    "isCDE",
    "additionalinformation",
    "dependenttabs",
    "isMasterData",
    "isDataMigrationInternal",
]
WORKDAY_ORACLE_MAPPING_TABLE = (
    "map_maestro.workdayoraclemap"  # UPDATE with your actual table name
)


def fetch_workday_oracle_mapping(conn):
    """Fetch Workday-Oracle bidirectional mapping from your existing table."""
    query = f"""
        SELECT 
            "Column Number" as column_number,
            "Workday Table Name" as workday_table_name,
            "Workday Field Name" as workday_field_name,
            "Workday Field Description" as workday_field_description,
            "Workday Data Type" as workday_data_type,
            "Required?" as required,
            "Example Target Data Value" as example_target_data_value,
            "Oracle Table Name" as oracle_table_name,
            "Oracle Field Name" as oracle_field_name,
            "Oracle Translation Rules" as oracle_translation_rules,
            "Cross References" as cross_references,
            "Comments" as comments,
            "Column Status" as column_status,
            "Path" as path,
            "Parameter" as parameter
        FROM {WORKDAY_ORACLE_MAPPING_TABLE}
        WHERE "Column Status" IS NULL OR "Column Status" != 'Inactive'
        ORDER BY "Column Number"
    """

    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query)
            rows = cur.fetchall()
            logger.info(f"Fetched {len(rows)} Workday-Oracle mapping rows")
            return rows
    except psycopg2.Error as e:
        logger.error(f"Error fetching Workday-Oracle mapping: {e.pgerror or e}")
        return []


def check_transitivity_availability(conn, legacy_system, current_target):
    """
    Check if transitivity mapping is available for the selected legacy and target systems.
    Returns list of available transitivity targets.
    """
    available_targets = []

    # If current target is neither Oracle nor Workday, transitivity is possible
    if current_target not in ["Oracle", "Workday"]:
        # Check if we have mappings to Oracle or Workday for this legacy system
        query = """
            SELECT DISTINCT targetsystem
            FROM map_maestro.parentprojectdetails
            WHERE legacysystem = %s AND targetsystem IN ('Oracle', 'Workday')
        """

        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(query, (legacy_system,))
                intermediates = [r["targetsystem"] for r in cur.fetchall()]

                # Check if WO mapping table exists and has data
                wo_mappings = fetch_workday_oracle_mapping(conn)
                if wo_mappings:
                    if "Oracle" in intermediates:
                        available_targets.append(
                            "Workday"
                        )  # Can map to Workday via Oracle
                    if "Workday" in intermediates:
                        available_targets.append(
                            "Oracle"
                        )  # Can map to Oracle via Workday
        except Exception as e:
            logger.error(f"Error checking transitivity: {e}")

    return available_targets


def apply_transitivity_mapping_debug(
    legacy_rows, wo_mappings, source_target, final_target
):
    """
    Apply transitivity with enhanced debugging to see why mappings aren't matching.
    """
    transitive_rows = []
    translation_stats = {
        "total": 0,
        "mapped": 0,
        "no_mapping": 0,
        "multiple_mappings": 0,
    }

    # Sample data for debugging
    logger.info("=" * 80)
    logger.info("DEBUGGING TRANSITIVITY MAPPING")
    logger.info("=" * 80)

    if final_target == "Oracle" and source_target == "Workday":
        # Log sample of WO mapping data
        logger.info("\nWorkday-Oracle Mapping Table (first 5 rows):")
        for i, row in enumerate(wo_mappings[:5]):
            logger.info(
                f"  {i+1}. Workday: '{row['workday_table_name']}' . '{row['workday_field_name']}' -> Oracle: '{row['oracle_table_name']}' . '{row['oracle_field_name']}'"
            )

        # Log sample of source data
        logger.info("\nSource PeopleSoft->Workday Mappings (first 10 rows):")
        for i, row in enumerate(legacy_rows[:10]):
            logger.info(
                f"  {i+1}. Table: '{row.get('targettablename')}' , Field: '{row.get('targetfieldname')}'"
            )

        # Get unique table/field combinations from both datasets
        wo_tables = set(
            row["workday_table_name"]
            for row in wo_mappings
            if row["workday_table_name"]
        )
        wo_fields = set(
            row["workday_field_name"]
            for row in wo_mappings
            if row["workday_field_name"]
        )

        source_tables = set(
            row.get("targettablename")
            for row in legacy_rows
            if row.get("targettablename")
        )
        source_fields = set(
            row.get("targetfieldname")
            for row in legacy_rows
            if row.get("targetfieldname")
        )

        logger.info(
            f"\nWorkday-Oracle mapping has {len(wo_tables)} unique tables and {len(wo_fields)} unique fields"
        )
        logger.info(
            f"Source data has {len(source_tables)} unique tables and {len(source_fields)} unique fields"
        )

        # Check for any overlap
        table_overlap = wo_tables.intersection(source_tables)
        field_overlap = wo_fields.intersection(source_fields)

        logger.info(f"\nTable name overlap: {len(table_overlap)} tables")
        if table_overlap:
            logger.info(f"  Matching tables: {list(table_overlap)[:5]}")
        else:
            logger.info(f"  NO TABLE OVERLAP!")
            logger.info(f"  WO Tables (sample): {list(wo_tables)[:5]}")
            logger.info(f"  Source Tables (sample): {list(source_tables)[:5]}")

        logger.info(f"\nField name overlap: {len(field_overlap)} fields")
        if field_overlap:
            logger.info(f"  Matching fields: {list(field_overlap)[:10]}")
        else:
            logger.info(f"  NO FIELD OVERLAP!")
            logger.info(f"  WO Fields (sample): {list(wo_fields)[:10]}")
            logger.info(f"  Source Fields (sample): {list(source_fields)[:10]}")

        # Create Workday to Oracle lookup
        wo_lookup = {}
        for row in wo_mappings:
            # Try with exact match
            key = (row["workday_table_name"], row["workday_field_name"])
            if key not in wo_lookup:
                wo_lookup[key] = []
            wo_lookup[key].append(
                {
                    "oracle_table_name": row["oracle_table_name"],
                    "oracle_field_name": row["oracle_field_name"],
                    "oracle_translation_rules": row.get("oracle_translation_rules", ""),
                    "cross_references": row.get("cross_references", ""),
                }
            )

            # Also try case-insensitive and trimmed versions
            key_lower = (
                (
                    str(row["workday_table_name"]).strip().lower()
                    if row["workday_table_name"]
                    else ""
                ),
                (
                    str(row["workday_field_name"]).strip().lower()
                    if row["workday_field_name"]
                    else ""
                ),
            )
            wo_lookup[key_lower] = wo_lookup[key]

        logger.info(f"\nCreated lookup with {len(wo_lookup)} entries")
        logger.info("=" * 80)

        # Track first few mismatches for debugging
        mismatch_samples = []

        for idx, row in enumerate(legacy_rows):
            translation_stats["total"] += 1

            # Try exact match
            lookup_key = (row["targettablename"], row["targetfieldname"])

            # Try case-insensitive match
            lookup_key_lower = (
                (
                    str(row["targettablename"]).strip().lower()
                    if row.get("targettablename")
                    else ""
                ),
                (
                    str(row["targetfieldname"]).strip().lower()
                    if row.get("targetfieldname")
                    else ""
                ),
            )

            new_row = dict(row)

            # Try both exact and case-insensitive
            if lookup_key in wo_lookup:
                mappings = wo_lookup[lookup_key]
                found = True
            elif lookup_key_lower in wo_lookup:
                mappings = wo_lookup[lookup_key_lower]
                found = True
            else:
                found = False
                mappings = []

            if found and mappings:
                if len(mappings) == 1:
                    trans = mappings[0]
                    new_row["targettablename"] = trans["oracle_table_name"]
                    new_row["targetfieldname"] = trans["oracle_field_name"]
                    new_row["targetsystem"] = "Oracle"

                    existing_trans = new_row.get("translationrule", "") or ""
                    oracle_trans = trans["oracle_translation_rules"] or ""
                    if existing_trans and oracle_trans:
                        new_row["translationrule"] = (
                            f"{existing_trans} -> {oracle_trans}"
                        )
                    elif oracle_trans:
                        new_row["translationrule"] = oracle_trans

                    if trans["cross_references"]:
                        existing_refs = new_row.get("crossreferences", "") or ""
                        new_row["crossreferences"] = (
                            f"{existing_refs}; {trans['cross_references']}".strip("; ")
                        )

                    new_row["translation_status"] = "Mapped"
                    new_row["translation_notes"] = (
                        f"Successfully mapped from Workday to Oracle"
                    )
                    translation_stats["mapped"] += 1
                else:
                    trans = mappings[0]
                    new_row["targettablename"] = trans["oracle_table_name"]
                    new_row["targetfieldname"] = trans["oracle_field_name"]
                    new_row["targetsystem"] = "Oracle"

                    new_row["translation_status"] = "Multiple Mappings"
                    new_row["translation_notes"] = (
                        f"WARNING: {len(mappings)} Oracle mappings found. Using first match."
                    )
                    translation_stats["multiple_mappings"] += 1
            else:
                new_row["translation_status"] = "No Mapping"
                new_row["translation_notes"] = (
                    f"No Workday->Oracle mapping found for {lookup_key[0]}.{lookup_key[1]}"
                )
                new_row["targetsystem"] = "Oracle"
                translation_stats["no_mapping"] += 1

                # Save first few mismatches for debugging
                if len(mismatch_samples) < 5:
                    mismatch_samples.append(
                        {
                            "table": lookup_key[0],
                            "field": lookup_key[1],
                            "table_lower": lookup_key_lower[0],
                            "field_lower": lookup_key_lower[1],
                        }
                    )

            existing_comments = new_row.get("comments", "") or ""
            trans_note = f"[Transitive: Legacy→Workday→Oracle]"
            new_row["comments"] = f"{existing_comments}\n{trans_note}".strip()

            transitive_rows.append(new_row)

        # Log mismatch samples
        if mismatch_samples:
            logger.info("\nSample of fields that didn't match:")
            for i, mm in enumerate(mismatch_samples, 1):
                logger.info(
                    f"  {i}. Table: '{mm['table']}' (lower: '{mm['table_lower']}'), Field: '{mm['field']}' (lower: '{mm['field_lower']}')"
                )

    elif final_target == "Workday" and source_target == "Oracle":
        # Similar logic for Oracle -> Workday
        logger.info("\nOracle-Workday Mapping (reverse direction)")
        logger.info("This path uses the same WO mapping table in reverse")

        wo_lookup = {}
        for row in wo_mappings:
            key = (row["oracle_table_name"], row["oracle_field_name"])
            if key not in wo_lookup:
                wo_lookup[key] = []
            wo_lookup[key].append(
                {
                    "workday_table_name": row["workday_table_name"],
                    "workday_field_name": row["workday_field_name"],
                    "workday_field_description": row.get(
                        "workday_field_description", ""
                    ),
                    "workday_data_type": row.get("workday_data_type", ""),
                    "oracle_translation_rules": row.get("oracle_translation_rules", ""),
                    "cross_references": row.get("cross_references", ""),
                }
            )

        for row in legacy_rows:
            translation_stats["total"] += 1
            lookup_key = (row["targettablename"], row["targetfieldname"])
            new_row = dict(row)

            if lookup_key in wo_lookup:
                mappings = wo_lookup[lookup_key]

                if len(mappings) == 1:
                    trans = mappings[0]
                    new_row["targettablename"] = trans["workday_table_name"]
                    new_row["targetfieldname"] = trans["workday_field_name"]
                    new_row["targetfielddescription"] = trans.get(
                        "workday_field_description", ""
                    )
                    new_row["datatype"] = trans.get(
                        "workday_data_type", new_row.get("datatype", "")
                    )
                    new_row["targetsystem"] = "Workday"

                    existing_trans = new_row.get("translationrule", "") or ""
                    if trans["oracle_translation_rules"]:
                        new_row["translationrule"] = (
                            f"{existing_trans} -> Reverse: {trans['oracle_translation_rules']}".strip(
                                " -> "
                            )
                        )

                    if trans["cross_references"]:
                        existing_refs = new_row.get("crossreferences", "") or ""
                        new_row["crossreferences"] = (
                            f"{existing_refs}; {trans['cross_references']}".strip("; ")
                        )

                    new_row["translation_status"] = "Mapped"
                    new_row["translation_notes"] = (
                        f"Successfully mapped from Oracle to Workday"
                    )
                    translation_stats["mapped"] += 1
                else:
                    trans = mappings[0]
                    new_row["targettablename"] = trans["workday_table_name"]
                    new_row["targetfieldname"] = trans["workday_field_name"]
                    new_row["targetsystem"] = "Workday"

                    new_row["translation_status"] = "Multiple Mappings"
                    new_row["translation_notes"] = (
                        f"WARNING: {len(mappings)} Workday mappings found. Using first match."
                    )
                    translation_stats["multiple_mappings"] += 1
            else:
                new_row["translation_status"] = "No Mapping"
                new_row["translation_notes"] = (
                    f"No Oracle->Workday mapping found for {lookup_key[0]}.{lookup_key[1]}"
                )
                new_row["targetsystem"] = "Workday"
                translation_stats["no_mapping"] += 1

            existing_comments = new_row.get("comments", "") or ""
            trans_note = f"[Transitive: Legacy→Oracle→Workday]"
            new_row["comments"] = f"{existing_comments}\n{trans_note}".strip()

            transitive_rows.append(new_row)

    # Log statistics
    logger.info("\n" + "=" * 80)
    logger.info(f"Translation Statistics:")
    logger.info(f"  Total records: {translation_stats['total']}")
    logger.info(f"  Successfully mapped: {translation_stats['mapped']}")
    logger.info(f"  No mapping found: {translation_stats['no_mapping']}")
    logger.info(f"  Multiple mappings: {translation_stats['multiple_mappings']}")
    logger.info(
        f"  Success rate: {translation_stats['mapped']/translation_stats['total']*100:.1f}%"
    )
    logger.info("=" * 80)

    return transitive_rows


def apply_transitivity_mapping(legacy_rows, wo_mappings, source_target, final_target):
    """
    Apply transitivity to convert legacy mappings to new target system.
    Now includes translation status tracking.
    """
    transitive_rows = []
    translation_stats = {
        "total": 0,
        "mapped": 0,
        "no_mapping": 0,
        "multiple_mappings": 0,
    }

    if final_target == "Oracle" and source_target == "Workday":
        # Create Workday to Oracle lookup
        wo_lookup = {}
        for row in wo_mappings:
            key = (row["workday_table_name"], row["workday_field_name"])
            if key not in wo_lookup:
                wo_lookup[key] = []
            wo_lookup[key].append(
                {
                    "oracle_table_name": row["oracle_table_name"],
                    "oracle_field_name": row["oracle_field_name"],
                    "oracle_translation_rules": row.get("oracle_translation_rules", ""),
                    "cross_references": row.get("cross_references", ""),
                }
            )

        for row in legacy_rows:
            translation_stats["total"] += 1
            lookup_key = (row["targettablename"], row["targetfieldname"])
            new_row = dict(row)

            # Add translation status field
            if lookup_key in wo_lookup:
                mappings = wo_lookup[lookup_key]

                if len(mappings) == 1:
                    # Single mapping found - ideal case
                    trans = mappings[0]
                    new_row["targettablename"] = trans["oracle_table_name"]
                    new_row["targetfieldname"] = trans["oracle_field_name"]
                    new_row["targetsystem"] = "Oracle"

                    # Combine translation rules
                    existing_trans = new_row.get("translationrule", "") or ""
                    oracle_trans = trans["oracle_translation_rules"] or ""
                    if existing_trans and oracle_trans:
                        new_row["translationrule"] = (
                            f"{existing_trans} -> {oracle_trans}"
                        )
                    elif oracle_trans:
                        new_row["translationrule"] = oracle_trans

                    # Add cross references
                    if trans["cross_references"]:
                        existing_refs = new_row.get("crossreferences", "") or ""
                        new_row["crossreferences"] = (
                            f"{existing_refs}; {trans['cross_references']}".strip("; ")
                        )

                    # Set translation status
                    new_row["translation_status"] = "Mapped"
                    new_row["translation_notes"] = (
                        f"Successfully mapped from Workday to Oracle"
                    )
                    translation_stats["mapped"] += 1

                else:
                    # Multiple mappings found
                    trans = mappings[0]  # Use first one
                    new_row["targettablename"] = trans["oracle_table_name"]
                    new_row["targetfieldname"] = trans["oracle_field_name"]
                    new_row["targetsystem"] = "Oracle"

                    new_row["translation_status"] = "Multiple Mappings"
                    new_row["translation_notes"] = (
                        f"WARNING: {len(mappings)} Oracle mappings found for Workday field {lookup_key}. Using first match."
                    )
                    translation_stats["multiple_mappings"] += 1

                    logger.warning(
                        f"Multiple mappings for {lookup_key}: {len(mappings)}"
                    )
            else:
                # No mapping found
                new_row["translation_status"] = "No Mapping"
                new_row["translation_notes"] = (
                    f"WARNING: No Workday->Oracle mapping found for {lookup_key[0]}.{lookup_key[1]}"
                )
                new_row["targetsystem"] = "Oracle"
                translation_stats["no_mapping"] += 1

                logger.warning(f"No Workday->Oracle mapping found for {lookup_key}")

            # Update comments with transitivity note
            existing_comments = new_row.get("comments", "") or ""
            trans_note = f"[Transitive: Legacy→Workday→Oracle]"
            new_row["comments"] = f"{existing_comments}\n{trans_note}".strip()

            transitive_rows.append(new_row)

    elif final_target == "Workday" and source_target == "Oracle":
        # Create Oracle to Workday lookup
        wo_lookup = {}
        for row in wo_mappings:
            key = (row["oracle_table_name"], row["oracle_field_name"])
            if key not in wo_lookup:
                wo_lookup[key] = []
            wo_lookup[key].append(
                {
                    "workday_table_name": row["workday_table_name"],
                    "workday_field_name": row["workday_field_name"],
                    "workday_field_description": row.get(
                        "workday_field_description", ""
                    ),
                    "workday_data_type": row.get("workday_data_type", ""),
                    "oracle_translation_rules": row.get("oracle_translation_rules", ""),
                    "cross_references": row.get("cross_references", ""),
                }
            )

        for row in legacy_rows:
            translation_stats["total"] += 1
            lookup_key = (row["targettablename"], row["targetfieldname"])
            new_row = dict(row)

            if lookup_key in wo_lookup:
                mappings = wo_lookup[lookup_key]

                if len(mappings) == 1:
                    trans = mappings[0]
                    new_row["targettablename"] = trans["workday_table_name"]
                    new_row["targetfieldname"] = trans["workday_field_name"]
                    new_row["targetfielddescription"] = trans.get(
                        "workday_field_description", ""
                    )
                    new_row["datatype"] = trans.get(
                        "workday_data_type", new_row.get("datatype", "")
                    )
                    new_row["targetsystem"] = "Workday"

                    # Handle translation rules
                    existing_trans = new_row.get("translationrule", "") or ""
                    if trans["oracle_translation_rules"]:
                        new_row["translationrule"] = (
                            f"{existing_trans} -> Reverse: {trans['oracle_translation_rules']}".strip(
                                " -> "
                            )
                        )

                    # Update cross references
                    if trans["cross_references"]:
                        existing_refs = new_row.get("crossreferences", "") or ""
                        new_row["crossreferences"] = (
                            f"{existing_refs}; {trans['cross_references']}".strip("; ")
                        )

                    new_row["translation_status"] = "Mapped"
                    new_row["translation_notes"] = (
                        f"Successfully mapped from Oracle to Workday"
                    )
                    translation_stats["mapped"] += 1
                else:
                    trans = mappings[0]
                    new_row["targettablename"] = trans["workday_table_name"]
                    new_row["targetfieldname"] = trans["workday_field_name"]
                    new_row["targetsystem"] = "Workday"

                    new_row["translation_status"] = "Multiple Mappings"
                    new_row["translation_notes"] = (
                        f"WARNING: {len(mappings)} Workday mappings found for Oracle field {lookup_key}. Using first match."
                    )
                    translation_stats["multiple_mappings"] += 1
            else:
                new_row["translation_status"] = "No Mapping"
                new_row["translation_notes"] = (
                    f"WARNING: No Oracle->Workday mapping found for {lookup_key[0]}.{lookup_key[1]}"
                )
                new_row["targetsystem"] = "Workday"
                translation_stats["no_mapping"] += 1

            existing_comments = new_row.get("comments", "") or ""
            trans_note = f"[Transitive: Legacy→Oracle→Workday]"
            new_row["comments"] = f"{existing_comments}\n{trans_note}".strip()

            transitive_rows.append(new_row)

    # Log statistics
    logger.info(f"Translation Statistics:")
    logger.info(f"  Total records: {translation_stats['total']}")
    logger.info(f"  Successfully mapped: {translation_stats['mapped']}")
    logger.info(f"  No mapping found: {translation_stats['no_mapping']}")
    logger.info(f"  Multiple mappings: {translation_stats['multiple_mappings']}")
    logger.info(
        f"  Success rate: {translation_stats['mapped']/translation_stats['total']*100:.1f}%"
    )

    return transitive_rows


# Updated transitivity functions for your existing table structure
# Add these to your MapMaestro application


def fetch_workday_oracle_mapping(conn, mapping_type="workday_to_oracle"):
    """
    Fetch Workday-Oracle bidirectional mapping from your existing table.
    mapping_type: "workday_to_oracle" or "oracle_to_workday"
    """
    # Adjust table name based on your actual table
    table_name = (
        "map_maestro.workday_oracle_mappings"  # Replace with your actual table name
    )

    query = f"""
        SELECT 
            column_number,
            workday_table_name,
            workday_field_name,
            workday_field_description,
            workday_data_type,
            required,
            example_target_data_value,
            oracle_table_name,
            oracle_field_name,
            oracle_translation_rules,
            cross_references,
            comments,
            column_status,
            path,
            parameter
        FROM {table_name}
        WHERE column_status = 'Active'  -- Assuming you have active/inactive status
    """

    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query)
            rows = cur.fetchall()
            logger.info(f"Fetched {len(rows)} Workday-Oracle mapping rows")
            return rows
    except psycopg2.Error as e:
        logger.error(f"Error fetching Workday-Oracle mapping: {e.pgerror or e}")
        return []


def fetch_legacy_mapping_for_transitivity(
    conn, legacy_system, intermediate_system, filters=None
):
    """
    Fetch existing legacy system mapping to either Workday or Oracle.
    """
    query = """
        SELECT DISTINCT
            c.*,
            p.legacysystem,
            p.targetsystem,
            p.filename,
            p.projectname,
            p.domainname,
            p.subdomainname
        FROM map_maestro.childattributemappingdetails c
        JOIN map_maestro.parentprojectdetails p ON c.projectid = p.projectid
        WHERE p.legacysystem = %s AND p.targetsystem = %s
    """

    params = [legacy_system, intermediate_system]

    # Add additional filters if provided
    if filters:
        for col, vals in filters.items():
            if vals and col not in ["legacysystem", "targetsystem"]:
                placeholders = ",".join(["%s"] * len(vals))
                query += f" AND p.{col} IN ({placeholders})"
                params.extend(vals)

    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query, params)
            rows = cur.fetchall()
            logger.info(
                f"Fetched {len(rows)} mapping rows for {legacy_system} -> {intermediate_system}"
            )
            return rows
    except psycopg2.Error as e:
        logger.error(f"Error fetching legacy mapping: {e.pgerror or e}")
        raise


def generate_transitive_mapping_with_existing_table(
    legacy_mapping, wo_mapping, target_system, legacy_system
):
    """
    Generate transitive mapping using your existing Workday-Oracle mapping table.
    """
    transitive_rows = []

    # Create lookup dictionaries based on direction
    if target_system == "Oracle":
        # Legacy -> Workday exists, need to map Workday fields to Oracle
        wo_lookup = {
            (row["workday_table_name"], row["workday_field_name"]): {
                "oracle_table_name": row["oracle_table_name"],
                "oracle_field_name": row["oracle_field_name"],
                "oracle_translation_rules": row.get("oracle_translation_rules", ""),
                "cross_references": row.get("cross_references", ""),
                "workday_data_type": row.get("workday_data_type", ""),
                "column_status": row.get("column_status", "Active"),
            }
            for row in wo_mapping
        }

        # Process each legacy->Workday mapping row
        for row in legacy_mapping:
            lookup_key = (row["targettablename"], row["targetfieldname"])

            if lookup_key in wo_lookup:
                trans = wo_lookup[lookup_key]

                # Create new transitive row
                new_row = dict(row)

                # Update with Oracle target information
                new_row["targettablename"] = trans["oracle_table_name"]
                new_row["targetfieldname"] = trans["oracle_field_name"]
                new_row["targetsystem"] = "Oracle"

                # Combine translation rules
                existing_trans = new_row.get("translationrule", "") or ""
                oracle_trans = trans["oracle_translation_rules"] or ""
                if existing_trans and oracle_trans:
                    new_row["translationrule"] = f"{existing_trans} -> {oracle_trans}"
                elif oracle_trans:
                    new_row["translationrule"] = oracle_trans

                # Update cross references
                if trans["cross_references"]:
                    existing_refs = new_row.get("crossreferences", "") or ""
                    new_row["crossreferences"] = (
                        f"{existing_refs}; {trans['cross_references']}".strip("; ")
                    )

                # Add transitivity note to comments
                existing_comments = new_row.get("comments", "") or ""
                trans_note = f"[Transitive: {legacy_system}→Workday→Oracle]"
                new_row["comments"] = f"{existing_comments}\n{trans_note}".strip()

                transitive_rows.append(new_row)
            else:
                # No translation found - include with warning
                logger.warning(f"No Workday->Oracle mapping found for {lookup_key}")
                new_row = dict(row)
                new_row["targetsystem"] = "Oracle"
                existing_comments = new_row.get("comments", "") or ""
                new_row["comments"] = (
                    f"{existing_comments}\n[WARNING: No Workday->Oracle translation found for {lookup_key}]".strip()
                )
                transitive_rows.append(new_row)

    else:  # target_system == "Workday"
        # Legacy -> Oracle exists, need to map Oracle fields to Workday
        wo_lookup = {
            (row["oracle_table_name"], row["oracle_field_name"]): {
                "workday_table_name": row["workday_table_name"],
                "workday_field_name": row["workday_field_name"],
                "workday_field_description": row.get("workday_field_description", ""),
                "workday_data_type": row.get("workday_data_type", ""),
                "oracle_translation_rules": row.get("oracle_translation_rules", ""),
                "cross_references": row.get("cross_references", ""),
            }
            for row in wo_mapping
        }

        # Process each legacy->Oracle mapping row
        for row in legacy_mapping:
            lookup_key = (row["targettablename"], row["targetfieldname"])

            if lookup_key in wo_lookup:
                trans = wo_lookup[lookup_key]

                # Create new transitive row
                new_row = dict(row)

                # Update with Workday target information
                new_row["targettablename"] = trans["workday_table_name"]
                new_row["targetfieldname"] = trans["workday_field_name"]
                new_row["targetfielddescription"] = trans.get(
                    "workday_field_description", ""
                )
                new_row["datatype"] = trans.get(
                    "workday_data_type", new_row.get("datatype", "")
                )
                new_row["targetsystem"] = "Workday"

                # Handle translation rules
                existing_trans = new_row.get("translationrule", "") or ""
                if trans["oracle_translation_rules"]:
                    new_row["translationrule"] = (
                        f"{existing_trans} -> Reverse: {trans['oracle_translation_rules']}".strip(
                            " -> "
                        )
                    )

                # Update cross references
                if trans["cross_references"]:
                    existing_refs = new_row.get("crossreferences", "") or ""
                    new_row["crossreferences"] = (
                        f"{existing_refs}; {trans['cross_references']}".strip("; ")
                    )

                # Add transitivity note
                existing_comments = new_row.get("comments", "") or ""
                trans_note = f"[Transitive: {legacy_system}→Oracle→Workday]"
                new_row["comments"] = f"{existing_comments}\n{trans_note}".strip()

                transitive_rows.append(new_row)
            else:
                # No translation found
                logger.warning(f"No Oracle->Workday mapping found for {lookup_key}")
                new_row = dict(row)
                new_row["targetsystem"] = "Workday"
                existing_comments = new_row.get("comments", "") or ""
                new_row["comments"] = (
                    f"{existing_comments}\n[WARNING: No Oracle->Workday translation found for {lookup_key}]".strip()
                )
                transitive_rows.append(new_row)

    return transitive_rows


def run_transitivity_with_existing_table(
    conn, legacy_system, target_system, filters, output_path
):
    """
    Execute transitivity mapping generation using existing table structure.
    """
    try:
        # Determine intermediate system
        intermediate_system = "Workday" if target_system == "Oracle" else "Oracle"

        # Step 1: Get legacy to intermediate mapping
        legacy_to_intermediate = fetch_legacy_mapping_for_transitivity(
            conn, legacy_system, intermediate_system, filters
        )

        if not legacy_to_intermediate:
            raise ValueError(
                f"No mapping found from {legacy_system} to {intermediate_system}. "
                f"Cannot generate transitive mapping."
            )

        # Step 2: Get Workday-Oracle bidirectional mapping
        wo_mapping = fetch_workday_oracle_mapping(conn)

        if not wo_mapping:
            raise ValueError(
                f"No Workday-Oracle mapping found in the translation table. "
                f"Please ensure the workday_oracle_mappings table is populated."
            )

        # Step 3: Generate transitive mapping
        transitive_rows = generate_transitive_mapping_with_existing_table(
            legacy_to_intermediate, wo_mapping, target_system, legacy_system
        )

        if not transitive_rows:
            raise ValueError(
                "Failed to generate transitive mapping. No matching fields found."
            )

        # Step 4: Export to Excel (using your existing export functions)
        export_transitive_mapping_enhanced(
            transitive_rows, legacy_system, target_system, output_path
        )

        return len(transitive_rows)

    except Exception as e:
        logger.error(f"Error in transitivity generation: {str(e)}")
        raise


def clean_rows_for_export(rows, group_by_column="filename"):
    """
    Clean rows for Excel export by removing duplicate/conflicting columns
    and ensuring consistent data within each group.
    """
    if not rows:
        return rows

    # Columns that should be consistent within a group (from parent table)
    parent_columns = [
        "targetsystem",
        "legacysystem",
        "projectname",
        "domainname",
        "subdomainname",
    ]

    # For each group, keep only the first value of parent columns
    cleaned_rows = []
    for row in rows:
        cleaned_row = dict(row)
        # These parent columns are already in the row, no need to do anything special
        # The grouping in pandas will handle duplicates
        cleaned_rows.append(cleaned_row)

    return cleaned_rows


def export_summarized_excel_with_status(rows, filename: str):
    if not rows:
        logger.warning("No data to export.")
        return

    # Clean rows before creating DataFrame
    cleaned_rows = clean_rows_for_export(rows)
    df = pd.DataFrame(cleaned_rows)

    # Check if this is a transitive export (has translation_status column)
    is_transitive = "translation_status" in df.columns

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # If transitive, create a Translation Status summary sheet first
    if is_transitive:
        status_ws = wb.create_sheet("Translation Status", 0)

        # Summary statistics
        status_ws.cell(row=1, column=1, value="Translation Summary Report").font = Font(
            bold=True, size=16
        )
        status_ws.cell(
            row=2,
            column=1,
            value=f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}",
        )

        status_counts = df["translation_status"].value_counts()
        total = len(df)

        status_ws.cell(row=4, column=1, value="Overall Statistics").font = Font(
            bold=True, size=14
        )
        status_ws.cell(row=5, column=1, value="Total Records:")
        status_ws.cell(row=5, column=2, value=total)

        row_num = 6
        for status, count in status_counts.items():
            pct = count / total * 100
            status_ws.cell(row=row_num, column=1, value=f"{status}:")
            status_ws.cell(row=row_num, column=2, value=f"{count} ({pct:.1f}%)")
            row_num += 1

        # Detail list of unmapped fields
        if "No Mapping" in status_counts:
            unmapped = df[df["translation_status"] == "No Mapping"]

            status_ws.cell(
                row=row_num + 2, column=1, value="Fields Without Mapping:"
            ).font = Font(bold=True, size=14)

            # Headers
            headers = [
                "Source Table",
                "Source Field",
                "Target Table",
                "Target Field",
                "Notes",
            ]
            for col_idx, header in enumerate(headers, 1):
                cell = status_ws.cell(row=row_num + 3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
                cell.font = Font(bold=True, color="FFFFFFFF")

            # Data
            data_row = row_num + 4
            for _, row in unmapped.iterrows():
                status_ws.cell(
                    row=data_row, column=1, value=row.get("legacytablename", "")
                )
                status_ws.cell(
                    row=data_row, column=2, value=row.get("legacyfieldname", "")
                )
                status_ws.cell(
                    row=data_row, column=3, value=row.get("targettablename", "")
                )
                status_ws.cell(
                    row=data_row, column=4, value=row.get("targetfieldname", "")
                )
                status_ws.cell(
                    row=data_row, column=5, value=row.get("translation_notes", "")
                )
                data_row += 1

            # Auto-adjust column widths
            for col_idx in range(1, 6):
                status_ws.column_dimensions[get_column_letter(col_idx)].width = 25

    # Create summary sheet
    summary_cols = [
        "targettablename",
        "targettabledescription",
        "targettableinscope",
        "targettablerequired",
        "primarytable",
        "functionaltargettablename",
        "dependenttabs",
        "mappingid",
    ]

    if is_transitive:
        summary_cols.insert(2, "translation_status")

    available_summary_cols = [col for col in summary_cols if col in df.columns]
    summary_df = (
        df[available_summary_cols]
        .drop_duplicates(subset=["targettablename"])
        .reset_index(drop=True)
    )

    if "targettabledescription" in summary_df.columns:
        summary_df = summary_df.rename(
            columns={"targettabledescription": "Description"}
        )

    target_names = sorted(summary_df["targettablename"].unique())

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_df.insert(0, "No.", range(1, len(summary_df) + 1))
    if "mappingid" in summary_df.columns:
        summary_df.sort_values("mappingid", inplace=True)
        summary_df.drop(columns=["mappingid"], inplace=True, errors="ignore")

    header_row = 3
    data_start = header_row + 1
    fill = PatternFill(fill_type="solid", fgColor="FF16365C")

    for col_idx, col_name in enumerate(summary_df.columns, 1):
        display_map = {
            "No.": "No.",
            "targettablename": "Tab",
            "Description": "Description",
            "translation_status": "Translation Status",
            "targettableinscope": "In Scope?",
            "targettablerequired": "Basic Setup",
            "dependenttabs": "Dependent Tabs",
            "primarytable": "Primary Table",
            "functionaltargettablename": "Functional Table Name",
        }
        display_name = display_map.get(col_name, col_name)
        cell = summary_ws.cell(row=header_row, column=col_idx, value=display_name)
        cell.font = Font(bold=True, size=14, color="FFFFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        max_length = max(
            summary_df[col_name].astype(str).map(len).max(), len(display_name)
        )
        summary_ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    for i, row in enumerate(
        summary_df.itertuples(index=False, name=None), start=data_start
    ):
        for col_idx, value in enumerate(row, 1):
            cell = summary_ws.cell(row=i, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left")
            # Color code by translation status
            if is_transitive and col_idx == 3:  # translation_status column
                if value == "Mapped":
                    cell.fill = PatternFill(fill_type="solid", fgColor="FF90EE90")
                elif value == "No Mapping":
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFFFA07A")

    summary_ws.freeze_panes = f"A{data_start}"

    # Create sheets per target table
    for tname in target_names:
        group = df[df["targettablename"] == tname]
        ws = wb.create_sheet(title=str(tname))

        # Mapping specs at top
        ws.cell(row=1, column=1, value="Conversion Name:").font = Font(bold=True)
        ws.cell(row=1, column=2, value=tname)

        desc = summary_df.loc[
            summary_df["targettablename"] == tname, "Description"
        ].values
        desc_val = desc[0] if len(desc) > 0 else ""
        ws.cell(row=2, column=1, value="Description:").font = Font(bold=True)
        ws.cell(row=2, column=2, value=desc_val)

        ws.cell(row=3, column=1, value="Conversion Cycle:").font = Font(bold=True)
        ws.cell(row=3, column=2, value="")

        sel = (
            group["selectionlogic"].iloc[0] if "selectionlogic" in group.columns else ""
        )
        ws.cell(row=4, column=1, value="Selection Rules:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=sel)

        data_start = 7

        # Data columns - add translation status if transitive
        data_cols = [
            "columnnumber",
            "targetfieldname",
            "targetfielddescription",
            "datatype",
            "required",
            "LOV",
            "exampletargetdatavalue",
            "legacytablename",
            "legacyfieldname",
            "legacyfielddescription",
            "translationrule",
            "crossreferences",
            "comments",
            "columnstatus",
            "additionalinformation",
        ]

        if is_transitive:
            data_cols.insert(1, "translation_status")

        available_data_cols = [col for col in data_cols if col in group.columns]

        if "columnnumber" in group.columns:
            group = group.sort_values("columnnumber").reset_index(drop=True)
        else:
            group.insert(0, "columnnumber", range(1, len(group) + 1))

        # Headers
        data_display_map = {
            "columnnumber": "Column Number",
            "translation_status": "Translation Status",
            "targetfieldname": "Field Name",
            "targetfielddescription": "Field Description",
            "datatype": "Data Type",
            "required": "Required?",
            "LOV": "LOV?",
            "exampletargetdatavalue": "Example Target Data Value",
            "legacytablename": "Legacy Table Name",
            "legacyfieldname": "Legacy Field Name",
            "legacyfielddescription": "Legacy Field Description",
            "translationrule": "Translation Rules",
            "crossreferences": "Cross References",
            "comments": "Comments",
            "columnstatus": "Column Status",
            "additionalinformation": "Additional Information",
        }

        for col_idx, col_name in enumerate(available_data_cols, 1):
            display_name = data_display_map.get(col_name, col_name)
            cell = ws.cell(row=data_start, column=col_idx, value=display_name)
            cell.font = Font(bold=True, size=14, color="FFFFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            if col_name in group.columns:
                max_length = max(
                    group[col_name].astype(str).map(len).max(), len(display_name)
                )
            else:
                max_length = len(display_name)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

        # Write rows
        for row_idx, row in enumerate(
            group[available_data_cols].itertuples(index=False, name=None),
            data_start + 1,
        ):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left")

                # Color code translation status
                if (
                    is_transitive
                    and available_data_cols[col_idx - 1] == "translation_status"
                ):
                    if value == "Mapped":
                        cell.fill = PatternFill(fill_type="solid", fgColor="FF90EE90")
                    elif value == "No Mapping":
                        cell.fill = PatternFill(fill_type="solid", fgColor="FFFFA07A")
                    elif value == "Multiple Mappings":
                        cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")

        ws.freeze_panes = f"A{data_start + 1}"

    wb.save(filename)
    logger.info(f"Exported Excel file with status tracking: {filename}")


# Update the run method in your ElegantMapMaestroApp class
def run_transitivity_updated(self, legacy_system, target_system, output_path, filters):
    """
    Updated run_transitivity method for the ElegantMapMaestroApp class.
    """
    conn = get_connection()
    try:
        row_count = run_transitivity_with_existing_table(
            conn, legacy_system, target_system, filters, output_path
        )
        return row_count
    finally:
        conn.close()


def export_transitive_mapping_enhanced(
    rows, legacy_system, target_system, filename=None
):
    """
    Enhanced export function for transitive mappings with summary statistics.
    """
    if not filename:
        filename = f"Transitive_{legacy_system}_to_{target_system}.xlsx"

    if not rows:
        logger.warning("No transitive mapping data to export.")
        return

    df = pd.DataFrame(rows)

    wb = Workbook()

    # Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"

    # Title and metadata
    summary_ws.cell(row=1, column=1, value="Transitive Mapping Report").font = Font(
        bold=True, size=16
    )
    summary_ws.cell(row=2, column=1, value=f"Source: {legacy_system}").font = Font(
        bold=True
    )
    summary_ws.cell(row=3, column=1, value=f"Target: {target_system}").font = Font(
        bold=True
    )
    summary_ws.cell(
        row=4,
        column=1,
        value=f"Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}",
    )

    # Statistics
    summary_ws.cell(row=6, column=1, value="Statistics:").font = Font(bold=True)
    summary_ws.cell(row=7, column=1, value="Total Mappings:")
    summary_ws.cell(row=7, column=2, value=len(rows))

    # Count warnings
    warnings_count = sum(1 for r in rows if "WARNING" in r.get("comments", ""))
    summary_ws.cell(row=8, column=1, value="Mappings with Warnings:")
    summary_ws.cell(row=8, column=2, value=warnings_count)

    # Unique tables
    unique_legacy_tables = (
        df["legacytablename"].nunique() if "legacytablename" in df.columns else 0
    )
    unique_target_tables = (
        df["targettablename"].nunique() if "targettablename" in df.columns else 0
    )
    summary_ws.cell(row=9, column=1, value="Unique Legacy Tables:")
    summary_ws.cell(row=9, column=2, value=unique_legacy_tables)
    summary_ws.cell(row=10, column=1, value="Unique Target Tables:")
    summary_ws.cell(row=10, column=2, value=unique_target_tables)

    # Detailed mapping sheet
    detail_ws = wb.create_sheet("Detailed Mappings")

    # Define columns for detailed sheet
    detail_cols = [
        "targettablename",
        "targetfieldname",
        "targetfielddescription",
        "datatype",
        "required",
        "legacytablename",
        "legacyfieldname",
        "legacyfielddescription",
        "translationrule",
        "crossreferences",
        "comments",
        "columnstatus",
        "domainname",
        "subdomainname",
    ]

    # Filter columns to only those that exist
    available_cols = [col for col in detail_cols if col in df.columns]

    # Write headers
    fill = PatternFill(fill_type="solid", fgColor="FF16365C")
    for col_idx, col_name in enumerate(available_cols, 1):
        cell = detail_ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row_idx, row in enumerate(
        df[available_cols].itertuples(index=False, name=None), 2
    ):
        for col_idx, value in enumerate(row, 1):
            cell = detail_ws.cell(row=row_idx, column=col_idx, value=value)
            # Highlight warnings in red
            if value and isinstance(value, str) and "WARNING" in value:
                cell.font = Font(color="FF0000")

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(available_cols, 1):
        max_length = max(df[col_name].astype(str).map(len).max(), len(col_name))
        detail_ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max_length + 2, 50
        )

    # Freeze panes
    detail_ws.freeze_panes = "A2"

    wb.save(filename)
    logger.info(f"Exported transitive mapping to: {filename}")
    return filename


def parse_uploaded_excel(file_path):
    """
    Parse uploaded Excel file and extract mapping data.
    Returns a dictionary with sheet names as keys and dataframes as values.
    """
    try:
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        sheets_data = {}

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            # Clean column names - remove extra spaces and convert to lowercase for matching
            df.columns = df.columns.str.strip()
            sheets_data[sheet_name] = df
            logger.info(f"Loaded sheet '{sheet_name}' with {len(df)} rows")

        return sheets_data
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise


def validate_excel_structure(sheets_data):
    """
    Validate that the uploaded Excel has required columns.
    Returns validation result and any missing columns.
    """
    required_columns = [
        "targettablename",
        "targetfieldname",
    ]

    optional_columns = [
        "legacyfieldname",
        "legacytablename",
        "datatype",
        "required",
        "translationrule",
        "comments",
        "targetfielddescription",
        "legacyfielddescription",
    ]

    validation_results = {}

    for sheet_name, df in sheets_data.items():
        sheet_cols = [
            col.lower().replace(" ", "").replace("_", "") for col in df.columns
        ]

        missing_required = []
        for req_col in required_columns:
            normalized_req = req_col.lower().replace("_", "")
            if not any(normalized_req in col for col in sheet_cols):
                missing_required.append(req_col)

        available_optional = []
        for opt_col in optional_columns:
            normalized_opt = opt_col.lower().replace("_", "")
            if any(normalized_opt in col for col in sheet_cols):
                available_optional.append(opt_col)

        validation_results[sheet_name] = {
            "valid": len(missing_required) == 0,
            "missing_required": missing_required,
            "available_optional": available_optional,
            "total_rows": len(df),
        }

    return validation_results


def transform_excel_to_mapping_format(df, target_system, legacy_system=None):
    """
    Transform uploaded Excel dataframe to match the internal mapping format.
    Fetches actual mappingid from database by matching target table/field names.
    """

    column_mapping = {
        # Column number variations
        "column number": "columnnumber",
        "columnnumber": "columnnumber",
        "column_number": "columnnumber",
        "col number": "columnnumber",
        # Legacy table/field
        "legacy table name": "legacytablename",
        "legacytablename": "legacytablename",
        "legacy_table_name": "legacytablename",
        "legacy table": "legacytablename",
        "legacy field name": "legacyfieldname",
        "legacyfieldname": "legacyfieldname",
        "legacy_field_name": "legacyfieldname",
        "legacy field": "legacyfieldname",
        "legacy field description": "legacyfielddescription",
        "legacyfielddescription": "legacyfielddescription",
        "legacy_field_description": "legacyfielddescription",
        # Target table/field
        "target table name": "targettablename",
        "targettablename": "targettablename",
        "target_table_name": "targettablename",
        "target table": "targettablename",
        "target field name": "targetfieldname",
        "targetfieldname": "targetfieldname",
        "target_field_name": "targetfieldname",
        "target field": "targetfieldname",
        "target field description": "targetfielddescription",
        "targetfielddescription": "targetfielddescription",
        "target_field_description": "targetfielddescription",
        "description": "targetfielddescription",
        # Data type and attributes
        "data type": "datatype",
        "datatype": "datatype",
        "data_type": "datatype",
        "required": "required",
        "required?": "required",
        "lov": "LOV",
        "lov?": "LOV",
        "list of values": "LOV",
        "example target data value": "exampletargetdatavalue",
        "exampletargetdatavalue": "exampletargetdatavalue",
        "example_target_data_value": "exampletargetdatavalue",
        "example value": "exampletargetdatavalue",
        "example data value": "exampletargetdatavalue",
        # Rules and references
        "translation rule": "translationrule",
        "translation rules": "translationrule",
        "translationrule": "translationrule",
        "translation_rule": "translationrule",
        "translation_rules": "translationrule",
        "cross reference": "crossreferences",
        "cross references": "crossreferences",
        "crossreferences": "crossreferences",
        "cross_references": "crossreferences",
        "comments": "comments",
        "comment": "comments",
        "column status": "columnstatus",
        "columnstatus": "columnstatus",
        "column_status": "columnstatus",
        "status": "columnstatus",
        "additional information": "additionalinformation",
        "additionalinformation": "additionalinformation",
        "additional_information": "additionalinformation",
        "additional info": "additionalinformation",
    }

    normalized_cols = {}
    for col in df.columns:
        normalized = (
            col.lower().strip().replace("_", "").replace(" ", "").replace("?", "")
        )
        if normalized in column_mapping:
            normalized_cols[col] = column_mapping[normalized]
        else:
            if col.lower() in column_mapping:
                normalized_cols[col] = column_mapping[col.lower()]
            else:
                normalized_cols[col] = col

    # Rename columns
    df_renamed = df.rename(columns=normalized_cols)

    df_renamed["targetsystem"] = target_system

    if legacy_system:
        df_renamed["legacysystem"] = legacy_system

    logger.info("Fetching mappingids from childattributemappingdetails table.")
    if (
        "targettablename" in df_renamed.columns
        and "targetfieldname" in df_renamed.columns
    ):
        target_pairs = df_renamed[
            ["targettablename", "targetfieldname"]
        ].drop_duplicates()

        conn = None
        try:
            conn = get_connection()
            mappingid_lookup = {}

            for _, row in target_pairs.iterrows():
                target_table = row["targettablename"]
                target_field = row["targetfieldname"]

                query = """
                    SELECT c.mappingid, c.targettablename, c.targetfieldname, 
                           p.legacysystem, p.targetsystem
                    FROM map_maestro.childattributemappingdetails c
                    JOIN map_maestro.parentprojectdetails p ON c.projectid = p.projectid
                    WHERE c.targettablename = %s 
                      AND c.targetfieldname = %s
                      AND p.targetsystem = %s
                """
                params = [target_table, target_field, target_system]

                if legacy_system:
                    query += " AND p.legacysystem = %s"
                    params.append(legacy_system)

                query += " LIMIT 1"

                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute(query, params)
                    result = cur.fetchone()

                    if result:
                        key = (target_table, target_field)
                        mappingid_lookup[key] = result["mappingid"]

            logger.info(
                f"Found mappingids for {len(mappingid_lookup)} out of {len(target_pairs)} target field pairs"
            )

            # Map mappingids to dataframe rows
            mappingids = []
            not_found_count = 0

            for idx, row in df_renamed.iterrows():
                target_table = row.get("targettablename")
                target_field = row.get("targetfieldname")
                key = (target_table, target_field)

                if key in mappingid_lookup:
                    mappingids.append(mappingid_lookup[key])
                else:
                    # Generate fallback ID if not found in database
                    timestamp = int(time.time() * 1000)
                    unique_str = f"{target_table}_{target_field}_{target_system}_{timestamp}_{idx}"
                    hash_obj = hashlib.md5(unique_str.encode())
                    hash_int = int(hash_obj.hexdigest()[:15], 16)
                    mappingids.append(hash_int)
                    not_found_count += 1
                    logger.warning(
                        f"No mappingid found for {target_table}.{target_field}, generated fallback ID: {hash_int}"
                    )

            df_renamed["mappingid"] = mappingids

            if not_found_count > 0:
                logger.warning(
                    f"Generated fallback mappingids for {not_found_count} rows not found in database"
                )
            else:
                logger.info(
                    f"Successfully mapped all {len(mappingids)} rows to existing mappingids"
                )

        except Exception as e:
            logger.error(f"Error fetching mappingids from database: {e}", exc_info=True)
            # Fallback: generate IDs if database query fails
            logger.warning("Falling back to generated mappingids due to database error")
            timestamp = int(time.time() * 1000)
            mappingids = []
            for idx, row in df_renamed.iterrows():
                target_table = str(row.get("targettablename", ""))
                target_field = str(row.get("targetfieldname", ""))
                unique_str = (
                    f"{target_table}_{target_field}_{target_system}_{timestamp}_{idx}"
                )
                hash_obj = hashlib.md5(unique_str.encode())
                hash_int = int(hash_obj.hexdigest()[:15], 16)
                mappingids.append(hash_int)
            df_renamed["mappingid"] = mappingids

        finally:
            if conn:
                conn.close()
    else:
        logger.error(
            "Cannot fetch mappingids: targettablename or targetfieldname column missing"
        )
        # Generate fallback IDs
        timestamp = int(time.time() * 1000)
        mappingids = []
        for idx in range(len(df_renamed)):
            unique_str = f"unknown_{target_system}_{timestamp}_{idx}"
            hash_obj = hashlib.md5(unique_str.encode())
            hash_int = int(hash_obj.hexdigest()[:15], 16)
            mappingids.append(hash_int)
        df_renamed["mappingid"] = mappingids

    # Add default values for missing columns
    if "datatype" not in df_renamed.columns:
        df_renamed["datatype"] = "VARCHAR(1000)"
    if "required" not in df_renamed.columns:
        df_renamed["required"] = "N"
    if "comments" not in df_renamed.columns:
        df_renamed["comments"] = ""
    if "LOV" not in df_renamed.columns:
        df_renamed["LOV"] = "N"
    if "columnstatus" not in df_renamed.columns:
        df_renamed["columnstatus"] = "Active"

    return df_renamed


class ElegantMapMaestroApp:
    def __init__(self, root):
        self.root = root
        self.setup_window()
        self.setup_styles()
        self.load_data()
        self.create_ui()
        self.setup_bindings()

        self.uploaded_file_path = None
        self.uploaded_data = None
        self.validation_results = None

    def setup_window(self):
        """Configure the main window"""
        try:
            icon_paths = ["logo_resized.ico"]

            for icon_path in icon_paths:
                if os.path.exists(icon_path):
                    self.root.iconbitmap(icon_path)
                    print(f"Icon loaded: {icon_path}")
                    break
            else:
                print("No icon file found - using default system icon")

        except Exception as e:
            print(f"Failed to load icon: {e}")
            pass

        self.root.title("MapMaestro")
        self.root.geometry("1200x800")
        self.root.state("zoomed")
        self.root.minsize(1200, 800)

        try:
            self.root.attributes("-zoomed", True)
        except:
            pass

        # Fallback: manual screen size
        self.root.update_idletasks()  # Ensure window is ready
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.geometry(f"{screen_width-100}x{screen_height-100}+0+0")

        self.root.minsize(1200, 800)

        # CRITICAL: Configure the root grid properly
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        # Set background color
        self.root.configure(bg="#f8f9fa")

    def setup_styles(self):
        """Create elegant styles for the application"""
        self.style = ttk.Style()

        # Configure modern theme
        self.style.theme_use("clam")

        # Color scheme
        self.colors = {
            "primary": "#667eea",
            "primary_dark": "#5a67d8",
            "secondary": "#764ba2",
            "success": "#00b894",
            "danger": "#e74c3c",
            "light": "#f8f9fa",
            "white": "#ffffff",
            "dark": "#2c3e50",
            "muted": "#7f8c8d",
            "border": "#e1e8ed",
        }

        # Fonts
        self.fonts = {
            "title": tkFont.Font(family="Segoe UI", size=28, weight="bold"),
            "subtitle": tkFont.Font(family="Segoe UI", size=18),
            "header": tkFont.Font(family="Segoe UI", size=24, weight="bold"),
            "body": tkFont.Font(family="Segoe UI", size=14),
            "button": tkFont.Font(family="Segoe UI", size=14, weight="bold"),
            "system_label": tkFont.Font(family="Segoe UI", size=20, weight="bold"),
            "combobox": tkFont.Font(family="Segoe UI", size=16), 
            "status": tkFont.Font(family="Segoe UI", size=15),
        }

        # Configure styles
        self.configure_styles()
        
        # Configure large radio button style for PII tab
        self.style.configure(
            "Large.TRadiobutton",
            font=("Segoe UI", 15, "bold")
        )

    def configure_styles(self):
        """Configure all the custom styles"""
        # Main frame style
        self.style.configure(
            "Main.TFrame", background=self.colors["light"], relief="flat"
        )

        # Header frame style
        self.style.configure(
            "Header.TFrame",
            background=self.colors["white"],
            relief="flat",
            borderwidth=0,
        )

        # Card frame style
        self.style.configure(
            "Card.TFrame", background=self.colors["white"], relief="flat", borderwidth=1
        )

        # Filter frame style
        self.style.configure(
            "Filter.TFrame",
            background=self.colors["white"],
            relief="flat",
            borderwidth=1,
        )

        # Title label style
        self.style.configure(
            "Title.TLabel",
            background=self.colors["white"],
            foreground=self.colors["dark"],
            font=self.fonts["title"],
        )

        # Subtitle label style
        self.style.configure(
            "Subtitle.TLabel",
            background=self.colors["white"],
            foreground=self.colors["muted"],
            font=self.fonts["subtitle"],
        )

        # Header label style
        self.style.configure(
            "Header.TLabel",
            background=self.colors["white"],
            foreground=self.colors["dark"],
            font=self.fonts["header"],
        )

        # Filter label style
        self.style.configure(
            "FilterLabel.TLabel",
            background=self.colors["white"],
            foreground=self.colors["dark"],
            font=self.fonts["body"],
        )

        # Primary button style
        self.style.configure(
            "Primary.TButton",
            font=self.fonts["button"],
            borderwidth=0,
            focuscolor="none",
        )

        self.style.map(
            "Primary.TButton",
            background=[
                ("active", self.colors["primary_dark"]),
                ("!active", self.colors["primary"]),
            ],
            foreground=[("active", "white"), ("!active", "white")],
        )

        # Success button style
        self.style.configure(
            "Success.TButton",
            font=self.fonts["button"],
            borderwidth=0,
            focuscolor="none",
        )

        self.style.map(
            "Success.TButton",
            background=[("active", "#00a085"), ("!active", self.colors["success"])],
            foreground=[("active", "white"), ("!active", "white")],
        )
        self.style.configure(
            "SystemLabel.TLabel",
            background=self.colors["white"],
            foreground=self.colors["dark"],
            font=self.fonts["system_label"],
        )
        # Danger button style
        self.style.configure(
            "Danger.TButton",
            font=self.fonts["button"],
            borderwidth=0,
            focuscolor="none",
        )

        self.style.map(
            "Danger.TButton",
            background=[("active", "#c0392b"), ("!active", self.colors["danger"])],
            foreground=[("active", "white"), ("!active", "white")],
        )

        # Combobox style
        self.style.configure(
            "Modern.TCombobox",
            fieldbackground=self.colors["white"],
            borderwidth=1,
            relief="solid",
        )

        # Entry style
        self.style.configure(
            "Modern.TEntry",
            fieldbackground=self.colors["white"],
            borderwidth=1,
            relief="solid",
        )
        
        self.style.configure(
            "Custom.TRadiobutton",
            font=("Segoe UI", 15),
            background=self.colors["white"],
        )

        self.style.map(
            "Custom.TRadiobutton",
            background=[("active", self.colors["white"])],
        )

    def load_data(self):
        """Load data from database with improved null handling"""
        try:
            conn = get_connection()
            rows = fetch_table_a(conn)
            conn.close()

            # Ensure we have valid rows
            self.rows_a = []
            for row in rows:
                if row is not None and isinstance(row, dict):
                    self.rows_a.append(row)

            # Extract unique values for each filter column with null safety
            self.all_values = {}
            for col in FILTER_COLUMNS:
                values = set()
                for r in self.rows_a:
                    if r is not None and isinstance(r, dict):
                        val = r.get(col)
                        if val is not None:
                            values.add(val)
                self.all_values[col] = sorted(values)

        except Exception as e:
            logger.error(f"Error loading data: {e}", exc_info=True)
            messagebox.showerror("Database Error", f"Failed to load data: {str(e)}")
            self.root.destroy()

    def create_ui(self):
        """Create the main UI with header and tabs below"""
        # Main container
        main_container = ttk.Frame(self.root, style="Main.TFrame")
        main_container.pack(fill="both", expand=True)

        # Create header that stays at top (shared across all tabs)
        self.create_main_header(main_container)

        # Create notebook for tabs below header
        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # Tab 1: Generate Mappings 
        self.main_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.main_tab, text="   Generate Mappings  ")
        self.create_main_tab_content(self.main_tab)

        # Tab 2: Upload Excel 
        self.upload_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.upload_tab, text="   Upload Excel  ")
        self.create_upload_tab_content(self.upload_tab)

        # Tab 3: Detect PII Fields 
        self.pii_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.pii_tab, text="   Detect PII Fields  ")
        self.create_pii_tab_content(self.pii_tab)

    def create_main_header(self, parent):
        """Create the main header that appears above tabs"""
        header_container = ttk.Frame(parent, style="Header.TFrame")
        header_container.pack(fill="x", padx=20, pady=20)

        header_frame = ttk.Frame(header_container, style="Header.TFrame", padding=20)
        header_frame.pack(fill="x")
        header_frame.grid_columnconfigure(1, weight=1)

        # Logo
        logo_frame = tk.Frame(
            header_frame, width=100, height=100, bg=self.colors["white"]
        )
        logo_frame.grid(row=0, column=0, padx=(0, 16), rowspan=2)
        logo_frame.grid_propagate(False)

        try:
            logo_paths = ["logo_resized.ico"]
            logo_loaded = False
            for logo_path in logo_paths:
                if os.path.exists(logo_path):
                    try:
                        from PIL import Image, ImageTk

                        image = Image.open(logo_path)
                        image = image.resize((100, 100), Image.Resampling.LANCZOS)
                        photo = ImageTk.PhotoImage(image)
                        logo_label = tk.Label(
                            logo_frame, image=photo, bg=self.colors["white"]
                        )
                        logo_label.image = photo
                        logo_label.place(relx=0.5, rely=0.5, anchor="center")
                        logo_loaded = True
                        break
                    except:
                        pass

            if not logo_loaded:
                logo_label = tk.Label(
                    logo_frame,
                    text="M",
                    fg="white",
                    bg=self.colors["primary"],
                    font=("Segoe UI", 40, "bold"),
                )
                logo_label.place(relx=0.5, rely=0.5, anchor="center")
        except:
            logo_label = tk.Label(
                logo_frame,
                text="M",
                fg="white",
                bg=self.colors["primary"],
                font=("Segoe UI", 40, "bold"),
            )
            logo_label.place(relx=0.5, rely=0.5, anchor="center")

        # Title and subtitle
        title_label = ttk.Label(header_frame, text="MapMaestro", style="Title.TLabel")
        title_label.grid(row=0, column=1, sticky="w")

        subtitle_label = ttk.Label(
            header_frame,
            text="Orchestrate Your Maps with Precision",
            style="Subtitle.TLabel",
        )
        subtitle_label.grid(row=1, column=1, sticky="w")

        # Separator
        separator = ttk.Separator(parent, orient="horizontal")
        separator.pack(fill="x", padx=20, pady=(0, 10))

    def create_main_tab_content(self, parent):
        """Create content for Generate Mappings tab"""
        # Create scrollable content
        self.canvas = tk.Canvas(parent, bg=self.colors["light"], highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(
            parent, orient="vertical", command=self.canvas.yview
        )
        self.scrollable_frame = ttk.Frame(self.canvas, style="Main.TFrame")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")),
        )

        self.canvas_frame = self.canvas.create_window(
            (0, 0), window=self.scrollable_frame, anchor="nw"
        )
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        self.scrollbar.grid(row=0, column=1, sticky="nsew", pady=0)

        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_columnconfigure(1, weight=0)

        main_content = ttk.Frame(self.scrollable_frame, style="Main.TFrame", padding=20)
        main_content.grid(row=0, column=0, sticky="nsew")
        main_content.grid_columnconfigure(0, weight=1)

        self.scrollable_frame.grid_columnconfigure(0, weight=1)
        self.scrollable_frame.grid_rowconfigure(0, weight=1)

        self.canvas.bind("<Configure>", self.on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self.on_mousewheel)

        # Create existing sections (without header as it's now above tabs)
        self.create_filters_section(main_content)
        self.create_action_section(main_content)

        self.root.after(100, self.update_scroll_region)

    def create_upload_tab_content(self, parent):
        """Create content for Upload Excel tab"""
        # Create scrollable content (same structure as main tab)
        self.upload_canvas = tk.Canvas(
            parent, bg=self.colors["light"], highlightthickness=0
        )
        self.upload_scrollbar = ttk.Scrollbar(
            parent, orient="vertical", command=self.upload_canvas.yview
        )
        self.upload_scrollable_frame = ttk.Frame(
            self.upload_canvas, style="Main.TFrame"
        )

        self.upload_scrollable_frame.bind(
            "<Configure>",
            lambda e: self.upload_canvas.configure(
                scrollregion=self.upload_canvas.bbox("all")
            ),
        )

        self.upload_canvas_frame = self.upload_canvas.create_window(
            (0, 0), window=self.upload_scrollable_frame, anchor="nw"
        )
        self.upload_canvas.configure(yscrollcommand=self.upload_scrollbar.set)

        self.upload_canvas.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        self.upload_scrollbar.grid(row=0, column=1, sticky="nsew", pady=0)

        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_columnconfigure(1, weight=0)

        main_content = ttk.Frame(
            self.upload_scrollable_frame, style="Main.TFrame", padding=20
        )
        main_content.grid(row=0, column=0, sticky="nsew")
        main_content.grid_columnconfigure(0, weight=1)

        self.upload_scrollable_frame.grid_columnconfigure(0, weight=1)
        self.upload_scrollable_frame.grid_rowconfigure(0, weight=1)

        self.upload_canvas.bind("<Configure>", self.on_upload_canvas_configure)
        self.upload_canvas.bind_all("<MouseWheel>", self.on_upload_mousewheel)

        # Create sections
        self.create_upload_file_section(main_content)
        self.create_upload_info_and_filters_section(main_content)
        self.create_upload_action_section_full(main_content)

        self.root.after(100, self.update_upload_scroll_region)

    def on_upload_canvas_configure(self, event):
        """Handle upload canvas resize"""
        canvas_width = event.width
        self.upload_canvas.itemconfig(self.upload_canvas_frame, width=canvas_width)

    def update_upload_scroll_region(self):
        """Update the upload scroll region"""
        self.upload_canvas.configure(scrollregion=self.upload_canvas.bbox("all"))

    def on_upload_mousewheel(self, event):
        """Handle mouse wheel scrolling for upload tab"""
        # Check if we're on the upload tab
        if self.notebook.select() == str(self.upload_tab):
            # Check if the canvas is scrollable
            if (
                self.upload_canvas.winfo_reqheight()
                < self.upload_scrollable_frame.winfo_reqheight()
            ):
                # Windows and MacOS
                if hasattr(event, "delta"):
                    delta = event.delta
                    self.upload_canvas.yview_scroll(int(-1 * (delta / 120)), "units")
                # Linux
                elif event.num == 4:
                    self.upload_canvas.yview_scroll(-1, "units")
                elif event.num == 5:
                    self.upload_canvas.yview_scroll(1, "units")

    def create_upload_file_section(self, parent):
        """Create the file selection section - full width"""
        # Section header
        section_header = ttk.Label(parent, text="File Selection", style="Header.TLabel")
        section_header.grid(row=0, column=0, sticky="w", pady=(0, 15))

        # File selection container
        file_container = ttk.Frame(parent, style="Card.TFrame", padding=20)
        file_container.grid(row=1, column=0, sticky="ew", pady=(0, 20))
        file_container.grid_columnconfigure(0, weight=1)

        # Simplified instruction - single line
        instructions = ttk.Label(
            file_container,
            text="Upload an Excel File containing the Target Field Mappings.",
            style="FilterLabel.TLabel",
            foreground=self.colors["muted"],
        )
        instructions.grid(row=0, column=0, sticky="w", pady=(0, 15))

        # Excel File subheader
        file_subheader = ttk.Label(
            file_container,
            text="Excel File",
            style="FilterLabel.TLabel",
            font=("Segoe UI", 15, "bold"),
        )
        file_subheader.grid(row=1, column=0, sticky="w", pady=(0, 8))

        # File input frame (browse button)
        file_input_frame = ttk.Frame(file_container)
        file_input_frame.grid(row=2, column=0, sticky="ew", pady=(0, 15))
        file_input_frame.grid_columnconfigure(0, weight=1)

        self.upload_file_path = ttk.Entry(
            file_input_frame, style="Modern.TEntry", font=self.fonts["body"]
        )
        self.upload_file_path.insert(0, "No file selected")
        self.upload_file_path.config(state="readonly")
        self.upload_file_path.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        browse_btn = ttk.Button(
            file_input_frame,
            text="Browse...",
            command=self.browse_upload_file,
            style="Primary.TButton",
        )
        browse_btn.grid(row=0, column=1)

        # Validation status
        self.validation_label = ttk.Label(
            file_container, text="", 
            foreground=self.colors["muted"],
            font=self.fonts["status"],
        )
        self.validation_label.grid(row=3, column=0, sticky="w", pady=(0, 0))

    def create_upload_info_and_filters_section(self, parent):
        """Create file info and filter selection side by side"""
        # Container for side-by-side layout
        container = ttk.Frame(parent, style="Main.TFrame")
        container.grid(row=2, column=0, sticky="nsew", pady=(0, 20))
        container.grid_columnconfigure(0, weight=1)
        container.grid_columnconfigure(1, weight=1)

        # LEFT SIDE: File Information
        info_section = ttk.Frame(container, style="Card.TFrame", padding=20)
        info_section.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        info_section.grid_columnconfigure(0, weight=1)
        info_section.grid_rowconfigure(1, weight=1)

        info_header = ttk.Label(
            info_section, text="File Information", style="Header.TLabel"
        )
        info_header.grid(row=0, column=0, sticky="w", pady=(0, 15))

        # File info text box
        info_text_frame = ttk.Frame(info_section)
        info_text_frame.grid(row=1, column=0, sticky="nsew")
        info_text_frame.grid_columnconfigure(0, weight=1)
        info_text_frame.grid_rowconfigure(0, weight=1)

        info_scrollbar = ttk.Scrollbar(info_text_frame)
        info_scrollbar.grid(row=0, column=1, sticky="ns")

        self.file_info_text = tk.Text(
            info_text_frame,
            height=10,
            font=self.fonts["body"],
            wrap="word",
            bg=self.colors["light"],
            relief="flat",
            borderwidth=0,
            yscrollcommand=info_scrollbar.set,
        )
        self.file_info_text.grid(row=0, column=0, sticky="nsew")
        info_scrollbar.config(command=self.file_info_text.yview)

        self.file_info_text.insert(
            "1.0", "No file loaded yet.\n\nPlease select an Excel file to begin."
        )
        self.file_info_text.config(state="disabled")

        # RIGHT SIDE: Filter Selection
        filter_section = ttk.Frame(container, style="Card.TFrame", padding=20)
        filter_section.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        filter_section.grid_columnconfigure(0, weight=1)

        filter_header = ttk.Label(
            filter_section, text="Filter Selection", style="Header.TLabel"
        )
        filter_header.grid(row=0, column=0, sticky="w", pady=(0, 15))

        # Legacy System
        self.create_upload_filter_group(
            filter_section, "legacysystem", "Legacy System", 1
        )

        # Target System
        self.create_upload_filter_group(
            filter_section, "targetsystem", "Target System", 2
        )

        # Project (Optional)
        self.create_upload_filter_group(
            filter_section, "projectname", "Project (Optional)", 3
        )

        # Transitivity option frame (initially hidden)
        self.upload_transitivity_frame = ttk.Frame(filter_section)
        self.upload_transitivity_frame.grid(row=4, column=0, sticky="ew", pady=(15, 0))
        self.upload_transitivity_frame.grid_columnconfigure(1, weight=1)

        self.upload_enable_transitivity = tk.BooleanVar(value=False)
        self.upload_transitivity_check = ttk.Checkbutton(
            self.upload_transitivity_frame,
            text="Apply transitivity mapping",
            variable=self.upload_enable_transitivity,
            command=self.on_upload_transitivity_toggle,
        )
        self.upload_transitivity_check.pack(side="left", padx=(0, 10))

        self.upload_transitivity_target = ttk.Combobox(
            self.upload_transitivity_frame,
            values=[],
            state="disabled",
            style="Modern.TCombobox",
            font=self.fonts["combobox"],
            width=20,
        )
        self.upload_transitivity_target.pack(side="left")

        # Initially hide transitivity
        self.upload_transitivity_frame.grid_remove()
        # Setup bindings after creating all listboxes
        self.setup_upload_listbox_bindings()

    def create_upload_filter_group(self, parent, column, label_text, grid_row):
        """Create a filter group with label and listbox for upload tab"""
        # Filter frame
        filter_frame = ttk.Frame(parent, style="Filter.TFrame", padding=15)
        filter_frame.grid(row=grid_row, column=0, sticky="ew", pady=(0, 15))
        filter_frame.grid_columnconfigure(0, weight=1)
        filter_frame.grid_rowconfigure(1, weight=1)

        # Label
        label = ttk.Label(
            filter_frame,
            text=label_text,
            style="FilterLabel.TLabel",
            font=("Segoe UI", 15, "bold"),
        )
        label.grid(row=0, column=0, sticky="w", pady=(0, 8))

        # Listbox with scrollbar
        listbox_frame = tk.Frame(filter_frame, bg=self.colors["white"])
        listbox_frame.grid(row=1, column=0, sticky="nsew")
        listbox_frame.grid_columnconfigure(0, weight=1)
        listbox_frame.grid_rowconfigure(0, weight=1)

        # Scrollbar
        scrollbar = ttk.Scrollbar(listbox_frame)
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Listbox
        listbox = tk.Listbox(
            listbox_frame,
            selectmode=tk.MULTIPLE,
            exportselection=False,
            height=6,
            bg=self.colors["white"],
            fg=self.colors["dark"],
            font=self.fonts["body"],
            selectbackground=self.colors["primary"],
            selectforeground="white",
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightcolor=self.colors["primary"],
            yscrollcommand=scrollbar.set,
        )
        listbox.grid(row=0, column=0, sticky="nsew")
        scrollbar.config(command=listbox.yview)

        # Populate listbox
        listbox.insert(tk.END, "All")
        for value in self.all_values.get(column, []):
            listbox.insert(tk.END, value)

        # Store reference with upload prefix
        if not hasattr(self, "upload_listboxes"):
            self.upload_listboxes = {}
        self.upload_listboxes[column] = listbox

    def setup_upload_listbox_bindings(self):
        """Setup event bindings for upload listboxes"""
        if hasattr(self, "upload_listboxes"):
            legacy_listbox = self.upload_listboxes.get("legacysystem")
            if legacy_listbox:
                legacy_listbox.bind("<<ListboxSelect>>", self.on_upload_legacy_change)

    def create_upload_action_section_full(self, parent):
        """Create action section for upload tab"""
        # Section header
        section_header = ttk.Label(parent, text="Export Options", style="Header.TLabel")
        section_header.grid(row=3, column=0, sticky="w", pady=(0, 15))

        # Action container
        action_container = ttk.Frame(parent, style="Card.TFrame", padding=20)
        action_container.grid(row=4, column=0, sticky="ew")
        action_container.grid_columnconfigure(1, weight=1)

        # Action dropdown
        action_label = ttk.Label(
            action_container, text="Action", style="FilterLabel.TLabel"
        )
        action_label.grid(row=0, column=0, sticky="w", padx=(0, 15), pady=(0, 15))

        self.upload_action_cb = ttk.Combobox(
            action_container,
            values=[
                "Standard Excel",
                "Generate mapping spec excels",
                "Export into Data Lineage App",
            ],
            state="readonly",
            style="Modern.TCombobox",
            font=self.fonts["combobox"],
        )
        self.upload_action_cb.current(0)
        self.upload_action_cb.grid(
            row=0, column=1, sticky="ew", padx=(0, 15), pady=(0, 15)
        )

        # Output path
        path_label = ttk.Label(
            action_container, text="Output Path", style="FilterLabel.TLabel"
        )
        path_label.grid(row=1, column=0, sticky="w", padx=(0, 15), pady=(0, 15))

        path_frame = ttk.Frame(action_container)
        path_frame.grid(row=1, column=1, sticky="ew", padx=(0, 15), pady=(0, 15))
        path_frame.grid_columnconfigure(0, weight=1)

        self.upload_output_path = ttk.Entry(
            path_frame, style="Modern.TEntry", font=self.fonts["body"]
        )
        self.upload_output_path.insert(0, "uploaded_mapping_output.xlsx")
        self.upload_output_path.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        browse_btn = ttk.Button(
            path_frame,
            text="Browse...",
            command=self.browse_upload_output,
            style="Primary.TButton",
        )
        browse_btn.grid(row=0, column=1)

        # Action buttons
        button_frame = ttk.Frame(action_container)
        button_frame.grid(row=2, column=0, columnspan=2, pady=(20, 0))

        self.upload_process_btn = ttk.Button(
            button_frame,
            text="Run",
            command=self.process_upload,
            style="Success.TButton",
            width=20,
            state="disabled",
        )
        self.upload_process_btn.pack(side="left", padx=(0, 10))

        clear_btn = ttk.Button(
            button_frame,
            text="Clear",
            command=self.clear_upload,
            style="Danger.TButton",
            width=20,
        )
        clear_btn.pack(side="left")

        # Status label
        self.upload_status_label = ttk.Label(
            action_container,
            text="",
            foreground=self.colors["success"],
            font=self.fonts["status"],
        )
        self.upload_status_label.grid(row=3, column=0, columnspan=2, pady=(15, 0))

    def browse_upload_file(self):
        """Handle Excel file selection for upload"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm")],
        )

        if file_path:
            self.uploaded_file_path = file_path
            self.upload_file_path.config(state="normal")
            self.upload_file_path.delete(0, tk.END)
            self.upload_file_path.insert(0, os.path.basename(file_path))
            self.upload_file_path.config(state="readonly")

            # Parse and validate
            try:
                self.uploaded_data = parse_uploaded_excel(file_path)
                self.validation_results = validate_excel_structure(self.uploaded_data)

                # Update file info
                self.file_info_text.config(state="normal")
                self.file_info_text.delete("1.0", tk.END)

                info_lines = []
                info_lines.append(f"File: {os.path.basename(file_path)}")
                info_lines.append(f"Sheets: {len(self.uploaded_data)}")

                for sheet_name, results in self.validation_results.items():
                    status = "Valid" if results["valid"] else "Invalid"
                    info_lines.append(f"\n{status} {sheet_name}:")
                    info_lines.append(f"  Rows: {results['total_rows']}")
                    if results["missing_required"]:
                        info_lines.append(
                            f"  Missing: {', '.join(results['missing_required'][:3])}"
                        )
                    if results["valid"]:
                        info_lines.append(f"  Status: All required columns present")
                    info_lines.append("")

                self.file_info_text.insert("1.0", "\n".join(info_lines))
                self.file_info_text.config(state="disabled")

                # Enable process button if valid
                valid_sheets = sum(
                    1 for v in self.validation_results.values() if v["valid"]
                )
                if valid_sheets > 0:
                    self.upload_process_btn.config(state="normal")
                    self.validation_label.config(
                        text=f" {valid_sheets} valid sheet(s) ready to process",
                        foreground=self.colors["success"],
                    )
                else:
                    self.validation_label.config(
                        text="No valid sheets found", foreground=self.colors["danger"]
                    )

            except Exception as e:
                self.validation_label.config(
                    text=f"Error: {str(e)}", foreground=self.colors["danger"]
                )

    def browse_upload_output(self):
        """Browse for output file location"""
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=self.upload_output_path.get(),
        )
        if path:
            self.upload_output_path.delete(0, tk.END)
            self.upload_output_path.insert(0, path)

    def on_upload_legacy_change(self, event=None):
        """Check transitivity options when legacy system changes - works with listbox"""
        if not hasattr(self, "upload_listboxes"):
            return

        legacy_listbox = self.upload_listboxes.get("legacysystem")
        if not legacy_listbox:
            return

        selected_indices = legacy_listbox.curselection()
        if not selected_indices:
            self.upload_transitivity_frame.grid_remove()
            return

        selected_values = [legacy_listbox.get(i) for i in selected_indices]

        # Check if exactly one legacy system is selected (excluding 'All')
        if "All" in selected_values or len(selected_values) != 1:
            self.upload_transitivity_frame.grid_remove()
            return

        legacy = selected_values[0]

        # Check for transitivity options
        conn = get_connection()
        try:
            available = check_transitivity_availability(conn, legacy, None)
            if available:
                self.upload_transitivity_frame.grid()
                self.upload_transitivity_check.pack(side="left", padx=(0, 10))
                self.upload_transitivity_target.pack(side="left")
                self.upload_transitivity_target["values"] = available
                if available:
                    self.upload_transitivity_target.current(0)
            else:
                self.upload_transitivity_frame.grid_remove()
        finally:
            conn.close()

    def on_upload_transitivity_toggle(self):
        """Handle transitivity checkbox toggle in upload tab"""
        if self.upload_enable_transitivity.get():
            self.upload_transitivity_target.config(state="readonly")
        else:
            self.upload_transitivity_target.config(state="disabled")

    def process_upload(self):
        """Process the uploaded Excel file"""
        if not self.uploaded_file_path:
            messagebox.showerror("Error", "No file uploaded")
            return

        try:
            self.upload_status_label.config(
                text="Processing.", foreground=self.colors["primary"]
            )
            self.root.update()

            # Get selections from listboxes
            if hasattr(self, "upload_listboxes"):
                # Legacy System
                legacy_listbox = self.upload_listboxes.get("legacysystem")
                legacy_selected = (
                    [legacy_listbox.get(i) for i in legacy_listbox.curselection()]
                    if legacy_listbox
                    else []
                )
                legacy_system = (
                    None
                    if not legacy_selected or "All" in legacy_selected
                    else legacy_selected[0] if len(legacy_selected) == 1 else None
                )

                # Target System
                target_listbox = self.upload_listboxes.get("targetsystem")
                target_selected = (
                    [target_listbox.get(i) for i in target_listbox.curselection()]
                    if target_listbox
                    else []
                )
                target_system = (
                    None
                    if not target_selected or "All" in target_selected
                    else target_selected[0] if len(target_selected) == 1 else None
                )

                # Project
                project_listbox = self.upload_listboxes.get("projectname")
                project_selected = (
                    [project_listbox.get(i) for i in project_listbox.curselection()]
                    if project_listbox
                    else []
                )
                project = (
                    None
                    if not project_selected or "All" in project_selected
                    else project_selected[0] if len(project_selected) == 1 else None
                )
            else:
                legacy_system = None
                target_system = None
                project = None
            action = self.upload_action_cb.get()

            if not project or project == "All":
                project_name = None
            else:
                project_name = project

            if not target_system or target_system == "All":
                messagebox.showerror("Error", "Please select a target system")
                return

            if project == "All":
                project = None

            # BRANCH 1: Standard Excel / Generate mapping spec
            if action in ("Standard Excel", "Generate mapping spec excels"):
                if not legacy_system or legacy_system == "All":
                    messagebox.showerror(
                        "Error", "Please select a specific Legacy System."
                    )
                    return

                export_type = "standard" if action == "Standard Excel" else "summarized"
                output_path = self.upload_output_path.get().strip() or None

                success, final_out, record_count = process_mapping_request_from_file(
                    file_path=self.uploaded_file_path,
                    legacy_system=legacy_system,
                    target_system=target_system,
                    export_type=export_type,
                    output_path=output_path,
                    project_name=project_name,
                )

                if success:
                    msg = (
                        f"Mapping completed.\n"
                        f"Records: {record_count}\n"
                        f"Output: {final_out}"
                    )
                    self.upload_status_label.config(
                        text=msg, foreground=self.colors["success"]
                    )
                    messagebox.showinfo("Mapping Complete", msg)
                else:
                    msg = "No mappings found or process failed.\n\n"
                    self.upload_status_label.config(
                        text=msg, foreground=self.colors["danger"]
                    )
                    messagebox.showerror("Lookup Error", msg)

                return

            # BRANCH 2: Export into Data Lineage App
            if not self.uploaded_data:
                messagebox.showerror(
                    "Error",
                    "No parsed Excel data available. Please re-select the file.",
                )
                return

            if legacy_system == "All":
                legacy_system = None
            if project == "All":
                project = None

            all_processed_rows = []
            logger.info(
                f"Processing {len(self.uploaded_data)} sheet(s) from uploaded file"
            )

            for sheet_name, df in self.uploaded_data.items():
                logger.info(
                    f"Processing sheet: {sheet_name}, valid: {self.validation_results[sheet_name]['valid']}"
                )

                if self.validation_results[sheet_name]["valid"]:
                    logger.info(
                        f"Sheet '{sheet_name}' has {len(df)} rows before transformation"
                    )

                    transformed_df = transform_excel_to_mapping_format(
                        df, target_system, legacy_system
                    )

                    logger.info(
                        f"Sheet '{sheet_name}' has {len(transformed_df)} rows after transformation"
                    )

                    if project:
                        transformed_df["projectname"] = project

                    records = transformed_df.to_dict("records")
                    logger.info(
                        f"Converting {len(records)} records to dict for sheet '{sheet_name}'"
                    )

                    all_processed_rows.extend(records)
                    logger.info(
                        f"Total processed rows so far: {len(all_processed_rows)}"
                    )

            logger.info(
                f"FINAL: all_processed_rows has {len(all_processed_rows)} total rows before database insert"
            )

            # Apply transitivity if enabled
            if (
                self.upload_enable_transitivity.get()
                and self.upload_transitivity_target.get()
            ):
                # Apply transitivity logic here if/when needed
                pass

            if action == "Export into Data Lineage App":
                logger.info(f"About to insert {len(all_processed_rows)} rows")

                if not all_processed_rows:
                    logger.error(
                        "ERROR: all_processed_rows is EMPTY! Nothing to insert."
                    )
                    self.upload_status_label.config(
                        text="Error: No data to export",
                        foreground=self.colors["danger"],
                    )
                    return
                conn_output = get_outputconnection()
                try:
                    store_output_to_db(
                        all_processed_rows,
                        conn_output,
                        table_name=postgres_outputtable,
                        column_mapping=COLUMN_MAPPING,
                        columns_to_insert=desired,
                    )
                    self.upload_status_label.config(
                        text="Exported successfully into Data Lineage App",
                        foreground=self.colors["success"],
                    )
                finally:
                    conn_output.close()

        except Exception as e:
            self.upload_status_label.config(
                text=f"Error: {str(e)}", foreground=self.colors["danger"]
            )
            logger.error("[UPLOAD] Error in process_upload", exc_info=True)
            messagebox.showerror("Processing Error", str(e))

    def clear_upload(self):
        """Clear upload data and reset UI"""
        self.uploaded_file_path = None
        self.uploaded_data = None
        self.validation_results = None

        self.upload_file_path.config(state="normal")
        self.upload_file_path.delete(0, tk.END)
        self.upload_file_path.insert(0, "No file selected")
        self.upload_file_path.config(state="readonly")

        self.file_info_text.config(state="normal")
        self.file_info_text.delete("1.0", tk.END)
        self.file_info_text.insert("1.0", "No file loaded yet.")
        self.file_info_text.config(state="disabled")

        self.validation_label.config(text="")
        self.upload_status_label.config(text="")
        self.upload_process_btn.config(state="disabled")

        # Reset listbox selections
        if hasattr(self, "upload_listboxes"):
            for listbox in self.upload_listboxes.values():
                listbox.selection_clear(0, tk.END)
        self.upload_action_cb.current(0)

    def create_header(self, parent):
        """Create the header section with logo and title"""
        header_frame = ttk.Frame(parent, style="Header.TFrame", padding=20)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        header_frame.grid_columnconfigure(1, weight=1)

        # Logo placeholder
        logo_frame = tk.Frame(header_frame, width=100, height=100)
        logo_frame.grid(row=0, column=0, padx=(0, 16), rowspan=2)
        logo_frame.grid_propagate(False)

        try:
            logo_paths = ["logo_resized.ico"]

            logo_loaded = False
            for logo_path in logo_paths:
                if os.path.exists(logo_path):
                    try:
                        from PIL import Image, ImageTk

                        image = Image.open(logo_path)
                        image = image.resize((100, 100), Image.Resampling.LANCZOS)
                        photo = ImageTk.PhotoImage(image)

                        logo_label = tk.Label(logo_frame, image=photo)
                        logo_label.image = photo  # Keep a reference
                        logo_label.place(relx=0.5, rely=0.5, anchor="center")
                        logo_loaded = True
                        print(f"Logo loaded: {logo_path}")
                        break
                    except ImportError:
                        print("PIL not available, using text logo")
                        break
                    except Exception as e:
                        print(f"Failed to load {logo_path}: {e}")
                        continue

            # Fallback to text logo if no image found
            if not logo_loaded:
                logo_label = tk.Label(
                    logo_frame,
                    text="M",
                    fg="white",
                    bg=self.colors["primary"],
                    font=("Segoe UI", 20, "bold"),
                )
                logo_label.place(relx=0.5, rely=0.5, anchor="center")

        except Exception as e:
            print(f"Logo loading error: {e}")
            # Ultimate fallback - text logo
            logo_label = tk.Label(
                logo_frame,
                text="M",
                fg="white",
                bg=self.colors["primary"],
                font=("Segoe UI", 20, "bold"),
            )
            logo_label.place(relx=0.5, rely=0.5, anchor="center")

        # Title
        title_label = ttk.Label(header_frame, text="MapMaestro", style="Title.TLabel")
        title_label.grid(row=0, column=1, sticky="w")

        # Subtitle
        subtitle_label = ttk.Label(
            header_frame,
            text="Orchestrate Your Maps with Precision",
            style="Subtitle.TLabel",
        )
        subtitle_label.grid(row=1, column=1, sticky="w")

        # Add separator
        separator = ttk.Separator(parent, orient="horizontal")
        separator.grid(row=1, column=0, sticky="ew", pady=(0, 20))

    def create_filters_section(self, parent):
        """Create the filters section"""
        # Filters header
        filters_header = ttk.Label(parent, text="🔍 Filters", style="Header.TLabel")
        filters_header.grid(row=2, column=0, sticky="w", pady=(0, 15))

        # Filters container - MAKE IT EXPAND
        filters_container = ttk.Frame(parent, style="Card.TFrame", padding=20)
        filters_container.grid(
            row=3, column=0, sticky="nsew", pady=(0, 20)
        )  # Changed to nsew

        # CRITICAL: Configure the filters container to expand
        filters_container.grid_columnconfigure(0, weight=1)
        filters_container.grid_columnconfigure(1, weight=1)
        filters_container.grid_rowconfigure(0, weight=0)  # Systems row
        filters_container.grid_rowconfigure(1, weight=1)  # First filter row
        filters_container.grid_rowconfigure(2, weight=1)  # Second filter row

        # CRITICAL: Configure parent to allow expansion
        parent.grid_rowconfigure(3, weight=1)  # Make the filters section expandable

        # Systems row (Legacy and Target side by side)
        systems_frame = ttk.Frame(filters_container)
        systems_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 20))
        systems_frame.grid_columnconfigure(0, weight=1)
        systems_frame.grid_columnconfigure(1, weight=1)

        # Legacy System
        self.create_filter_group(
            systems_frame, "legacysystem", "🏛️ Legacy System", 0, 0, is_system=True
        )

        # Target System
        self.create_filter_group(
            systems_frame, "targetsystem", "🎯 Target System", 0, 1, is_system=True
        )

        # Remaining filters in grid
        filter_positions = [
            ("projectname", "📁 Project Name", 1, 0),
            ("domainname", "🌐 Domain Name", 1, 1),
            ("subdomainname", "🔗 Subdomain Name", 2, 0),
            ("filename", "📄 Filename", 2, 1),
        ]

        for col, label, row, column in filter_positions:
            self.create_filter_group(
                filters_container, col, label, row, column, is_system=True
            )

    def create_filter_group(
        self, parent, column, label_text, grid_row, grid_col, is_system=True
    ):
        """Create a filter group with label and listbox"""
        # Filter frame - MAKE IT EXPAND
        if is_system:
            filter_frame = ttk.Frame(parent, style="Filter.TFrame", padding=15)
            filter_frame.grid(
                row=grid_row, column=grid_col, sticky="nsew", pady=(0, 15)
            )
        else:
            filter_frame = ttk.Frame(parent, style="Filter.TFrame", padding=15)
            filter_frame.grid(
                row=grid_row,
                column=grid_col,
                sticky="nsew",
                padx=(0, 10) if grid_col == 0 else (10, 0),
                pady=(0, 15),
            )

        # CRITICAL: Make filter frame expand
        filter_frame.grid_columnconfigure(0, weight=1)
        filter_frame.grid_rowconfigure(1, weight=1)  # Row with listbox should expand

        # CRITICAL: Configure parent grid weights
        parent.grid_rowconfigure(grid_row, weight=1 if not is_system else 0)
        parent.grid_columnconfigure(grid_col, weight=1)

        # Label - Use different style for system labels
        label_style = "SystemLabel.TLabel" if is_system else "FilterLabel.TLabel"
        label = ttk.Label(filter_frame, text=label_text, style=label_style)
        label.grid(row=0, column=0, sticky="w", pady=(0, 8))  # Changed to grid

        # Listbox with scrollbar - MAKE IT EXPAND
        listbox_frame = tk.Frame(filter_frame, bg=self.colors["white"])
        listbox_frame.grid(row=1, column=0, sticky="nsew")  # Changed to grid and nsew
        listbox_frame.grid_columnconfigure(0, weight=1)
        listbox_frame.grid_rowconfigure(0, weight=1)

        # Scrollbar
        scrollbar = ttk.Scrollbar(listbox_frame)
        scrollbar.grid(row=0, column=1, sticky="ns")  # Changed to grid

        # Listbox
        listbox = tk.Listbox(
            listbox_frame,
            selectmode=tk.MULTIPLE,
            exportselection=False,
            height=8 if not is_system else 6,  # Make them taller
            bg=self.colors["white"],
            fg=self.colors["dark"],
            font=self.fonts["body"],
            selectbackground=self.colors["primary"],
            selectforeground="white",
            relief="flat",
            borderwidth=0,
            highlightthickness=1,
            highlightcolor=self.colors["primary"],
            yscrollcommand=scrollbar.set,
        )
        listbox.grid(row=0, column=0, sticky="nsew")  # Changed to grid and nsew
        scrollbar.config(command=listbox.yview)

        # Populate listbox
        listbox.insert(tk.END, "All")
        for value in self.all_values[column]:
            listbox.insert(tk.END, value)

        # Store reference
        if not hasattr(self, "listboxes"):
            self.listboxes = {}
        self.listboxes[column] = listbox

    def create_action_section(self, parent):
        """Enhanced action section with conditional transitivity option"""
        # Action container
        action_container = ttk.Frame(parent, style="Card.TFrame", padding=20)
        action_container.grid(row=4, column=0, sticky="ew")
        action_container.grid_columnconfigure(1, weight=1)

        # Action dropdown - keep original options
        action_label = ttk.Label(
            action_container, text="Action", style="FilterLabel.TLabel"
        )
        action_label.grid(row=0, column=0, sticky="w", padx=(0, 15), pady=(0, 15))

        self.action_cb = ttk.Combobox(
            action_container,
            values=[
                "Standard Excel",
                "Generate mapping spec excels for all subdomains",
                "Export into Data Lineage App",
                "PII Field Detection",
            ],
            state="readonly",
            style="Modern.TCombobox",
            font=self.fonts["combobox"],
        )
        self.action_cb.current(0)
        self.action_cb.grid(row=0, column=1, sticky="ew", padx=(0, 15), pady=(0, 15))

        # Frame for transitivity option (initially hidden)
        self.transitivity_frame = ttk.Frame(action_container)
        self.transitivity_frame.grid(
            row=1, column=0, columnspan=2, sticky="ew", pady=(0, 10)
        )
        self.transitivity_frame.grid_columnconfigure(1, weight=1)

        # Transitivity checkbox and dropdown
        self.enable_transitivity = tk.BooleanVar(value=False)
        self.transitivity_check = ttk.Checkbutton(
            self.transitivity_frame,
            text="Generate mapping to additional target system",
            variable=self.enable_transitivity,
            command=self.on_transitivity_toggle,
        )

        self.transitivity_target = ttk.Combobox(
            self.transitivity_frame,
            values=[],
            state="disabled",
            style="Modern.TCombobox",
            font=self.fonts["combobox"],
            width=15,
        )

        # Initially hide transitivity options
        self.transitivity_frame.grid_remove()

        # Excel path
        path_label = ttk.Label(
            action_container, text="Excel Path", style="FilterLabel.TLabel"
        )
        path_label.grid(row=2, column=0, sticky="w", padx=(0, 15), pady=(0, 15))

        path_frame = ttk.Frame(action_container)
        path_frame.grid(row=2, column=1, sticky="ew", padx=(0, 15), pady=(0, 15))
        path_frame.grid_columnconfigure(0, weight=1)

        self.excel_path = ttk.Entry(
            path_frame, style="Modern.TEntry", font=self.fonts["body"]
        )
        self.excel_path.insert(0, "output.xlsx")
        self.excel_path.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        browse_btn = ttk.Button(
            path_frame, text="Browse...", command=self.browse, style="Primary.TButton"
        )
        browse_btn.grid(row=0, column=1)

        # Action buttons
        button_frame = ttk.Frame(action_container)
        button_frame.grid(row=3, column=0, columnspan=2, pady=(20, 0))

        self.run_btn = ttk.Button(
            button_frame,
            text="Run",
            command=self.run,
            style="Success.TButton",
            width=15,
        )
        self.run_btn.pack(side="left", padx=(0, 10))

        cancel_btn = ttk.Button(
            button_frame,
            text="Cancel",
            command=self.root.destroy,
            style="Danger.TButton",
            width=15,
        )
        cancel_btn.pack(side="left")

        # Status label
        self.status_label = ttk.Label(
            action_container, text="", 
            foreground=self.colors["success"],
            font=self.fonts["status"]
        )
        self.status_label.grid(row=4, column=0, columnspan=2, pady=(15, 0))

    def on_canvas_configure(self, event):
        """Handle canvas resize to adjust the scrollable frame width"""
        canvas_width = event.width
        self.canvas.itemconfig(self.canvas_frame, width=canvas_width)

    def on_mousewheel(self, event):
        """Handle mouse wheel scrolling"""
        # Check if the canvas is scrollable (content height > canvas height)
        if self.canvas.winfo_reqheight() < self.scrollable_frame.winfo_reqheight():
            # Windows and MacOS
            if hasattr(event, "delta"):
                delta = event.delta
                self.canvas.yview_scroll(int(-1 * (delta / 120)), "units")
            # Linux
            elif event.num == 4:
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.canvas.yview_scroll(1, "units")

    def update_scroll_region(self):
        """Update the scroll region to encompass all content"""
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def scroll_to_top(self):
        """Scroll to the top of the window"""
        self.canvas.yview_moveto(0)

    def scroll_to_bottom(self):
        """Scroll to the bottom of the window"""
        self.canvas.yview_moveto(1)


    def setup_bindings(self):
        """Setup all event bindings including mouse wheel for ALL tabs"""
        try:
            # Bind filter listboxes
            for i, col in enumerate(FILTER_COLUMNS):
                if col in self.listboxes:
                    listbox = self.listboxes.get(col)
                    if listbox:
                        listbox.bind(
                            "<<ListboxSelect>>",
                            lambda e, column=col, index=i: self.on_filter_select(
                                column, index
                            ),
                        )

            # FIXED: Use single unified mouse wheel handler
            self.root.bind_all("<MouseWheel>", self._on_unified_mousewheel)
            self.root.bind_all("<Button-4>", self._on_unified_mousewheel)
            self.root.bind_all("<Button-5>", self._on_unified_mousewheel)

        except Exception as e:
            logger.error(f"Error setting up bindings: {e}")


    def _on_unified_mousewheel(self, event):
        """Unified mouse wheel handler that routes to correct canvas based on active tab"""
        try:
            active_tab = self.notebook.select()

            # Determine scroll direction and amount
            if hasattr(event, "delta"):
                delta = int(-1 * (event.delta / 120))
            elif event.num == 4:
                delta = -1
            elif event.num == 5:
                delta = 1
            else:
                return

            # Route to appropriate canvas based on active tab
            if active_tab == str(self.main_tab) and hasattr(self, "canvas"):
                self.canvas.yview_scroll(delta, "units")
            elif active_tab == str(self.upload_tab) and hasattr(self, "upload_canvas"):
                self.upload_canvas.yview_scroll(delta, "units")
            elif active_tab == str(self.pii_tab) and hasattr(self, "pii_canvas"):
                self.pii_canvas.yview_scroll(delta, "units")
        except Exception as e:
            logger.error(f"Mouse wheel error: {e}")

    def check_transitivity_options(self):
        """Check and update transitivity options based on current selections"""
        try:
            # Get current filter selections
            filters = self.get_current_filters()

            # Check if exactly one legacy system is selected
            legacy_systems = filters.get("legacysystem", [])
            target_systems = filters.get("targetsystem", [])

            if len(legacy_systems) == 1 and len(target_systems) <= 1:
                legacy_system = legacy_systems[0]
                current_target = target_systems[0] if target_systems else None

                # Check available transitivity targets
                conn = get_connection()
                try:
                    available = check_transitivity_availability(
                        conn, legacy_system, current_target
                    )

                    if available:
                        # Show transitivity options
                        self.transitivity_frame.grid()
                        self.transitivity_check.grid(
                            row=0, column=0, sticky="w", padx=(0, 10)
                        )
                        self.transitivity_target.grid(row=0, column=1, sticky="w")

                        # Update dropdown values
                        self.transitivity_target["values"] = available
                        if available:
                            self.transitivity_target.current(0)

                        # Update label text
                        label_text = f"Generate mapping to additional target system (via {current_target or 'intermediate'})"
                        self.transitivity_check.config(text=label_text)
                    else:
                        # Hide transitivity options
                        self.transitivity_frame.grid_remove()
                        self.enable_transitivity.set(False)
                finally:
                    conn.close()
            else:
                # Hide transitivity options
                self.transitivity_frame.grid_remove()
                self.enable_transitivity.set(False)

        except Exception as e:
            logger.error(f"Error checking transitivity options: {e}")

    def on_select(self, changed_col, changed_index):
        """Handle cascading filter selection with null safety"""
        try:
            # Get current filters
            filters = {}
            for col, lb in self.listboxes.items():
                sel = [lb.get(i) for i in lb.curselection()]
                filters[col] = [] if (not sel or "All" in sel) else sel

            # Apply filters to get current dataset
            current = self.rows_a
            for col, vals in filters.items():
                if vals:
                    # Safe filtering with null checks
                    current = [
                        r
                        for r in current
                        if r is not None and isinstance(r, dict) and r.get(col) in vals
                    ]

            # Update downstream listboxes
            start = FILTER_COLUMNS.index(changed_col) + 1
            for col in FILTER_COLUMNS[start:]:
                lb = self.listboxes[col]
                lb.delete(0, tk.END)
                lb.insert(tk.END, "All")
                # Safe value extraction with null checks
                vals = sorted(
                    {
                        r.get(col)
                        for r in current
                        if r is not None
                        and isinstance(r, dict)
                        and r.get(col) is not None
                    }
                )
                for v in vals:
                    lb.insert(tk.END, v)
        except Exception as e:
            self.show_status(f"Filter error: {str(e)}", "error")

    def on_filter_select(self, changed_col, changed_index):
        """Enhanced filter selection handler with error handling"""
        try:
            # Call original on_select logic
            self.on_select(changed_col, changed_index)

            # Check if we should show transitivity options
            self.check_transitivity_options()
        except Exception as e:
            logger.error(f"Error in filter selection: {e}")

    def on_transitivity_toggle(self):
        """Handle transitivity checkbox toggle"""
        if self.enable_transitivity.get():
            self.transitivity_target.config(state="readonly")
        else:
            self.transitivity_target.config(state="disabled")

    def get_current_filters(self):
        """Get current filter selections"""
        filters = {}
        for col, lb in self.listboxes.items():
            sel = [lb.get(i) for i in lb.curselection()]
            filters[col] = [] if (not sel or "All" in sel) else sel

    def browse(self):
        """Open file dialog for Excel path"""
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=self.excel_path.get(),
        )
        if path:
            self.excel_path.delete(0, tk.END)
            self.excel_path.insert(0, path)

    def show_status(self, message, status_type="success"):
        """Show status message"""
        color = (
            self.colors["success"]
            if status_type == "success"
            else self.colors["danger"]
        )
        self.status_label.config(text=message, foreground=color)
        self.root.after(5000, lambda: self.status_label.config(text=""))


    def get_current_filters(self):
        """Get current filter selections with null safety"""
        filters = {}
        try:
            if hasattr(self, "listboxes"):
                for col, lb in self.listboxes.items():
                    if lb is not None:
                        sel = [lb.get(i) for i in lb.curselection()]
                        filters[col] = [] if (not sel or "All" in sel) else sel
                    else:
                        filters[col] = []
        except Exception as e:
            logger.error(f"Error getting filters: {e}")
        return filters

    def create_pii_tab_content(self, parent):
        """Create content for PII Detection tab"""
        # Create scrollable content
        self.pii_canvas = tk.Canvas(
            parent, bg=self.colors["light"], highlightthickness=0
        )
        self.pii_scrollbar = ttk.Scrollbar(
            parent, orient="vertical", command=self.pii_canvas.yview
        )
        self.pii_scrollable_frame = ttk.Frame(self.pii_canvas, style="Main.TFrame")

        self.pii_scrollable_frame.bind(
            "<Configure>",
            lambda e: self.pii_canvas.configure(
                scrollregion=self.pii_canvas.bbox("all")
            ),
        )

        self.pii_canvas_frame = self.pii_canvas.create_window(
            (0, 0), window=self.pii_scrollable_frame, anchor="nw"
        )
        self.pii_canvas.configure(yscrollcommand=self.pii_scrollbar.set)

        self.pii_canvas.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        self.pii_scrollbar.grid(row=0, column=1, sticky="nsew", pady=0)

        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_columnconfigure(1, weight=0)

        main_content = ttk.Frame(
            self.pii_scrollable_frame, style="Main.TFrame", padding=20
        )
        main_content.grid(row=0, column=0, sticky="nsew")
        main_content.grid_columnconfigure(0, weight=1)

        self.pii_scrollable_frame.grid_columnconfigure(0, weight=1)
        self.pii_scrollable_frame.grid_rowconfigure(0, weight=1)

        self.pii_canvas.bind("<Configure>", self.on_pii_canvas_configure)

        # Create sections
        self.create_pii_header_section(main_content)
        self.create_pii_file_section(main_content)
        self.create_pii_results_section(main_content)
        self.create_pii_action_section(main_content)

        self.root.after(100, self.update_pii_scroll_region)

    def on_pii_canvas_configure(self, event):
        """Handle PII canvas resize"""
        canvas_width = event.width
        self.pii_canvas.itemconfig(self.pii_canvas_frame, width=canvas_width)

    def update_pii_scroll_region(self):
        """Update the PII scroll region"""
        self.pii_canvas.configure(scrollregion=self.pii_canvas.bbox("all"))

    def create_pii_header_section(self, parent):
        """Create header for PII detection tab"""
        header_text = ttk.Label(
            parent, text=" PII Field Detection", style="Header.TLabel"
        )
        header_text.grid(row=0, column=0, sticky="w", pady=(0, 10))

        desc_text = ttk.Label(
            parent,
            text="Upload your mapping specification Excel file to automatically detect fields that may contain Personally Identifiable Information (PII) using AI-powered analysis.",
            style="FilterLabel.TLabel",
            foreground=self.colors["muted"],
            wraplength=1000,
        )
        desc_text.grid(row=1, column=0, sticky="w", pady=(0, 20))

        # Separator
        separator = ttk.Separator(parent, orient="horizontal")
        separator.grid(row=2, column=0, sticky="ew", pady=(0, 20))

    def create_pii_file_section(self, parent):
        """Create the file selection section with custom rules option"""
        # Section header
        section_header = ttk.Label(parent, text="File Selection", style="Header.TLabel")
        section_header.grid(row=3, column=0, sticky="w", pady=(0, 15))

        file_container = ttk.Frame(parent, style="Card.TFrame", padding=20)
        file_container.grid(row=4, column=0, sticky="ew", pady=(0, 20))
        file_container.grid_columnconfigure(0, weight=1)

        # Excel File subheader
        file_subheader = ttk.Label(
            file_container,
            text="Excel File",
            style="FilterLabel.TLabel",
            font=("Segoe UI", 15, "bold"),
        )
        file_subheader.grid(row=0, column=0, sticky="w", pady=(0, 8))

        # File input frame (browse button)
        file_input_frame = ttk.Frame(file_container)
        file_input_frame.grid(row=1, column=0, sticky="ew", pady=(0, 15))
        file_input_frame.grid_columnconfigure(0, weight=1)

        self.pii_file_path = ttk.Entry(
            file_input_frame, style="Modern.TEntry", font=self.fonts["body"]
        )
        self.pii_file_path.insert(0, "No file selected")
        self.pii_file_path.config(state="readonly")
        self.pii_file_path.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        browse_btn = ttk.Button(
            file_input_frame,
            text="Browse...",
            command=self.browse_pii_file,
            style="Primary.TButton",
        )
        browse_btn.grid(row=0, column=1)

        # Validation status
        self.pii_validation_label = ttk.Label(
            file_container, text="", 
            foreground=self.colors["muted"],
            font=self.fonts["status"],
        )
        self.pii_validation_label.grid(row=2, column=0, sticky="w", pady=(0, 15))

        # Separator
        separator = ttk.Separator(file_container, orient="horizontal")
        separator.grid(row=3, column=0, sticky="ew", pady=(0, 15))

        # Custom Rules Section (OPTIONAL)
        custom_rules_header = ttk.Label(
            file_container,
            text="Custom Rules (Optional)",
            style="FilterLabel.TLabel",
            font=("Segoe UI", 15, "bold"),
        )
        custom_rules_header.grid(row=4, column=0, sticky="w", pady=(0, 10))

        # Custom rules text box
        custom_rules_text_frame = ttk.Frame(file_container)
        custom_rules_text_frame.grid(row=5, column=0, sticky="ew", pady=(0, 15))
        custom_rules_text_frame.grid_columnconfigure(0, weight=1)

        custom_rules_scrollbar = ttk.Scrollbar(custom_rules_text_frame)
        custom_rules_scrollbar.grid(row=0, column=1, sticky="ns")

        self.pii_custom_rules_text = tk.Text(
            custom_rules_text_frame,
            height=6,
            font=self.fonts["body"],
            wrap="word",
            bg=self.colors["white"],
            relief="solid",
            borderwidth=1,
            yscrollcommand=custom_rules_scrollbar.set,
        )
        self.pii_custom_rules_text.grid(row=0, column=0, sticky="ew")
        custom_rules_scrollbar.config(command=self.pii_custom_rules_text.yview)

        # Placeholder text
        placeholder = "Example:\nsalary\ncompensation\nbonus_amount"
        self.pii_custom_rules_text.insert("1.0", placeholder)
        self.pii_custom_rules_text.config(foreground=self.colors["muted"])

        # Bind focus events for placeholder
        self.pii_custom_rules_text.bind("<FocusIn>", self.on_custom_rules_focus_in)
        self.pii_custom_rules_text.bind("<FocusOut>", self.on_custom_rules_focus_out)

        # Options frame (Scope and Action side by side)
        options_frame = tk.Frame(file_container, bg=self.colors["white"])
        options_frame.grid(row=6, column=0, sticky="ew", pady=(0, 15))
        options_frame.grid_columnconfigure(0, weight=1)
        options_frame.grid_columnconfigure(1, weight=1)

        # LEFT: Scope selection
        scope_frame = tk.Frame(options_frame, bg=self.colors["white"])
        scope_frame.grid(row=0, column=0, sticky="ew", padx=(0, 10))

        scope_label = ttk.Label(
            scope_frame,
            text="Apply To",
            style="FilterLabel.TLabel",
            font=("Segoe UI", 13, "bold"),
        )
        scope_label.pack(anchor="w", pady=(0, 5))

        self.pii_custom_scope = ttk.Combobox(
            scope_frame,
            values=["Field Name", "Field Description", "Table Name"],
            state="readonly",
            style="Modern.TCombobox",
            font=self.fonts["combobox"],
        )
        self.pii_custom_scope.current(0)  # Default: Field Name
        self.pii_custom_scope.pack(fill="x")

        # RIGHT: Action selection
        action_frame = tk.Frame(options_frame, bg=self.colors["white"])
        action_frame.grid(row=0, column=1, sticky="ew", padx=(10, 0))

        action_label = ttk.Label(
            action_frame,
            text="Mark As",
            style="FilterLabel.TLabel",
            font=("Segoe UI", 13, "bold"),
        )
        action_label.pack(anchor="w", pady=(0, 5))

        self.pii_custom_action = tk.StringVar(value="PII")

        action_radio_frame = tk.Frame(action_frame, bg=self.colors["white"])
        action_radio_frame.pack(fill="x")

        # Custom style for larger radio buttons
        radio_font = tkFont.Font(family="Segoe UI", size=15, weight="bold")
        
        pii_radio = ttk.Radiobutton(
            action_radio_frame,
            text="PII",
            variable=self.pii_custom_action,
            value="PII",
            command=self.on_custom_action_change,
        )
        pii_radio.pack(side="left", padx=(0, 15))
        
        # Apply custom font using configure
        pii_radio.configure(style="Large.TRadiobutton")

        non_pii_radio = ttk.Radiobutton(
            action_radio_frame,
            text="Non-PII",
            variable=self.pii_custom_action,
            value="Non-PII",
        )
        non_pii_radio.pack(side="left")
        non_pii_radio.configure(style="Large.TRadiobutton")

        # PII Type (only shown when "PII" is selected)
        self.pii_type_frame = tk.Frame(file_container, bg=self.colors["white"])
        self.pii_type_frame.grid(row=7, column=0, sticky="ew", pady=(0, 15))

        pii_type_label = ttk.Label(
            self.pii_type_frame,
            text="PII Category (optional)",
            style="FilterLabel.TLabel",
            font=("Segoe UI", 13, "bold"),
        )
        pii_type_label.grid(row=0, column=0, sticky="w", pady=(0, 5))

        self.pii_custom_type = ttk.Entry(
            self.pii_type_frame,
            style="Modern.TEntry",
            font=self.fonts["body"],
        )
        self.pii_custom_type.insert(0, "Custom_PII")
        self.pii_custom_type.grid(row=1, column=0, sticky="ew")

        # Detect button
        self.pii_detect_btn = ttk.Button(
            file_container,
            text="Run",
            command=self.detect_pii,
            style="Success.TButton",
            state="disabled",
            width=25,
        )
        self.pii_detect_btn.grid(row=8, column=0, pady=(10, 0))


    def on_custom_rules_focus_in(self, event):
        """Remove placeholder text on focus"""
        current_text = self.pii_custom_rules_text.get("1.0", "end-1c")
        if current_text.startswith("Example:"):
            self.pii_custom_rules_text.delete("1.0", tk.END)
            self.pii_custom_rules_text.config(foreground=self.colors["dark"])


    def on_custom_rules_focus_out(self, event):
        """Add placeholder text if empty"""
        current_text = self.pii_custom_rules_text.get("1.0", "end-1c").strip()
        if not current_text:
            placeholder = "Example:\nsalary\ncompensation\nbonus_amount"
            self.pii_custom_rules_text.insert("1.0", placeholder)
            self.pii_custom_rules_text.config(foreground=self.colors["muted"])


    def on_custom_action_change(self):
        """Show/hide PII type field based on action selection"""
        if self.pii_custom_action.get() == "PII":
            self.pii_type_frame.grid()
        else:
            self.pii_type_frame.grid_remove()

    def create_pii_results_section(self, parent):
        """Create results display section"""
        section_header = ttk.Label(
            parent, text="Detection Results", style="Header.TLabel"
        )
        section_header.grid(row=5, column=0, sticky="w", pady=(0, 15))

        results_container = ttk.Frame(parent, style="Card.TFrame", padding=20)
        results_container.grid(row=6, column=0, sticky="ew", pady=(0, 20))
        results_container.grid_columnconfigure(0, weight=1)

        # Summary statistics (initially hidden)
        self.pii_stats_frame = ttk.Frame(results_container)
        self.pii_stats_frame.grid(row=0, column=0, sticky="ew", pady=(0, 15))
        self.pii_stats_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        # Create 4 stat boxes
        self.pii_total_label = self.create_stat_box(
            self.pii_stats_frame, "Total Fields", "0", 0
        )
        self.pii_detected_label = self.create_stat_box(
            self.pii_stats_frame, "PII Detected", "0", 1
        )
        self.pii_clean_label = self.create_stat_box(
            self.pii_stats_frame, "Non-PII", "0", 2
        )
        self.pii_rate_label = self.create_stat_box(
            self.pii_stats_frame, "Detection Rate", "0%", 3
        )

        self.pii_stats_frame.grid_remove()  # Hide initially

        # Preview table
        preview_label = ttk.Label(
            results_container,
            text="Preview",
            style="FilterLabel.TLabel",
            font=("Segoe UI", 14, "bold"),
        )
        preview_label.grid(row=1, column=0, sticky="w", pady=(0, 10))

        # Create text widget for preview
        preview_frame = ttk.Frame(results_container)
        preview_frame.grid(row=2, column=0, sticky="nsew")
        preview_frame.grid_columnconfigure(0, weight=1)
        preview_frame.grid_rowconfigure(0, weight=1)

        preview_scrollbar = ttk.Scrollbar(preview_frame)
        preview_scrollbar.grid(row=0, column=1, sticky="ns")

        self.pii_preview_text = tk.Text(
            preview_frame,
            height=12,
            font=self.fonts["body"],
            wrap="none",
            bg=self.colors["light"],
            relief="flat",
            borderwidth=0,
            yscrollcommand=preview_scrollbar.set,
        )
        self.pii_preview_text.grid(row=0, column=0, sticky="nsew")
        preview_scrollbar.config(command=self.pii_preview_text.yview)

        self.pii_preview_text.insert(
            "1.0",
            "No detection results yet.\n\nPlease select a file and click 'Detect PII Fields' to begin.",
        )
        self.pii_preview_text.config(state="disabled")

    def create_stat_box(self, parent, label, value, column):
        """Create a statistics box with white background and subtle shadow"""
        # Container frame for shadow effect
        container = tk.Frame(parent, bg=self.colors["border"], padx=2, pady=2)
        container.grid(row=0, column=column, sticky="ew", padx=5)

        # Inner frame with white background
        frame = tk.Frame(container, bg=self.colors["white"], padx=15, pady=15)
        frame.pack(fill="both", expand=True)
    
        label_widget = tk.Label(
            frame,
            text=label,
            bg=self.colors["white"],
            fg=self.colors["muted"],
            font=self.fonts["body"]
        )
        label_widget.pack()
    
        value_widget = tk.Label(
            frame,
            text=value,
            bg=self.colors["white"],
            fg=self.colors["primary"],
            font=("Segoe UI", 24, "bold")
        )
        value_widget.pack()
    
        return value_widget

    def create_pii_action_section(self, parent):
        """Create action buttons section"""
        action_container = ttk.Frame(parent, style="Card.TFrame", padding=20)
        action_container.grid(row=7, column=0, sticky="ew")
        action_container.grid_columnconfigure(0, weight=1)

        # Download button (initially disabled)
        button_frame = ttk.Frame(action_container)
        button_frame.pack(pady=(0, 15))

        self.pii_download_btn = ttk.Button(
            button_frame,
            text="Download Results",
            command=self.download_pii_results,
            style="Primary.TButton",
            state="disabled",
            width=25,
        )
        self.pii_download_btn.pack(side="left", padx=(0, 10))

        clear_btn = ttk.Button(
            button_frame,
            text="Clear",
            command=self.clear_pii,
            style="Danger.TButton",
            width=15,
        )
        clear_btn.pack(side="left")

        # Status label
        self.pii_status_label = ttk.Label(
            action_container,
            text="",
            foreground=self.colors["success"],
            font=self.fonts["body"],
        )
        self.pii_status_label.pack()

    def browse_pii_file(self):
        """Handle PII file selection"""
        file_path = filedialog.askopenfilename(
            title="Select Mapping Specification Excel File",
            filetypes=[("Excel Files", "*.xlsx *.xls *.xlsm")],
        )

        if file_path:
            self.pii_selected_file = file_path
            self.pii_file_path.config(state="normal")
            self.pii_file_path.delete(0, tk.END)
            self.pii_file_path.insert(0, os.path.basename(file_path))
            self.pii_file_path.config(state="readonly")

            self.pii_detect_btn.config(state="normal")
            self.pii_status_label.config(text="")
            self.pii_validation_label.config(
                text=f"File loaded: {os.path.basename(file_path)}", 
                foreground=self.colors["success"]
            )

    def detect_pii(self):
        """Execute PII detection with optional custom rules"""
        if not hasattr(self, "pii_selected_file"):
            messagebox.showerror("Error", "No file selected")
            return

        # Check API key
        gemini_api_key = os.getenv("GEMINI_API_KEY")
        if not gemini_api_key:
            messagebox.showerror(
                "API Key Missing",
                "Gemini API key not configured.\n\nPlease add GEMINI_API_KEY to your .env file.",
            )
            return

        try:
            # Disable button
            self.pii_detect_btn.config(state="disabled")
            self.pii_status_label.config(
                text="Initializing PII detection...", foreground=self.colors["primary"]
            )
            self.root.update()

            # Extract custom rules from UI (if provided)
            custom_rules_list = self._extract_custom_rules()

            if custom_rules_list:
                logger.info(f"Custom rules detected: {len(custom_rules_list)} rule(s)")
            else:
                logger.info("No custom rules provided, using system rules only")

            # Create orchestrator
            conn = get_connection()  
            orchestrator = PIIDetectionOrchestrator(
                gemini_api_key=gemini_api_key,
                db_connection=conn,  
                batch_size=12,
                custom_rules=custom_rules_list
            )
            # Create progress callback
            def update_progress(current_batch, total_batches):
                progress_text = (
                    f"Processing Batch {current_batch} of {total_batches}..."
                )
                self.pii_status_label.config(text=progress_text)
                self.root.update()

            # Process file
            output_file, pii_count, total_count, preview_df = orchestrator.process_file(
                self.pii_selected_file, progress_callback=update_progress
            )

            # Store results
            self.pii_output_file = output_file
            self.pii_results = {
                "pii_count": pii_count,
                "total_count": total_count,
                "preview_df": preview_df,
            }

            # Update UI
            self.display_pii_results(pii_count, total_count, preview_df)

            # Enable download
            self.pii_download_btn.config(state="normal")

            # Success message
            custom_rules_msg = f" (including {len(custom_rules_list)} custom rule(s))" if custom_rules_list else ""
            self.pii_status_label.config(
                text=f"Detection completed{custom_rules_msg}. Found {pii_count} PII fields out of {total_count} total fields",
                foreground=self.colors["success"],
            )

        except Exception as e:
            self.pii_status_label.config(
                text=f"Error: {str(e)}", foreground=self.colors["danger"]
            )
            logger.error(f"PII detection error: {e}", exc_info=True)
            messagebox.showerror(
                "Detection Error",
                f"An error occurred during PII detection:\n\n{str(e)}",
            )
        finally:
            self.pii_detect_btn.config(state="normal")


    def _extract_custom_rules(self):
        """
        Extract custom rules from UI components.

        Returns:
            List of dict with custom rules, or empty list if no rules provided
        """
        # Get text from custom rules text box
        custom_text = self.pii_custom_rules_text.get("1.0", "end-1c").strip()

        # Check if it's just the placeholder
        if not custom_text or custom_text.startswith("Example:"):
            return []

        # Get scope and action
        scope_display = self.pii_custom_scope.get()  # "Field Name", "Field Description", "Table Name"
        action = self.pii_custom_action.get()  # "PII" or "Non-PII"
        pii_type = self.pii_custom_type.get().strip() if action == "PII" else "Custom_Non_PII"

        # Map display names to internal names
        scope_map = {
            "Field Name": "fieldname",
            "Field Description": "description",
            "Table Name": "tablename"
        }
        scope_internal = scope_map.get(scope_display, "fieldname")

        # Parse patterns (one per line)
        patterns = [line.strip() for line in custom_text.split('\n') if line.strip()]

        if not patterns:
            return []

        # Create rule object
        custom_rule = {
            'patterns': patterns,
            'scope': scope_internal,
            'action': action,
            'pii_type': pii_type if action == "PII" else None
        }

        logger.info(f"Custom rule extracted:")
        logger.info(f"  Patterns: {len(patterns)}")
        logger.info(f"  Scope: {scope_internal}")
        logger.info(f"  Action: {action}")
        if action == "PII":
            logger.info(f"  PII Type: {pii_type}")

        return [custom_rule]  # Return as list (could support multiple rules in future)


    def display_pii_results(self, pii_count, total_count, preview_df):
        """Display PII detection results with clean formatting"""
        # Show statistics
        self.pii_stats_frame.grid()

        self.pii_total_label.config(text=str(total_count))
        self.pii_detected_label.config(text=str(pii_count))
        self.pii_clean_label.config(text=str(total_count - pii_count))

        detection_rate = (pii_count / total_count * 100) if total_count > 0 else 0
        self.pii_rate_label.config(text=f"{detection_rate:.1f}%")

        # Update preview
        self.pii_preview_text.config(state="normal")
        self.pii_preview_text.delete("1.0", tk.END)

        if len(preview_df) > 0:
            preview_text = "Top 10 PII Fields:\n\n"

            for idx, row in preview_df.iterrows():
                # Extract column number
                col_num = "N/A"
                for col_name in ["Column Number", "columnnumber", "column_number"]:
                    if col_name in row.index and pd.notna(row[col_name]):
                        col_num = row[col_name]
                        break

                # Extract table name
                table_name = "N/A"
                for col_name in [
                    "Target Table Name", "targettablename", "target_table_name",
                    "Legacy Table Name", "legacytablename", "legacy_table_name",
                ]:
                    if col_name in row.index and pd.notna(row[col_name]) and row[col_name] != "":
                        table_name = row[col_name]
                        break

                # Extract field name
                field_name = "N/A"
                for col_name in [
                    "Target Field Name", "targetfieldname", "target_field_name",
                    "Legacy Field Name", "legacyfieldname", "legacy_field_name",
                ]:
                    if col_name in row.index and pd.notna(row[col_name]) and row[col_name] != "":
                        field_name = row[col_name]
                        break
                    
                # Extract detection method - NEW
                detection_method = row.get('Detection_Method', 'N/A')

                # Simple clean format
                preview_text += f"#{idx + 1}\n"
                preview_text += f"  Column Number: {col_num}\n"
                preview_text += f"  Table:         {table_name}\n"
                preview_text += f"  Field:         {field_name}\n"
                preview_text += f"  PII Types:     {row.get('PII_Types', 'N/A')}\n"
                preview_text += f"  Confidence:    {row.get('Confidence_Level', 'N/A')}\n"
                preview_text += f"  Detection Method:  {detection_method}\n"  
                preview_text += f"{'-' * 80}\n\n"

            self.pii_preview_text.insert("1.0", preview_text)
        else:
            self.pii_preview_text.insert("1.0", "No PII fields detected in this file.")

        self.pii_preview_text.config(state="disabled")

    def download_pii_results(self):
        """Download PII detection results"""
        if not hasattr(self, "pii_output_file"):
            messagebox.showerror("Error", "No results to download")
            return

        # Ask where to save
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=os.path.basename(self.pii_output_file),
        )

        if save_path:
            import shutil

            shutil.copy(self.pii_output_file, save_path)
            messagebox.showinfo(
                "Download Complete", f"PII detection results saved to:\n\n{save_path}"
            )

    def clear_pii(self):
        """Clear PII detection data including custom rules"""
        if hasattr(self, "pii_selected_file"):
            delattr(self, "pii_selected_file")

        # Clear file selection
        self.pii_file_path.config(state="normal")
        self.pii_file_path.delete(0, tk.END)
        self.pii_file_path.insert(0, "No file selected")
        self.pii_file_path.config(state="readonly")

        # Clear custom rules text box
        self.pii_custom_rules_text.delete("1.0", tk.END)
        placeholder = "Example:\nsalary\ncompensation\nbonus_amount"
        self.pii_custom_rules_text.insert("1.0", placeholder)
        self.pii_custom_rules_text.config(foreground=self.colors["muted"])

        # Reset scope and action to defaults
        self.pii_custom_scope.current(0)  # Field Name
        self.pii_custom_action.set("PII")
        self.pii_custom_type.delete(0, tk.END)
        self.pii_custom_type.insert(0, "Custom_PII")

        # Show PII type frame (since default action is PII)
        self.pii_type_frame.grid()

        # Clear results
        self.pii_validation_label.config(text="")
        self.pii_status_label.config(text="")
        self.pii_detect_btn.config(state="disabled")
        self.pii_download_btn.config(state="disabled")

        self.pii_stats_frame.grid_remove()

        self.pii_preview_text.config(state="normal")
        self.pii_preview_text.delete("1.0", tk.END)
        self.pii_preview_text.insert(
            "1.0",
            "No detection results yet.\n\nPlease select a file and click 'Run' to begin.",
        )
        self.pii_preview_text.config(state="disabled")

        logger.info("PII detection cleared (including custom rules)")

    def run(self):
        """Enhanced run method with transitivity-only output when enabled"""
        try:
            # Disable run button during operation
            self.run_btn.config(state="disabled")
            self.show_status("Processing...", "success")
            self.root.update()

            # Build filters dict
            filters = self.get_current_filters()
            logger.info(f"Filters: {filters}")

            # Execute database operations
            conn = get_connection()
            try:
                pj_q, pj_p = build_projectid_query(filters)
                proj_rows = execute_query(conn, pj_q, pj_p)
                project_ids = [r["projectid"] for r in proj_rows]
                logger.info(f"Found {len(project_ids)} project IDs")

                tb_q, tb_p = build_tableb_query(project_ids)
                child_rows = execute_query(conn, tb_q, tb_p)
                logger.info(f"Found {len(child_rows)} child rows")

                # Check if transitivity is enabled
                if self.enable_transitivity.get() and self.transitivity_target.get():
                    logger.info("Transitivity is enabled")
                    transitivity_target = self.transitivity_target.get()
                    logger.info(f"Transitivity target: {transitivity_target}")

                    # Check if we have data
                    if not child_rows:
                        error_msg = "No mapping data found for the selected filters. Cannot apply transitivity."
                        self.show_status(f"{error_msg}", "error")
                        messagebox.showerror("No Data", error_msg)
                        return

                    # Filter to only rows with valid intermediate systems (Oracle or Workday)
                    # Group by targetsystem to see what we have
                    target_systems = {}
                    for row in child_rows:
                        ts = row.get("targetsystem")
                        if ts:
                            target_systems[ts] = target_systems.get(ts, 0) + 1

                    logger.info(f"Target systems found: {target_systems}")

                    # Filter to only Oracle or Workday as intermediate
                    valid_intermediate_rows = [
                        row
                        for row in child_rows
                        if row.get("targetsystem") in ["Oracle", "Workday"]
                    ]

                    if not valid_intermediate_rows:
                        error_msg = (
                            f"Cannot apply transitivity: found mappings to {list(target_systems.keys())}, "
                            f"but transitivity requires mappings to 'Oracle' or 'Workday' as intermediate systems.\n\n"
                            f"Your data contains:\n"
                            + "\n".join(
                                [
                                    f"  - {k}: {v} rows"
                                    for k, v in target_systems.items()
                                ]
                            )
                            + f"\n\nPlease select a target system filter for 'Oracle' or 'Workday' to use transitivity."
                        )
                        self.show_status(f"No valid intermediate system found", "error")
                        messagebox.showerror("Transitivity Error", error_msg)
                        return

                    # Use the filtered rows
                    child_rows = valid_intermediate_rows
                    intermediate_system = child_rows[0].get("targetsystem")
                    logger.info(
                        f"Using intermediate system: {intermediate_system} ({len(child_rows)} rows)"
                    )

                    # Verify the transitivity direction makes sense
                    if intermediate_system == transitivity_target:
                        error_msg = (
                            f"Cannot apply transitivity: intermediate system and target are both '{intermediate_system}'.\n\n"
                            f"Transitivity works by converting through an intermediate system to a different target."
                        )
                        self.show_status(f"Invalid transitivity configuration", "error")
                        messagebox.showerror("Transitivity Error", error_msg)
                        return

                    self.show_status(
                        f"Applying transitivity: {intermediate_system} → {transitivity_target}...",
                        "success",
                    )
                    self.root.update()

                    # Get WO mappings
                    wo_mappings = fetch_workday_oracle_mapping(conn)
                    logger.info(f"Found {len(wo_mappings)} Workday-Oracle mappings")

                    if not wo_mappings:
                        error_msg = (
                            "No Workday-Oracle mapping data found in the translation table.\n\n"
                            f"To use transitivity, the table '{WORKDAY_ORACLE_MAPPING_TABLE}' must be populated "
                            f"with bidirectional mappings between Workday and Oracle fields."
                        )
                        self.show_status("No Workday-Oracle mappings found", "error")
                        messagebox.showerror("Transitivity Error", error_msg)
                        return

                    # Apply transitivity
                    logger.info(
                        f"Applying transitivity mapping from {intermediate_system} to {transitivity_target}"
                    )
                    transitive_rows = apply_transitivity_mapping_debug(
                        child_rows,
                        wo_mappings,
                        intermediate_system,  # The current target (Oracle or Workday)
                        transitivity_target,  # The desired final target
                    )

                    if not transitive_rows:
                        error_msg = (
                            "Transitivity mapping produced no results.\n\n"
                            "This usually means there are no matching fields in the Workday-Oracle mapping table "
                            "for the fields in your source mappings."
                        )
                        self.show_status("No transitive mappings generated", "error")
                        messagebox.showerror("Transitivity Error", error_msg)
                        return

                    logger.info(f"Generated {len(transitive_rows)} transitive rows")

                    # Calculate translation statistics for user feedback
                    mapped_count = sum(
                        1
                        for r in transitive_rows
                        if r.get("translation_status") == "Mapped"
                    )
                    no_mapping_count = sum(
                        1
                        for r in transitive_rows
                        if r.get("translation_status") == "No Mapping"
                    )
                    multiple_count = sum(
                        1
                        for r in transitive_rows
                        if r.get("translation_status") == "Multiple Mappings"
                    )
                    success_rate = (
                        (mapped_count / len(transitive_rows) * 100)
                        if transitive_rows
                        else 0
                    )

                    # Process ONLY transitive output
                    action = self.action_cb.get()

                    if action == "Generate mapping spec excels for all subdomains":
                        # Generate transitive versions only
                        df = pd.DataFrame(transitive_rows)
                        count = 0
                        for sub, grp in df.groupby("filename"):
                            fname = f"{sub}_to_{transitivity_target}.xlsx"
                            export_summarized_excel_with_status(
                                grp.to_dict("records"), fname
                            )
                            count += 1
                            logger.info(f"Exported: {fname}")

                        # Show summary
                        summary_msg = (
                            f"Generated {count} transitive Excel files!\n\n"
                            f"Translation Summary:\n"
                            f"  • Total records: {len(transitive_rows)}\n"
                            f"  • Successfully mapped: {mapped_count} ({success_rate:.1f}%)\n"
                            f"  • No mapping found: {no_mapping_count}\n"
                            f"  • Multiple mappings: {multiple_count}\n\n"
                            f"Check the 'Translation Status' sheet in each Excel file for details."
                        )

                        self.show_status(
                            f"Generated {count} files - {success_rate:.1f}% success rate",
                            "success",
                        )
                        messagebox.showinfo("Export Complete", summary_msg)

                    elif action == "Standard Excel":
                        # Generate single transitive file only
                        base_path = self.excel_path.get().strip()
                        trans_path = base_path.replace(
                            ".xlsx", f"_to_{transitivity_target}.xlsx"
                        )
                        export_standard_excel(transitive_rows, trans_path)
                        logger.info(f"Exported: {trans_path}")

                        summary_msg = (
                            f"Transitive mapping saved!\n\n"
                            f"Translation Summary:\n"
                            f"  • Total records: {len(transitive_rows)}\n"
                            f"  • Successfully mapped: {mapped_count} ({success_rate:.1f}%)\n"
                            f"  • No mapping found: {no_mapping_count}\n"
                            f"  • Multiple mappings: {multiple_count}\n\n"
                            f"File: {trans_path}"
                        )

                        self.show_status(
                            f"Export complete - {success_rate:.1f}% success rate",
                            "success",
                        )
                        messagebox.showinfo("Export Complete", summary_msg)

                    else:  # Export into Data Lineage App
                        # Store transitive only
                        conn_output = get_outputconnection()
                        store_output_to_db(
                            transitive_rows,
                            conn_output,
                            table_name=postgres_outputtable,
                            column_mapping=COLUMN_MAPPING,
                            columns_to_insert=desired,
                        )
                        conn_output.close()
                        logger.info(f"Exported {len(transitive_rows)} rows to database")

                        summary_msg = (
                            f"Transitive mappings exported to database!\n\n"
                            f"Translation Summary:\n"
                            f"  • Total records: {len(transitive_rows)}\n"
                            f"  • Successfully mapped: {mapped_count} ({success_rate:.1f}%)\n"
                            f"  • No mapping found: {no_mapping_count}\n"
                            f"  • Multiple mappings: {multiple_count}"
                        )

                        self.show_status(
                            f"Database export complete - {success_rate:.1f}% success rate",
                            "success",
                        )
                        messagebox.showinfo("Export Complete", summary_msg)

                    # Early return - don't process original output
                    return

                # Normal processing (when transitivity is NOT enabled)
                logger.info("Normal processing (no transitivity)")
                action = self.action_cb.get()

                if action == "Generate mapping spec excels for all subdomains":
                    df = pd.DataFrame(child_rows)
                    count = 0
                    for sub, grp in df.groupby("filename"):
                        fname = f"{sub}.xlsx"
                        export_summarized_excel(grp.to_dict("records"), fname)
                        count += 1
                    self.show_status(
                        f"Generated {count} Excel files successfully!", "success"
                    )

                elif action == "Standard Excel":
                    path = self.excel_path.get().strip()
                    export_standard_excel(child_rows, path)
                    self.show_status(f"Excel saved to: {path}", "success")

                elif action == "Export into Data Lineage App":
                    conn_output = get_outputconnection()
                    store_output_to_db(
                        child_rows,
                        conn_output,
                        table_name=postgres_outputtable,
                        column_mapping=COLUMN_MAPPING,
                        columns_to_insert=desired,
                    )
                    conn_output.close()
                    self.show_status(
                        "Data exported to Data Lineage App successfully", "success"
                    )

                elif action == "PII Field Detection":
                    # Step 1: Generate mapping spec Excel files (one per subdomain)
                    logger.info("PII Field Detection: Step 1 - Generating mapping spec files")
                    self.show_status("Generating mapping spec Excel files...", "success")
                    self.root.update()
                    
                    generated_files = []
                    
                    try:
                        # Generate multiple Excel files (one per subdomain)
                        df = pd.DataFrame(child_rows)
                        count = 0
                        for sub, grp in df.groupby("filename"):
                            fname = f"{sub}.xlsx"
                            export_summarized_excel(grp.to_dict("records"), fname)
                            generated_files.append(fname)
                            count += 1
                            logger.info(f"Generated mapping file: {fname}")
                        
                        self.show_status(
                            f"Generated {count} mapping files. Starting PII detection...", 
                            "success"
                        )
                        self.root.update()
                        
                    except Exception as e:
                        error_msg = f"Failed to generate mapping spec files: {str(e)}"
                        self.show_status(error_msg, "error")
                        messagebox.showerror("Generation Error", error_msg)
                        logger.error(f"Mapping generation error: {e}", exc_info=True)
                        return  # Stop here, don't proceed to PII detection
                    
                    # Step 2: Run PII detection on each generated file
                    logger.info(f"PII Field Detection: Step 2 - Running PII detection on {len(generated_files)} files")
                    
                    # Check API key first
                    gemini_api_key = os.getenv("GEMINI_API_KEY")
                    if not gemini_api_key:
                        error_msg = (
                            "Gemini API key not configured.\n\n"
                            "Please add GEMINI_API_KEY to your .env file.\n\n"
                            f"Mapping files have been saved ({count} files), "
                            "but PII detection was skipped."
                        )
                        self.show_status("API key missing", "error")
                        messagebox.showwarning("API Key Missing", error_msg)
                        return
                    
                    try:
                        self.show_status("Initializing PII detection...", "success")
                        self.root.update()
                        
                        # Create orchestrator (NO custom rules as per requirement)
                        conn_pii = get_connection()
                        orchestrator = PIIDetectionOrchestrator(
                            gemini_api_key=gemini_api_key,
                            db_connection=conn_pii,
                            batch_size=12,
                            custom_rules=None  # Skip custom rules as requested
                        )
                        
                        # Track results for each file
                        pii_results = []
                        total_pii_count = 0
                        total_field_count = 0
                        
                        # Process each generated file
                        for idx, input_file in enumerate(generated_files, 1):
                            logger.info(f"Processing file {idx}/{len(generated_files)}: {input_file}")
                            
                            # Define progress callback for this file
                            def update_progress(current_batch, total_batches):
                                progress_text = (
                                    f"PII Detection [{idx}/{len(generated_files)}]: "
                                    f"Processing {input_file} - Batch {current_batch}/{total_batches}"
                                )
                                self.show_status(progress_text, "success")
                                self.root.update()
                            
                            try:
                                # Process this file
                                pii_output_file, pii_count, field_count, preview_df = orchestrator.process_file(
                                    input_file,
                                    progress_callback=update_progress
                                )
                                
                                # Track results
                                pii_results.append({
                                    'mapping_file': input_file,
                                    'pii_file': pii_output_file,
                                    'pii_count': pii_count,
                                    'total_count': field_count
                                })
                                
                                total_pii_count += pii_count
                                total_field_count += field_count
                                
                                logger.info(
                                    f"Completed {input_file}: "
                                    f"{pii_count}/{field_count} PII fields detected"
                                )
                                
                            except Exception as e:
                                logger.error(f"PII detection failed for {input_file}: {e}", exc_info=True)
                                pii_results.append({
                                    'mapping_file': input_file,
                                    'pii_file': None,
                                    'error': str(e)
                                })
                        
                        # Build summary message
                        success_count = sum(1 for r in pii_results if r.get('pii_file'))
                        failed_count = len(pii_results) - success_count
                        
                        # Individual file details
                        details = "Individual File Results:\n\n"
                        for r in pii_results:
                            if r.get('pii_file'):
                                details += f"  → PII Results: {r['pii_file']}\n"
                                details += f"  → Found: {r['pii_count']} PII / {r['total_count']} total fields\n\n"
                            else:
                                details += f" {r['mapping_file']}\n"
                                details += f"  → Error: {r.get('error', 'Unknown error')}\n\n"
                        
                        # Combined summary
                        summary = "\n\n" 
                        summary += "Combined Summary:\n"
                        summary += f"Total Files Processed:     {len(generated_files)}\n"
                        summary += f"Successful:                {success_count}\n"
                        summary += f"Failed:                    {failed_count}\n"
                        summary += f"\nTotal PII Fields Found:    {total_pii_count}\n"
                        summary += f"Total Fields Analyzed:     {total_field_count}\n"
                        
                        if total_field_count > 0:
                            detection_rate = (total_pii_count / total_field_count * 100)
                            summary += f"Overall Detection Rate:    {detection_rate:.1f}%\n"
                        
                        # Final message
                        final_message = details + summary
                        
                        # Update status
                        if failed_count == 0:
                            status_msg = (
                                f"PII Detection Complete! {success_count}/{len(generated_files)} files processed. "
                                f"Found {total_pii_count}/{total_field_count} PII fields"
                            )
                            self.show_status(status_msg, "success")
                            messagebox.showinfo("PII Detection Complete", final_message)
                        else:
                            status_msg = (
                                f"PII Detection Completed with errors. "
                                f"{success_count}/{len(generated_files)} successful. "
                                f"Found {total_pii_count}/{total_field_count} PII fields"
                            )
                            self.show_status(status_msg, "success")
                            messagebox.showwarning("PII Detection Completed", final_message)
                        
                        logger.info("Cleaning up: Deleting mapping files, keeping only PII results")
                        for mapping_file in generated_files:
                            try:
                                if os.path.exists(mapping_file):
                                    os.remove(mapping_file)
                                    logger.info(f"Deleted mapping file: {mapping_file}")
                            except Exception as e:
                                logger.warning(f"Could not delete {mapping_file}: {e}")

                        logger.info(f"PII detection complete: {success_count} successful, {failed_count} failed")
                        
                    except Exception as e:
                        error_msg = f"PII detection initialization failed: {str(e)}"
                        self.show_status(error_msg, "error")
                        warning_msg = (
                            f"Mapping spec files were generated successfully ({count} files):\n"
                            + "\n".join([f"  - {f}" for f in generated_files]) + "\n\n"
                            f"However, PII detection failed:\n{str(e)}"
                        )
                        messagebox.showwarning("PII Detection Error", warning_msg)
                        logger.error(f"PII detection error: {e}", exc_info=True)
                        # Note: Mapping files are kept, only PII detection failed


            finally:
                conn.close()

        except Exception as e:
            error_msg = f"Operation failed: {str(e)}"
            self.show_status(f"{error_msg}", "error")
            messagebox.showerror("Error", error_msg)
            logger.error(f"Run error: {e}", exc_info=True)
        finally:
            self.run_btn.config(state="normal")


WORKDAY_ORACLE_MAPPING_TABLE = (
    "map_maestro.workdayoraclemap"  # <-- CHANGE THIS TO YOUR ACTUAL TABLE NAME
)


# Update the fetch function with your actual table name
def fetch_workday_oracle_mapping(conn, mapping_type="bidirectional"):
    """
    Fetch Workday-Oracle bidirectional mapping from your existing table.
    Update WORKDAY_ORACLE_MAPPING_TABLE with your actual table name.
    """
    query = f"""
        SELECT 
            "Column Number" as column_number,
            "Workday Table Name" as workday_table_name,
            "Workday Field Name" as workday_field_name,
            "Workday Field Description" as workday_field_description,
            "Workday Data Type" as workday_data_type,
            "Required?" as required,
            "Example Target Data Value" as example_target_data_value,
            "Oracle Table Name" as oracle_table_name,
            "Oracle Field Name" as oracle_field_name,
            "Oracle Translation Rules" as oracle_translation_rules,
            "Cross References" as cross_references,
            "Comments" as comments,
            "Column Status" as column_status,
            "Path" as path,
            "Parameter" as parameter
        FROM {WORKDAY_ORACLE_MAPPING_TABLE}
        WHERE "Column Status" IS NULL OR "Column Status" != 'Inactive'
        ORDER BY "Column Number"
    """

    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query)
            rows = cur.fetchall()
            logger.info(
                f"Fetched {len(rows)} Workday-Oracle mapping rows from {WORKDAY_ORACLE_MAPPING_TABLE}"
            )
            return rows
    except psycopg2.Error as e:
        logger.error(f"Error fetching Workday-Oracle mapping: {e.pgerror or e}")
        # Provide helpful error message
        if "does not exist" in str(e):
            logger.error(
                f"Table {WORKDAY_ORACLE_MAPPING_TABLE} does not exist. Please check the table name."
            )
        return []


# SQL to check your existing table structure
def verify_table_structure(conn):
    """
    Helper function to verify your Workday-Oracle mapping table structure.
    Run this to confirm the table name and columns.
    """
    query = """
        SELECT 
            table_schema,
            table_name,
            column_name,
            data_type,
            ordinal_position
        FROM information_schema.columns
        WHERE table_name LIKE '%workday%' 
           OR table_name LIKE '%oracle%'
           OR table_name LIKE '%translation%'
           OR table_name LIKE '%mapping%'
        ORDER BY table_schema, table_name, ordinal_position;
    """

    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query)
            tables = cur.fetchall()

            print("\n=== Found tables with relevant names ===")
            current_table = None
            for row in tables:
                table_full = f"{row['table_schema']}.{row['table_name']}"
                if table_full != current_table:
                    current_table = table_full
                    print(f"\nTable: {table_full}")
                    print("Columns:")
                print(
                    f"  {row['ordinal_position']:2d}. {row['column_name']} ({row['data_type']})"
                )

            return tables
    except Exception as e:
        print(f"Error verifying table structure: {e}")
        return []


# Add a validation function to check if transitivity is possible
def validate_transitivity_setup(conn):
    """
    Validates if the transitivity feature can work with current data.
    Returns a report of what's available and what's missing.
    """
    report = {
        "workday_oracle_mapping": False,
        "legacy_to_workday_count": 0,
        "legacy_to_oracle_count": 0,
        "unique_legacy_systems": [],
        "ready_for_transitivity": [],
    }

    try:
        # Check Workday-Oracle mapping table
        wo_mappings = fetch_workday_oracle_mapping(conn)
        report["workday_oracle_mapping"] = len(wo_mappings) > 0
        report["workday_oracle_mapping_count"] = len(wo_mappings)

        # Check existing legacy mappings
        query = """
            SELECT DISTINCT 
                legacysystem, 
                targetsystem,
                COUNT(*) as mapping_count
            FROM map_maestro.parentprojectdetails
            GROUP BY legacysystem, targetsystem
            ORDER BY legacysystem, targetsystem
        """

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query)
            mappings = cur.fetchall()

            legacy_systems = set()
            for m in mappings:
                legacy_systems.add(m["legacysystem"])

                if m["targetsystem"] == "Workday":
                    report["legacy_to_workday_count"] += 1
                elif m["targetsystem"] == "Oracle":
                    report["legacy_to_oracle_count"] += 1

            report["unique_legacy_systems"] = list(legacy_systems)

            # Check which systems are ready for transitivity
            for legacy in legacy_systems:
                has_workday = any(
                    m["legacysystem"] == legacy and m["targetsystem"] == "Workday"
                    for m in mappings
                )
                has_oracle = any(
                    m["legacysystem"] == legacy and m["targetsystem"] == "Oracle"
                    for m in mappings
                )

                if has_workday and report["workday_oracle_mapping"]:
                    report["ready_for_transitivity"].append(
                        f"{legacy} → Oracle (via Workday)"
                    )
                if has_oracle and report["workday_oracle_mapping"]:
                    report["ready_for_transitivity"].append(
                        f"{legacy} → Workday (via Oracle)"
                    )

        # Print report
        print("\n=== Transitivity Readiness Report ===")
        print(
            f"Workday-Oracle Mapping Table: {'Found' if report['workday_oracle_mapping'] else 'Not Found'}"
        )
        if report["workday_oracle_mapping"]:
            print(f"  - Mappings available: {report['workday_oracle_mapping_count']}")
        print(f"\nLegacy Systems Found: {len(report['unique_legacy_systems'])}")
        for ls in report["unique_legacy_systems"]:
            print(f"  - {ls}")
        print(f"\nMappings Available:")
        print(f"  - Legacy → Workday: {report['legacy_to_workday_count']} systems")
        print(f"  - Legacy → Oracle: {report['legacy_to_oracle_count']} systems")
        print(f"\nReady for Transitivity:")
        if report["ready_for_transitivity"]:
            for r in report["ready_for_transitivity"]:
                print(f"  {r}")
        else:
            print("  No systems ready (need Workday-Oracle mapping table)")

        return report

    except Exception as e:
        print(f"Error validating transitivity setup: {e}")
        return report


# Test function to run standalone
if __name__ == "__main__":
    # Test the transitivity setup
    test_conn = get_connection()
    try:
        print("Verifying table structures...")
        verify_table_structure(test_conn)

        print("\n" + "=" * 50)
        print("Checking transitivity readiness...")
        validate_transitivity_setup(test_conn)

    finally:
        test_conn.close()


if __name__ == "__main__":
    # init_logging()
    root = tk.Tk()
    app = ElegantMapMaestroApp(root)
    root.mainloop()